import sys, os
root = r'D:\OneDrive\PROJETOFINANCEIRO.PY'
sys.path.insert(0, root)

import openpyxl
from financeiro.veiculo_helper import VeiculoHelper
from financeiro.cliente_helper import ClienteHelper

# Arquivo original
arquivo_original = os.path.join(root, 'financeiro', 'uploads', 'manifestos', 'Manifesto_Frete_09-25.xlsx')
print(f'🚛 INTEGRANDO AUTOMATICAMENTE: Manifesto_Frete_09-25.xlsx')

try:
    # Carregar workbook
    wb = openpyxl.load_workbook(arquivo_original, data_only=True)
    ws = wb.active
    
    print(f'📊 Arquivo carregado: {ws.max_row} linhas x {ws.max_column} colunas')
    
    # Mostrar TODAS as colunas para identificar
    print('\n📋 TODAS as colunas do arquivo:')
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(1, col).value
        print(f'  Col {col}: "{valor}"')
    
    # Adicionar cabeçalhos das novas colunas
    nova_col_status = ws.max_column + 1
    nova_col_tipologia = nova_col_status + 1
    nova_col_cliente = nova_col_status + 2
    
    ws.cell(1, nova_col_status, 'Status_Veiculo')
    ws.cell(1, nova_col_tipologia, 'Tipologia')
    ws.cell(1, nova_col_cliente, 'Cliente_Real')
    
    print(f'\n✅ Cabeçalhos adicionados nas colunas {nova_col_status}, {nova_col_tipologia}, {nova_col_cliente}')
    
    # Identificar colunas manualmente baseado no que encontramos
    col_placa = None
    col_cliente = None
    
    # Buscar coluna de placa/veículo
    for col in range(1, ws.max_column - 2):
        valor = str(ws.cell(1, col).value or '').upper()
        if any(palavra in valor for palavra in ['PLACA', 'VEICULO', 'VEHICLE', 'TRUCK']):
            col_placa = col
            print(f'🚚 Coluna PLACA encontrada: Col {col} = "{ws.cell(1, col).value}"')
            break
    
    # Buscar coluna de cliente
    for col in range(1, ws.max_column - 2):
        valor = str(ws.cell(1, col).value or '').upper()
        if any(palavra in valor for palavra in ['CLIENTE', 'CLIENT', 'DESTINATARIO', 'REMETENTE']):
            col_cliente = col
            print(f'👥 Coluna CLIENTE encontrada: Col {col} = "{ws.cell(1, col).value}"')
            break
    
    # Se não encontrou, usar algumas amostras de dados para tentar identificar
    if not col_placa or not col_cliente:
        print('\n🔍 Tentando identificar por padrão dos dados...')
        for row in range(2, min(6, ws.max_row + 1)):
            linha_dados = []
            for col in range(1, min(ws.max_column - 2, 15)):
                valor = ws.cell(row, col).value
                linha_dados.append(f'Col{col}:"{valor}"')
            print(f'  Linha {row}: {" | ".join(linha_dados)}')
    
    print(f'\n📍 RESULTADO - Placa: Col {col_placa}, Cliente: Col {col_cliente}')
    
    # Coletar amostra para testar
    placas_amostra = set()
    clientes_amostra = set()
    
    for row in range(2, min(20, ws.max_row + 1)):
        if col_placa:
            placa = ws.cell(row, col_placa).value
            if placa and str(placa).strip():
                placas_amostra.add(str(placa).upper().strip())
        
        if col_cliente:
            cliente = ws.cell(row, col_cliente).value
            if cliente and str(cliente).strip():
                clientes_amostra.add(str(cliente).upper().strip())
    
    print(f'\n🔍 Amostra: {len(placas_amostra)} placas, {len(clientes_amostra)} clientes')
    if placas_amostra:
        print(f'  Placas: {list(placas_amostra)[:3]}...')
    if clientes_amostra:
        print(f'  Clientes: {list(clientes_amostra)[:3]}...')
    
    # Buscar na base de dados
    dados_veiculos = {}
    dados_clientes = {}
    
    if placas_amostra:
        dados_veiculos = VeiculoHelper.buscar_multiplas_placas(list(placas_amostra))
        veiculos_encontrados = sum(1 for v in dados_veiculos.values() if v.get('encontrado', False))
        print(f'\n🚚 Veículos encontrados: {veiculos_encontrados}/{len(placas_amostra)}')
    
    if clientes_amostra:
        dados_clientes = ClienteHelper.buscar_multiplos_nomes_ajustados(list(clientes_amostra))
        clientes_encontrados = sum(1 for c in dados_clientes.values() if c.get('encontrado', False))
        print(f'👥 Clientes encontrados: {clientes_encontrados}/{len(clientes_amostra)}')
    
    # Se temos dados, vamos integrar TODAS as linhas
    print(f'\n🔄 Integrando TODAS as {ws.max_row - 1} linhas...')
    
    # Primeiro, coletar TODAS as placas e clientes únicos
    todas_placas = set()
    todos_clientes = set()
    
    for row in range(2, ws.max_row + 1):
        if col_placa:
            placa = ws.cell(row, col_placa).value
            if placa and str(placa).strip():
                todas_placas.add(str(placa).upper().strip())
        
        if col_cliente:
            cliente = ws.cell(row, col_cliente).value
            if cliente and str(cliente).strip():
                todos_clientes.add(str(cliente).upper().strip())
    
    print(f'📊 Total únicos: {len(todas_placas)} placas, {len(todos_clientes)} clientes')
    
    # Buscar TODOS os dados
    if todas_placas:
        print('🚚 Buscando TODOS os veículos...')
        dados_veiculos = VeiculoHelper.buscar_multiplas_placas(list(todas_placas))
    
    if todos_clientes:
        print('👥 Buscando TODOS os clientes...')
        dados_clientes = ClienteHelper.buscar_multiplos_nomes_ajustados(list(todos_clientes))
    
    # Integrar linha por linha
    linhas_processadas = 0
    for row in range(2, ws.max_row + 1):
        # Status e tipologia do veículo
        if col_placa:
            placa = ws.cell(row, col_placa).value
            if placa and str(placa).strip():
                placa_norm = str(placa).upper().strip()
                veiculo = dados_veiculos.get(placa_norm, {})
                ws.cell(row, nova_col_status, veiculo.get('status') or '0')
                ws.cell(row, nova_col_tipologia, veiculo.get('tipologia') or '0')
            else:
                ws.cell(row, nova_col_status, '0')
                ws.cell(row, nova_col_tipologia, '0')
        else:
            ws.cell(row, nova_col_status, '0')
            ws.cell(row, nova_col_tipologia, '0')
        
        # Cliente real
        if col_cliente:
            cliente = ws.cell(row, col_cliente).value
            if cliente and str(cliente).strip():
                cliente_norm = str(cliente).upper().strip()
                cliente_dados = dados_clientes.get(cliente_norm, {})
                ws.cell(row, nova_col_cliente, cliente_dados.get('nome_real') or '0')
            else:
                ws.cell(row, nova_col_cliente, '0')
        else:
            ws.cell(row, nova_col_cliente, '0')
        
        linhas_processadas += 1
        
        if linhas_processadas % 200 == 0:
            print(f'  Processadas {linhas_processadas} linhas...')
    
    # Salvar arquivo
    print('💾 Salvando arquivo...')
    wb.save(arquivo_original)
    
    print(f'\n🎉 INTEGRAÇÃO COMPLETA!')
    print(f'📁 Arquivo: {arquivo_original}')
    print(f'📊 {linhas_processadas} linhas integradas')
    print(f'➕ 3 colunas adicionadas: Status_Veiculo (Col {nova_col_status}), Tipologia (Col {nova_col_tipologia}), Cliente_Real (Col {nova_col_cliente})')
    print(f'🚚 {len(dados_veiculos)} placas processadas')
    print(f'👥 {len(dados_clientes)} clientes processados')
    
except Exception as e:
    print(f'❌ ERRO: {e}')
    import traceback
    traceback.print_exc()