from flask import Blueprint, render_template, request, flash, redirect, url_forfrom flask import Blueprint, render_template, request, redirect, url_for, flash, send_from_directoryfrom flask import Blueprint, render_template, request, redirect, url_for, flash, send_from_directory""""""

import os

import os

bp = Blueprint('frete', __name__, url_prefix='/frete')

import shutilimport os

@bp.route('/')

def index():from datetime import datetime

    return render_template('frete.html')

import shutilMódulo principal de FRETE - apenas rotas e coordenaçãoMódulo principal de FRETE - apenas rotas e coordenação

@bp.route('/upload', methods=['POST'])

def upload():# Importar lógicas específicas

    if 'arquivo' not in request.files:

        flash('Nenhum arquivo selecionado')from . import manifestofrom datetime import datetime

        return redirect(url_for('frete.index'))

    

    arquivo = request.files['arquivo']

    if arquivo.filename == '':bp = Blueprint('frete', __name__, url_prefix='/frete')from pathlib import PathLógicas específicas estão em manifesto.py e valencio.pyLógicas específicas estão em manifesto.py e valencio.py

        flash('Nenhum arquivo selecionado')

        return redirect(url_for('frete.index'))

    

    flash(f'Arquivo {arquivo.filename} recebido com sucesso!')

    return redirect(url_for('frete.index'))
@bp.route('/')

def index():# Importar lógicas específicas""""""

    return render_template('frete.html')

from . import manifesto



@bp.route('/importacao', methods=['GET', 'POST'])# from . import valencio  # Será criado depois

def importacao():

    if request.method == 'POST':

        files = request.files.getlist('files')

        if not files or files[0].filename == '':bp = Blueprint('frete', __name__, url_prefix='/frete')from flask import Blueprint, render_template, request, redirect, url_for, flash, send_from_directoryfrom flask import Blueprint, render_template, request, redirect, url_for, flash, send_from_directory

            flash('Nenhum arquivo selecionado.', 'error')

            return redirect(url_for('frete.importacao'))



        # Pegar o tipo escolhido pelo usuárioimport osimport os

        tipo_arquivo = request.form.get('tipo_arquivo', 'valencio')

        @bp.route('/')

        # Criar diretório de uploads

        uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))def index():import shutilimport shutil

        os.makedirs(uploads_dir, exist_ok=True)

    return render_template('frete.html')

        for file in files:

            if file.filename:from datetime import datetimefrom datetime import datetime

                temp_filepath = os.path.join(uploads_dir, file.filename)

                file.save(temp_filepath)

                

                try:@bp.route('/importacao', methods=['GET', 'POST'])from pathlib import Pathfrom pathlib import Path

                    if tipo_arquivo == 'manifesto':

                        # Processar MANIFESTOdef importacao():

                        result = manifesto.process_manifesto_file(temp_filepath, uploads_dir)

                            if request.method == 'POST':

                        if 'error' in result:

                            flash(f'{file.filename}: Erro - {result["error"]}', 'error')        if 'files' not in request.files:

                        else:

                            flash(f'{file.filename}: {result["message"]}', 'success')            flash('Nenhum arquivo enviado!', 'error')# Importar lógicas específicas# Importar lógicas específicas

                        

                        # Renomear arquivo original            return redirect(url_for('frete.importacao'))

                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

                        orig_name = f"MANIFESTO_{timestamp}_{file.filename}"from . import manifestofrom . import manifesto

                        final_path = os.path.join(uploads_dir, orig_name)

                        shutil.move(temp_filepath, final_path)        files = request.files.getlist('files')

                    

                    elif tipo_arquivo == 'valencio':        if not files or files[0].filename == '':# from . import valencio  # Será criado depois# from . import valencio  # Será criado depois

                        # Salvar arquivo VALENCIO (lógica será implementada depois)

                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')            flash('Nenhum arquivo selecionado.', 'error')

                        save_name = f"VALENCIO_{timestamp}_{file.filename}"

                        final_path = os.path.join(uploads_dir, save_name)            return redirect(url_for('frete.importacao'))

                        shutil.move(temp_filepath, final_path)

                        flash(f'{file.filename}: Arquivo Valencio salvo (processamento em desenvolvimento)', 'warning')

                    

                except Exception as e:        # Pegar o tipo de arquivo selecionado pelo usuáriobp = Blueprint('frete', __name__, url_prefix='/frete')bp = Blueprint('frete', __name__, url_prefix='/frete')

                    flash(f'{file.filename}: Erro - {str(e)}', 'error')

                    if os.path.exists(temp_filepath):        tipo_arquivo = request.form.get('tipo_arquivo', 'valencio')  # default: valencio

                        os.remove(temp_filepath)

        

        return redirect(url_for('frete.importacao'))

        # Criar diretório de uploads do frete se não existir

    # GET: Listar arquivos

    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))        uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))

    

    uploads = []        os.makedirs(uploads_dir, exist_ok=True)

    if os.path.isdir(uploads_dir):

        for fname in sorted(os.listdir(uploads_dir), reverse=True):@bp.route('/')@bp.route('/')

            fpath = os.path.join(uploads_dir, fname)

            try:        for file in files:

                mtime = os.path.getmtime(fpath)

                size = os.path.getsize(fpath)            if file.filename:def index():def index():

                uploads.append({

                    'name': fname,                # Salvar arquivo temporário para análise

                    'display_name': fname,

                    'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),                temp_filepath = os.path.join(uploads_dir, file.filename)    return render_template('frete.html')    return render_template('frete.html')

                    'size': size

                })                file.save(temp_filepath)

            except:

                continue                

    

    # Arquivos acumulados                try:

    accum_dir = os.path.join(uploads_dir, 'acumulados')

    accumulated = []                    if tipo_arquivo == 'manifesto':

    if os.path.isdir(accum_dir):

        for fname in sorted(os.listdir(accum_dir)):                        # LÓGICA MANIFESTO: Delegar para módulo específico

            fpath = os.path.join(accum_dir, fname)

            try:                        result = manifesto.process_manifesto_file(temp_filepath, uploads_dir)@bp.route('/importacao', methods=['GET', 'POST'])@bp.route('/importacao', methods=['GET', 'POST'])

                mtime = os.path.getmtime(fpath)

                size = os.path.getsize(fpath)                        

                accumulated.append({

                    'name': fname,                        if 'error' in result:def importacao():def importacao():

                    'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),

                    'size': size                            flash(f'{file.filename}: Erro - {result["error"]}', 'error')

                })

            except:                        else:    if request.method == 'POST':    if request.method == 'POST':

                continue

                            flash(f'{file.filename}: Manifesto processado - {result["message"]}', 'success')

    return render_template('frete_importacao.html', uploads=uploads, accumulated=accumulated)

                                if 'files' not in request.files:        if 'files' not in request.files:



@bp.route('/uploads/<path:filename>')                        # Manter arquivo original com timestamp

def uploaded_file(filename):

    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')            flash('Nenhum arquivo enviado!', 'error')            flash('Nenhum arquivo enviado!', 'error')

    return send_from_directory(uploads_dir, filename)

                        orig_name = f"MANIFESTO_{timestamp}_{file.filename}"



@bp.route('/acumulados/<path:filename>')                        final_path = os.path.join(uploads_dir, orig_name)            return redirect(url_for('frete.importacao'))            return redirect(url_for('frete.importacao'))

def accumulated_file(filename):

    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))                        if temp_filepath != final_path:

    accum_dir = os.path.join(uploads_dir, 'acumulados')

    return send_from_directory(accum_dir, filename)                            shutil.move(temp_filepath, final_path)

                    

                    elif tipo_arquivo == 'valencio':        files = request.files.getlist('files')        files = request.files.getlist('files')

                        # LÓGICA VALENCIO: Será implementada em valencio.py

                        # Por enquanto, apenas salvar o arquivo        if not files or files[0].filename == '':        if not files or files[0].filename == '':

                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

                        save_name = f"VALENCIO_{timestamp}_{file.filename}"            flash('Nenhum arquivo selecionado.', 'error')            flash('Nenhum arquivo selecionado.', 'error')

                        

                        final_path = os.path.join(uploads_dir, save_name)            return redirect(url_for('frete.importacao'))            return redirect(url_for('frete.importacao'))

                        if temp_filepath != final_path:

                            shutil.move(temp_filepath, final_path)

                        

                        flash(f'{file.filename}: Arquivo Valencio salvo (processamento será implementado)', 'warning')        # Pegar o tipo de arquivo selecionado pelo usuário        # Pegar o tipo de arquivo selecionado pelo usuário

                    

                    else:        tipo_arquivo = request.form.get('tipo_arquivo', 'valencio')  # default: valencio        tipo_arquivo = request.form.get('tipo_arquivo', 'valencio')  # default: valencio

                        flash(f'{file.filename}: Tipo de arquivo inválido', 'error')

                                    

                except Exception as e:

                    flash(f'{file.filename}: Erro na análise do arquivo ({str(e)})', 'error')        # Criar diretório de uploads do frete se não existir        # Criar diretório de uploads do frete se não existir

                    # Remover arquivo temporário em caso de erro

                    if os.path.exists(temp_filepath):        uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))        uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))

                        os.remove(temp_filepath)

        os.makedirs(uploads_dir, exist_ok=True)        os.makedirs(uploads_dir, exist_ok=True)

        return redirect(url_for('frete.importacao'))



    # GET: Listar arquivos

    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))        for file in files:        for file in files:

    

    # Lista de arquivos enviados (pasta uploads/frete)            if file.filename:            if file.filename:

    uploads = []

    if os.path.isdir(uploads_dir):                # Salvar arquivo temporário para análise                # Salvar arquivo temporário para análise

        for fname in sorted(os.listdir(uploads_dir), reverse=True):

            fpath = os.path.join(uploads_dir, fname)                temp_filepath = os.path.join(uploads_dir, file.filename)                temp_filepath = os.path.join(uploads_dir, file.filename)

            try:

                mtime = os.path.getmtime(fpath)                file.save(temp_filepath)                file.save(temp_filepath)

                size = os.path.getsize(fpath)

                                                

                uploads.append({

                    'name': fname,                try:                try:

                    'display_name': fname,

                    'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),                    if tipo_arquivo == 'manifesto':                    if tipo_arquivo == 'manifesto':

                    'size': size

                })                        # LÓGICA MANIFESTO: Delegar para módulo específico                        # LÓGICA MANIFESTO: Delegar para módulo específico

            except Exception:

                continue                        result = manifesto.process_manifesto_file(temp_filepath, uploads_dir)                        result = manifesto.process_manifesto_file(temp_filepath, uploads_dir)

    

    # Lista de arquivos acumulados                                                

    accum_dir = os.path.join(uploads_dir, 'acumulados')

    accumulated = []                        if 'error' in result:                        if 'error' in result:

    if os.path.isdir(accum_dir):

        for fname in sorted(os.listdir(accum_dir)):                            flash(f'{file.filename}: Erro - {result["error"]}', 'error')                            flash(f'{file.filename}: Erro - {result["error"]}', 'error')

            fpath = os.path.join(accum_dir, fname)

            try:                        else:                        else:

                mtime = os.path.getmtime(fpath)

                size = os.path.getsize(fpath)                            flash(f'{file.filename}: Manifesto processado - {result["message"]}', 'success')                            flash(f'{file.filename}: Manifesto processado - {result["message"]}', 'success')

                

                accumulated.append({                                                

                    'name': fname,

                    'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),                        # Manter arquivo original com timestamp                        # Manter arquivo original com timestamp

                    'size': size

                })                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            except Exception:

                continue                        orig_name = f"MANIFESTO_{timestamp}_{file.filename}"                        orig_name = f"MANIFESTO_{timestamp}_{file.filename}"



    return render_template('frete_importacao.html', uploads=uploads, accumulated=accumulated)                        final_path = os.path.join(uploads_dir, orig_name)                        final_path = os.path.join(uploads_dir, orig_name)



                        if temp_filepath != final_path:                        if temp_filepath != final_path:

@bp.route('/uploads/<path:filename>')

def uploaded_file(filename):                            shutil.move(temp_filepath, final_path)                            shutil.move(temp_filepath, final_path)

    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))

    return send_from_directory(uploads_dir, filename)                                        



                    elif tipo_arquivo == 'valencio':                    elif tipo_arquivo == 'valencio':

@bp.route('/acumulados/<path:filename>')

def accumulated_file(filename):                        # LÓGICA VALENCIO: Será implementada em valencio.py                        # LÓGICA VALENCIO: Será implementada em valencio.py

    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))

    accum_dir = os.path.join(uploads_dir, 'acumulados')                        # Por enquanto, apenas salvar o arquivo                        # Por enquanto, apenas salvar o arquivo

    return send_from_directory(accum_dir, filename)
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

                        save_name = f"VALENCIO_{timestamp}_{file.filename}"                        save_name = f"VALENCIO_{timestamp}_{file.filename}"

                                                

                        final_path = os.path.join(uploads_dir, save_name)                        final_path = os.path.join(uploads_dir, save_name)

                        if temp_filepath != final_path:                        if temp_filepath != final_path:

                            shutil.move(temp_filepath, final_path)                            shutil.move(temp_filepath, final_path)

                                                

                        flash(f'{file.filename}: Arquivo Valencio salvo (processamento será implementado)', 'warning')                        flash(f'{file.filename}: Arquivo Valencio salvo (processamento será implementado)', 'warning')

                                        

                    else:                    else:

                        flash(f'{file.filename}: Tipo de arquivo inválido', 'error')                        flash(f'{file.filename}: Tipo de arquivo inválido', 'error')

                                        

                except Exception as e:                except Exception as e:

                    flash(f'{file.filename}: Erro na análise do arquivo ({str(e)})', 'error')                    flash(f'{file.filename}: Erro na análise do arquivo ({str(e)})', 'error')

                    # Remover arquivo temporário em caso de erro                    # Remover arquivo temporário em caso de erro

                    if os.path.exists(temp_filepath):                    if os.path.exists(temp_filepath):

                        os.remove(temp_filepath)                        os.remove(temp_filepath)



        return redirect(url_for('frete.importacao'))        return redirect(url_for('frete.importacao'))



    # GET: Listar arquivos    # GET: Listar arquivos

    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))

        

    # Lista de arquivos enviados (pasta uploads/frete)    # Lista de arquivos enviados (pasta uploads/frete)

    uploads = []    uploads = []

    if os.path.isdir(uploads_dir):    if os.path.isdir(uploads_dir):

        for fname in sorted(os.listdir(uploads_dir), reverse=True):        for fname in sorted(os.listdir(uploads_dir), reverse=True):

            fpath = os.path.join(uploads_dir, fname)            fpath = os.path.join(uploads_dir, fname)

            try:            try:

                mtime = os.path.getmtime(fpath)                mtime = os.path.getmtime(fpath)

                size = os.path.getsize(fpath)                size = os.path.getsize(fpath)

                                

                uploads.append({                uploads.append({

                    'name': fname,                    'name': fname,

                    'display_name': fname,                    'display_name': fname,

                    'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),                    'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),

                    'size': size                    'size': size

                })                })

            except Exception:            except Exception:

                continue                continue

        

    # Lista de arquivos acumulados    # Lista de arquivos acumulados

    accum_dir = os.path.join(uploads_dir, 'acumulados')    accum_dir = os.path.join(uploads_dir, 'acumulados')

    accumulated = []    accumulated = []

    if os.path.isdir(accum_dir):    if os.path.isdir(accum_dir):

        for fname in sorted(os.listdir(accum_dir)):        for fname in sorted(os.listdir(accum_dir)):

            fpath = os.path.join(accum_dir, fname)            fpath = os.path.join(accum_dir, fname)

            try:            try:

                mtime = os.path.getmtime(fpath)                mtime = os.path.getmtime(fpath)

                size = os.path.getsize(fpath)                size = os.path.getsize(fpath)

                                

                accumulated.append({                accumulated.append({

                    'name': fname,                    'name': fname,

                    'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),                    'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),

                    'size': size                    'size': size

                })                })

            except Exception:            except Exception:

                continue                continue



    return render_template('frete_importacao.html', uploads=uploads, accumulated=accumulated)    return render_template('frete_importacao.html', uploads=uploads, accumulated=accumulated)





@bp.route('/uploads/<path:filename>')@bp.route('/uploads/<path:filename>')

def uploaded_file(filename):def uploaded_file(filename):

    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))

    return send_from_directory(uploads_dir, filename)    return send_from_directory(uploads_dir, filename)





@bp.route('/acumulados/<path:filename>')@bp.route('/acumulados/<path:filename>')

def accumulated_file(filename):def accumulated_file(filename):

    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))

    accum_dir = os.path.join(uploads_dir, 'acumulados')    accum_dir = os.path.join(uploads_dir, 'acumulados')

    return send_from_directory(accum_dir, filename)    return send_from_directory(accum_dir, filename)
    """Copia arquivo Manifesto para o acumulado (CTRL+C + CTRL+V REAL)."""
    try:
        # Diretório de acumulados
        uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))
        accum_dir = os.path.join(uploads_dir, 'acumulados')
        os.makedirs(accum_dir, exist_ok=True)
        
        # Arquivo acumulado principal
        accumulated_path = os.path.join(accum_dir, 'manifestos_acumulados.csv')
        
        # CTRL+C: Ler Excel preservando formato EXATO (sem quebras de linha)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Ler linha por linha, tratando quebras de linha internas
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            # Só adicionar se a linha tem pelo menos um valor não vazio
            if any(cell is not None and str(cell).strip() != '' for cell in row):
                # Converter tudo para string, REMOVENDO quebras de linha internas
                clean_row = []
                for cell in row:
                    if cell is None:
                        clean_row.append('')
                    elif hasattr(cell, 'strftime'):  # É data
                        clean_row.append(cell.strftime('%d/%m/%Y'))
                    else:
                        # CRITICAL: Remover quebras de linha que estragam o CSV
                        cell_str = str(cell).replace('\n', ' ').replace('\r', ' ').strip()
                        clean_row.append(cell_str)
                all_rows.append(clean_row)
        
        wb.close()
        
        # CTRL+V: Salvar EXATAMENTE como está (sem quebras de linha indevidas)
        with open(accumulated_path, 'w', encoding='utf-8-sig', newline='') as f:
            import csv
            writer = csv.writer(f, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            for row in all_rows:
                writer.writerow(row)
        
        return {
            'action': 'copied',
            'file': accumulated_path,
            'rows_added': len(all_rows),
            'message': f'CTRL+C + CTRL+V realizado: {len(all_rows)} linhas copiadas fielmente'
        }
        
    except Exception as e:
        return {'error': str(e)}


def read_excel_preserving_format(filepath: str) -> dict:
    """Lê Excel preservando formato original EXATAMENTE (CTRL+C)."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Ler todas as linhas preservando formato
        data = {'header': [], 'rows': []}
        
        # Header (primeira linha)
        header = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v is not None:
                header.append(str(v).strip())
            else:
                header.append('')
        data['header'] = header
        
        # Dados (linhas 2 em diante)
        for r in range(2, ws.max_row + 1):
            row = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    if isinstance(v, datetime):
                        row.append(v.strftime('%d/%m/%Y'))
                    else:
                        row.append(str(v).strip())
                else:
                    row.append('')
            
            # Só adicionar se a linha tem pelo menos um dado
            if any(cell.strip() for cell in row):
                data['rows'].append(row)
        
        return data
        
    except Exception as e:
        print(f"Erro ao ler Excel: {e}")
        return None


def save_data_to_csv(data: dict, filepath: str):
    """Salva dados para CSV preservando formato (CTRL+V)."""
    with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
        # Header
        f.write(';'.join(data['header']) + '\n')
        
        # Dados
        for row in data['rows']:
            f.write(';'.join(row) + '\n')


def merge_into_accumulated(new_data: dict, accumulated_path: str) -> dict:
    """Merge inteligente: substitui período existente ou adiciona novo."""
    try:
        # Ler dados existentes
        existing_data = read_csv_data(accumulated_path)
        if not existing_data:
            return {'error': 'Erro ao ler arquivo acumulado existente'}
        
        # Detectar períodos no novo arquivo
        new_periods = extract_periods_from_data(new_data)
        existing_periods = extract_periods_from_data(existing_data)
        
        # Criar novo conjunto de dados
        merged_rows = []
        replaced_periods = []
        
        # Manter linhas existentes que não conflitam
        for row in existing_data['rows']:
            row_period = extract_period_from_row(row, existing_data['header'])
            if row_period not in new_periods:
                merged_rows.append(row)
            else:
                if row_period not in replaced_periods:
                    replaced_periods.append(row_period)
        
        # Adicionar todas as linhas novas
        merged_rows.extend(new_data['rows'])
        
        # Salvar resultado
        merged_data = {
            'header': new_data['header'],  # Usar header do arquivo novo
            'rows': merged_rows
        }
        save_data_to_csv(merged_data, accumulated_path)
        
        return {
            'action': 'merged',
            'file': accumulated_path,
            'rows_added': len(new_data['rows']),
            'periods_replaced': replaced_periods
        }
        
    except Exception as e:
        return {'error': str(e)}


def read_csv_data(filepath: str) -> dict:
    """Lê dados de CSV preservando formato."""
    try:
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            lines = f.readlines()
        
        if not lines:
            return None
            
        header = lines[0].strip().split(';')
        rows = []
        
        for line in lines[1:]:
            if line.strip():
                rows.append(line.strip().split(';'))
        
        return {'header': header, 'rows': rows}
        
    except Exception as e:
        print(f"Erro ao ler CSV: {e}")
        return None


def extract_periods_from_data(data: dict) -> set:
    """Extrai períodos únicos dos dados (ano-mês)."""
    periods = set()
    header = data['header']
    
    # Encontrar coluna de data
    data_col_idx = None
    for i, col in enumerate(header):
        if 'data' in col.lower():
            data_col_idx = i
            break
    
    if data_col_idx is None:
        return periods
    
    for row in data['rows']:
        if len(row) > data_col_idx:
            period = extract_period_from_row(row, header)
            if period:
                periods.add(period)
    
    return periods


def extract_period_from_row(row: list, header: list) -> str:
    """Extrai período (YYYY-MM) de uma linha baseado na coluna Data."""
    try:
        # Encontrar coluna de data
        data_col_idx = None
        for i, col in enumerate(header):
            if 'data' in col.lower():
                data_col_idx = i
                break
        
        if data_col_idx is None or len(row) <= data_col_idx:
            return None
        
        date_str = row[data_col_idx].strip()
        if not date_str:
            return None
        
        # Tentar parsear data (formato dd/mm/yyyy)
        if '/' in date_str:
            parts = date_str.split('/')
            if len(parts) >= 3:
                day, month, year = parts[0], parts[1], parts[2]
                return f"{year}-{month.zfill(2)}"
        
        return None
        
    except Exception:
        return None

def process_frete_file(filepath: str) -> str:
    """Processa o arquivo de frete conforme regras do cliente e salva um arquivo processado.
    Retorna o caminho do arquivo processado."""
    p = Path(filepath)
    suffix = p.suffix.lower()

    # If Excel (.xlsx/.xls) use openpyxl: process workbook in-place (add RESULTADO_N column)
    if suffix in ('.xlsx', '.xls'):
        wb = openpyxl.load_workbook(str(p))
        ws = wb.active

        # find header row (where a cell equals 'Tipo' or contains 'Tipo')
        header_row = None
        for r in range(1, ws.max_row + 1):
            first = ws.cell(row=r, column=1).value
            if first is not None and str(first).strip().lower() == 'tipo':
                header_row = r
                break
        if header_row is None:
            # try to find any row that contains 'Tipo' in any cell
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is not None and 'tipo' in str(v).lower():
                        header_row = r
                        break
                if header_row is not None:
                    break

        if header_row is None:
            raise ValueError('Não foi possível localizar a linha de cabeçalho (Tipo) no arquivo Excel')

        # read header names
        hdr = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=c).value
            hdr.append(str(v).strip() if v is not None else '')

        # find indices for columns (case-insensitive, tolerant)
        def find_header(name):
            name_low = name.lower()
            for i, h in enumerate(hdr, start=1):
                if h.lower() == name_low:
                    return i
            # fallback: find header containing name
            for i, h in enumerate(hdr, start=1):
                if name_low in h.lower():
                    return i
            return None

        # Determine indices: prefer 'Kg Real' as source for calculations
        valor_idx = find_header('Kg Real') or find_header('KgReal') or find_header('Kg') or find_header('Valor') or find_header('Valor NF') or len(hdr)
        remet_idx = find_header('Remetente') or 4
        tipo_idx = find_header('Tipo') or 1

        # Ensure RESULTADO_N is in column N (Excel column letter 'N' => index 14)
        from openpyxl.utils import column_index_from_string
        desired_col_idx = column_index_from_string('N')
        # If worksheet has fewer columns than desired, extend by adding empty cols
        if ws.max_column < desired_col_idx:
            # nothing to shift, just ensure cells exist
            pass
        else:
            # if column N already has any non-empty cells below header, insert a column at N to avoid overwriting
            has_data = False
            for rr in range(header_row, ws.max_row + 1):
                if ws.cell(row=rr, column=desired_col_idx).value not in (None, ''):
                    has_data = True
                    break
            if has_data:
                ws.insert_cols(desired_col_idx)

        resultado_col = desired_col_idx
        ws.cell(row=header_row, column=resultado_col, value='RESULTADO_N')

        # iterate rows, collect blocks between Manifesto: and Total - Manifesto:
        r = header_row + 1
        while r <= ws.max_row:
            cellA = ws.cell(row=r, column=1).value
            if cellA is not None and str(cellA).strip().startswith('Manifesto:'):
                # start a block
                block_rows = []
                r += 1
                # skip possible header row inside block (Tipo ...)
                # collect until Total - Manifesto:
                while r <= ws.max_row:
                    a = ws.cell(row=r, column=1).value
                    if a is not None and str(a).strip().startswith('Total - Manifesto:'):
                        # process block_rows
                        has_valencio = False
                        # first pass: detect presence of Valencio
                        for rr in block_rows:
                            remet = ws.cell(row=rr, column=remet_idx).value
                            if remet is not None and str(remet).strip().upper() == 'FRIGORIFICO VALENCIO LTDA':
                                has_valencio = True
                                break

                        block_sum = 0.0
                        # second pass: compute results and write
                        for rr in block_rows:
                            raw_val = ws.cell(row=rr, column=valor_idx).value
                            # parse numeric value
                            val = 0.0
                            if raw_val is None or str(raw_val).strip() == '':
                                val = 0.0
                            else:
                                try:
                                    val = float(raw_val)
                                except:
                                    s = str(raw_val).replace('.', '').replace(',', '.')
                                    try:
                                        val = float(s)
                                    except:
                                        val = 0.0

                            remet = ws.cell(row=rr, column=remet_idx).value or ''
                            if has_valencio:
                                if str(remet).strip().upper() == 'FRIGORIFICO VALENCIO LTDA':
                                    result_val = round(val * 0.67 / 0.88, 2)
                                else:
                                    result_val = round(val, 2)
                                ws.cell(row=rr, column=resultado_col, value=float(f"{result_val:.2f}"))
                                block_sum += float(result_val)
                            else:
                                # leave blank
                                ws.cell(row=rr, column=resultado_col, value=None)

                        # write block total on the Total - Manifesto row
                        if has_valencio:
                            ws.cell(row=r, column=resultado_col, value=float(f"{round(block_sum,2):.2f}"))
                        else:
                            ws.cell(row=r, column=resultado_col, value=None)

                        # move past total line
                        r += 1
                        break

                    else:
                        # if this is a data row (has Tipo value 'Frete' or non-empty in Tipo column), collect
                        tipo_val = ws.cell(row=r, column=tipo_idx).value
                        # heuristic: collect rows where Tipo column is not empty
                        if tipo_val is not None and str(tipo_val).strip() != '':
                            block_rows.append(r)
                        r += 1
                continue
            else:
                r += 1

        # save processed workbook
        new_path = p.parent / (p.stem + '_processed.xlsx')
        wb.save(str(new_path))
        return str(new_path)
    else:
        text = p.read_text(encoding='utf-8')
        lines = [l.rstrip('\n') for l in text.splitlines()]

    blocks = []
    current_manifest = None
    current_rows = []
    header = None

    for line in lines:
        if not line.strip():
            continue
        if line.startswith('Manifesto:'):
            if current_manifest is not None:
                blocks.append({'manifesto': current_manifest, 'rows': current_rows, 'header': header})
            current_manifest = line
            current_rows = []
            header = None
            continue
        if line.startswith('Tipo;'):
            header = [c.strip() for c in line.split(';')]
            continue
        if line.startswith('Total - Manifesto:'):
            current_rows.append({'__total_line__': line})
            blocks.append({'manifesto': current_manifest, 'rows': current_rows, 'header': header})
            current_manifest = None
            current_rows = []
            header = None
            continue
        if header is None:
            continue
        fields = [c.strip() for c in line.split(';')]
        while len(fields) < len(header):
            fields.append('')
        row = dict(zip(header, fields))
        current_rows.append(row)

    if current_manifest is not None and current_rows:
        blocks.append({'manifesto': current_manifest, 'rows': current_rows, 'header': header})

    # We'll preserve the original file exactly, and only add/update the RESULTADO_N column.
    # Parse original header line (first line that starts with 'manifesto' or 'Tipo;') to keep columns.
    # We'll rebuild each original data line, appending RESULTADO_N as the last column.

    # Determine original header from the first header-like line in the file
    orig_header_line = None
    for l in lines:
        if l.lower().startswith('manifesto;') or l.startswith('Tipo;'):
            orig_header_line = l
            break

    if orig_header_line is None:
        # fallback: use the constructed header
        orig_header = [h.strip() for h in 'manifesto;Tipo;Número;Data;Remetente;Origem;Destinatário;Destino;Volumes;Kg Real;Kg Taxado;Valor NF;Valor'.split(';')]
    else:
        orig_header = [c.strip() for c in orig_header_line.split(';')]

    # we'll output all original lines, but when encountering data rows we append RESULTADO_N
    out_lines = []
    # keep track to know which block we're in
    block_idx = 0
    for block in blocks:
        hdr = block['header']
        if not hdr:
            continue

        # identify input column for calculation in CSV/text blocks: prefer 'Kg Real' then 'Valor'
        if any(h.lower() == 'kg real' for h in hdr):
            valor_key = next(h for h in hdr if h.lower() == 'kg real')
        else:
            candidates = [h for h in hdr if 'valor' in h.lower()]
            if any(h.lower() == 'valor' for h in hdr):
                valor_key = next(h for h in hdr if h.lower() == 'valor')
            elif len(candidates) > 0:
                valor_key = candidates[-1]
            else:
                valor_key = hdr[-1]

        # check if block contains FRIGORIFICO VALENCIO LTDA (normalized)
        has_valencio = any((r.get('Remetente','').upper().strip() == 'FRIGORIFICO VALENCIO LTDA') for r in block['rows'] if isinstance(r, dict))

        block_sum = 0.0
        for r in block['rows']:
            if isinstance(r, dict):
                remet = r.get('Remetente','').strip()
                raw_val = r.get(valor_key,'')
                val_str = raw_val.replace('.', '').replace(',', '.') if raw_val else ''
                try:
                    val = float(val_str) if val_str != '' else 0.0
                except:
                    val = 0.0

                result_N = ''
                if has_valencio:
                    if remet.upper() == 'FRIGORIFICO VALENCIO LTDA':
                        result_val = round(val * 0.67 / 0.88, 2)
                        result_N = f"{result_val:.2f}"
                    else:
                        result_val = round(val, 2)
                        result_N = f"{result_val:.2f}"
                    try:
                        block_sum += float(str(result_val))
                    except:
                        pass

                # Reconstruct the original line values in the original header order, preserving original field text
                # Use values from r when available, otherwise empty string
                ordered_vals = []
                for col in orig_header:
                    # Map column name variants to keys in r: try direct, then normalized versions
                    val_text = ''
                    if col in r:
                        val_text = r.get(col,'')
                    else:
                        # try lowercase/no-accent matches
                        for k in r.keys():
                            if k.lower().replace('ã','a').replace('á','a').replace('à','a').replace('é','e').replace('ê','e').replace('í','i').replace('ó','o').replace('õ','o').replace('ô','o').replace('ú','u').replace('ç','c') == col.lower().replace('ã','a').replace('á','a').replace('à','a').replace('é','e').replace('ê','e').replace('í','i').replace('ó','o').replace('õ','o').replace('ô','o').replace('ú','u').replace('ç','c'):
                                val_text = r.get(k,'')
                                break
                    ordered_vals.append(val_text)

                # Append RESULTADO_N
                ordered_vals.append(result_N)
                out_lines.append(';'.join(ordered_vals))
            else:
                total_text = r['__total_line__']
                if has_valencio:
                    out_lines.append(total_text + ';' + f"{round(block_sum,2):.2f}")
                else:
                    out_lines.append(total_text + ';')

        block_idx += 1

    # If there were any lines before the first Manifesto block (like header), preserve them at the top
    # Build final output: prefix any leading non-block lines from original file
    final_lines = []
    in_first_block = False
    block_line_idx = 0
    # We'll iterate original lines and replace data lines for blocks sequentially using out_lines
    out_iter = iter(out_lines)
    for line in lines:
        if line.startswith('Manifesto:'):
            in_first_block = True
            # output the Manifesto line as-is
            final_lines.append(line)
            continue
        if not in_first_block:
            final_lines.append(line)
            continue
        # For lines inside blocks we consume from out_iter until we hit the next Manifesto or end
        # To keep things simple, when we encounter a data line (contains ';' and not starting with 'Total - Manifesto:')
        # we pop from out_iter and use that as replacement; for total lines we also pop.
        if line.startswith('Tipo;'):
            # skip original header lines inside block
            final_lines.append(line)
            continue
        if line.startswith('Total - Manifesto:'):
            try:
                replacement = next(out_iter)
            except StopIteration:
                replacement = line
            final_lines.append(replacement)
            # next Manifesto will reset
            continue
        if ';' in line:
            # assume it's a data row
            try:
                replacement = next(out_iter)
            except StopIteration:
                replacement = line + ';'
            final_lines.append(replacement)
            continue
        # fallback
        final_lines.append(line)

    processed_name = f"processed_{p.name}"
    out_path = p.parent / processed_name

    # If original was Excel, write Excel preserving columns and adding RESULTADO_N
    if suffix in ('.xlsx', '.xls'):
        # create workbook and write each line splitting by ';'
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        for r_idx, out_line in enumerate(final_lines, start=1):
            parts = out_line.split(';')
            for c_idx, val in enumerate(parts, start=1):
                new_ws.cell(row=r_idx, column=c_idx, value=val)
        new_path = p.parent / (p.stem + '_processed.xlsx')
        new_wb.save(str(new_path))
        return str(new_path)
    else:
        out_path.write_text('\n'.join(final_lines), encoding='utf-8')
        return str(out_path)

bp = Blueprint('frete', __name__, url_prefix='/frete')


@bp.route('/')
def index():
    return render_template('frete.html')


@bp.route('/importacao', methods=['GET', 'POST'])
def importacao():
    if request.method == 'POST':
        if 'files' not in request.files:
            flash('Nenhum arquivo enviado!', 'error')
            return redirect(url_for('frete.importacao'))

        files = request.files.getlist('files')
        if not files or files[0].filename == '':
            flash('Nenhum arquivo selecionado.', 'error')
            return redirect(url_for('frete.importacao'))

        # Pegar o tipo de arquivo selecionado pelo usuário
        tipo_arquivo = request.form.get('tipo_arquivo', 'valencio')  # default: valencio
        
        # Criar diretório de uploads do frete se não existir
        uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))
        os.makedirs(uploads_dir, exist_ok=True)

        for file in files:
            if file.filename:
                # Salvar arquivo temporário para análise
                temp_filepath = os.path.join(uploads_dir, file.filename)
                file.save(temp_filepath)
                
                try:
                    # Usar tipo escolhido pelo usuário (não detectar automaticamente)
                    if tipo_arquivo == 'manifesto':
                        # LÓGICA MANIFESTO: CTRL+C + CTRL+V
                        result = copy_manifesto_to_accumulated(temp_filepath)
                        
                        if 'error' in result:
                            flash(f'{file.filename}: Erro - {result["error"]}', 'error')
                        else:
                            flash(f'{file.filename}: Manifesto processado - {result["message"]}', 'success')
                        
                        # Manter arquivo original com timestamp
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        orig_name = f"MANIFESTO_{timestamp}_{file.filename}"
                        final_path = os.path.join(uploads_dir, orig_name)
                        if temp_filepath != final_path:
                            shutil.move(temp_filepath, final_path)
                    
                    elif tipo_arquivo == 'valencio':
                        # LÓGICA VALENCIO: Processar com cálculos
                        orig_name = file.filename
                        
                        # Padronizar nome como Valencio_Ajustes
                        ext = os.path.splitext(orig_name)[1]
                        save_name = f"VALENCIO_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{orig_name}"
                        
                        final_path = os.path.join(uploads_dir, save_name)
                        if temp_filepath != final_path:
                            shutil.move(temp_filepath, final_path)
                        temp_filepath = final_path

                        # Processar arquivo Valencio
                        try:
                            processed = process_frete_file(temp_filepath)
                            flash(f'{orig_name}: Arquivo Valencio processado -> {os.path.basename(processed)}', 'success')
                        except Exception as e:
                            flash(f'{orig_name}: Erro no processamento Valencio ({str(e)})', 'error')
                    
                    else:
                        flash(f'{file.filename}: Tipo de arquivo inválido', 'error')
                    
                except Exception as e:
                    flash(f'{file.filename}: Erro na análise do arquivo ({str(e)})', 'error')
                    # Remover arquivo temporário em caso de erro
                    if os.path.exists(temp_filepath):
                        os.remove(temp_filepath)

        return redirect(url_for('frete.importacao'))

    # Lista de arquivos enviados (pasta uploads/frete)
    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))
    uploads = []
    if os.path.isdir(uploads_dir):
        for fname in sorted(os.listdir(uploads_dir), reverse=True):
            fpath = os.path.join(uploads_dir, fname)
            try:
                mtime = os.path.getmtime(fpath)
                size = os.path.getsize(fpath)
                display_name = fname
                
                uploads.append({
                    'name': fname,
                    'display_name': display_name,
                    'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),
                    'size': size
                })
            except Exception:
                continue
    
    # Lista de arquivos acumulados
    accum_dir = os.path.join(uploads_dir, 'acumulados')
    accumulated = []
    if os.path.isdir(accum_dir):
        for fname in sorted(os.listdir(accum_dir)):
            fpath = os.path.join(accum_dir, fname)
            try:
                mtime = os.path.getmtime(fpath)
                size = os.path.getsize(fpath)
                
                accumulated.append({
                    'name': fname,
                    'mtime': datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S'),
                    'size': size
                })
            except Exception:
                continue

    return render_template('frete_importacao.html', uploads=uploads, accumulated=accumulated)


@bp.route('/uploads/<path:filename>')
def uploaded_file(filename):
    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))
    return send_from_directory(uploads_dir, filename)


@bp.route('/acumulados/<path:filename>')
def accumulated_file(filename):
    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'frete'))
    accum_dir = os.path.join(uploads_dir, 'acumulados')
    return send_from_directory(accum_dir, filename)
