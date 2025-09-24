"""
Módulo para processamento de arquivos VALENCIO
Lógica de cálculos e processamento de dados de frete
"""
import os
import openpyxl
import csv
import re
from datetime import datetime

def tentar_converter_numero_brasileiro(texto):
    """
    Tenta converter texto que pode ser um número em formato brasileiro
    """
    if not texto or not isinstance(texto, str):
        return None
    
    texto_limpo = texto.strip().replace(' ', '')
    
    if not texto_limpo:
        return None
    
    if not re.match(r'^-?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d+)?$', texto_limpo):
        return None
    
    try:
        # Formato brasileiro "1.234,56"
        if ',' in texto_limpo and '.' in texto_limpo:
            pos_virgula = texto_limpo.rfind(',')
            pos_ponto = texto_limpo.rfind('.')
            
            if pos_virgula > pos_ponto:
                numero_str = texto_limpo.replace('.', '').replace(',', '.')
                return float(numero_str)
            else:
                numero_str = texto_limpo.replace(',', '')
                return float(numero_str)
        
        elif ',' in texto_limpo:
            numero_str = texto_limpo.replace(',', '.')
            return float(numero_str)
        
        elif '.' in texto_limpo:
            partes = texto_limpo.split('.')
            if len(partes[-1]) <= 2:
                return float(texto_limpo)
            else:
                numero_str = texto_limpo.replace('.', '')
                return float(numero_str)
        
        else:
            return float(texto_limpo)
            
    except (ValueError, AttributeError):
        return None

def processar_valencio(arquivo_upload):
    """
    Processa arquivo de valencio/cálculos
    
    Args:
        arquivo_upload: FileStorage object do Flask
        
    Returns:
        dict: {"success": bool, "message": str, "dados": dict}
    """
    try:
        # Determinar nome original quando arquivo_upload for caminho string
        if isinstance(arquivo_upload, str):
            nome_original = os.path.basename(arquivo_upload)
        else:
            nome_original = getattr(arquivo_upload, 'filename', 'uploaded_valencio.xlsx')

        # Ler dados do Excel
        dados_excel = ler_excel_valencio(arquivo_upload)
        
        # Aplicar nova lógica automática do Valencio (apenas blocos com Valencio)
        print("🔄 Aplicando lógica APENAS em blocos com Valencio...")
        calculos = processar_calculos_valencio(dados_excel)
        
        # Modificar arquivo ORIGINAL (não criar novo)
        resultado_excel = salvar_valencio_csv(calculos, nome_original, arquivo_upload if isinstance(arquivo_upload, str) else None)
        
        return {
            "success": True,
            "message": f"Valencio processado: {resultado_excel['blocos_valencio']} blocos com Valencio de {resultado_excel['blocos']} total, R$ {resultado_excel['total']:.2f} - Arquivo original modificado",
            "dados": {
                "arquivo_modificado": resultado_excel['arquivo'],
                "total_geral": resultado_excel['total'],
                "blocos_processados": resultado_excel['blocos'],
                "blocos_com_valencio": resultado_excel['blocos_valencio'],
                "linhas_processadas": resultado_excel['linhas']
            }
        }
        
    except Exception as e:
        return {
            "success": False,
            "message": f"Erro ao processar valencio: {str(e)}",
            "dados": {}
        }

def ler_excel_valencio(arquivo_upload):
    """
    Lê Excel específico para valencio/cálculos
    """
    wb = openpyxl.load_workbook(arquivo_upload, data_only=True)
    ws = wb.active
    
    dados = []
    colunas = []
    
    primeira_linha = True
    for linha in ws.iter_rows(values_only=True):
        if not any(linha):
            continue
            
        if primeira_linha:
            colunas = [str(cel) if cel is not None else f"Coluna_{i+1}" for i, cel in enumerate(linha)]
            primeira_linha = False
            continue
            
        linha_processada = []
        for cel in linha:
            if cel is None:
                linha_processada.append("")
            elif isinstance(cel, (int, float)):
                # Converter para inteiro (sem casa decimal)
                linha_processada.append(int(cel))
            elif isinstance(cel, datetime):
                linha_processada.append(cel.strftime("%d/%m/%Y"))
            else:
                valor_str = str(cel).replace('\n', ' ').replace('\r', '').strip()
                linha_processada.append(valor_str)
        
        dados.append(linha_processada)
    
    return {
        "colunas": colunas,
        "dados": dados
    }

def processar_calculos_valencio(dados_excel):
    """
    Processa cálculos específicos do Valencio com nova lógica:
    - Trabalha APENAS em blocos que contêm FRIGORIFICO VALENCIO LTDA
    - Outros blocos ficam vazios na coluna Valor_Real
    - Modifica arquivo original (não cria novo)
    """
    resultados = {
        "dados_originais": dados_excel['dados'],
        "colunas": dados_excel['colunas'],
        "calculos": [],
        "total_geral": 0,
        "blocos_processados": 0,
        "blocos_com_valencio": 0
    }
    
    print(f"🔍 ESTRUTURA DO ARQUIVO VALENCIO:")
    print(f"   Total de linhas: {len(dados_excel['dados'])}")
    
    # Usar coluna N (índice 13) que já existe no arquivo original
    col_n = 13  # Coluna N (base 0)
    
    # Processar dados linha por linha
    bloco_atual = []
    total_geral = 0
    linha_cabecalho = None
    
    i = 0
    while i < len(dados_excel['dados']):
        linha = dados_excel['dados'][i]
        
        # Verificar coluna A para identificar estrutura
        col_a_valor = str(linha[0]).strip() if len(linha) > 0 else ""
        
        if col_a_valor.startswith("Manifesto:"):
            # Início de novo bloco
            print(f"📦 Analisando bloco: {col_a_valor}")
            
            # Processar bloco anterior se existir
            if bloco_atual and linha_cabecalho:
                tem_valencio = verificar_se_bloco_tem_valencio(bloco_atual, linha_cabecalho, dados_excel['dados'])
                
                if tem_valencio:
                    print(f"   🏭 BLOCO COM VALENCIO - Processando...")
                    total_bloco = processar_bloco_valencio_real(bloco_atual, linha_cabecalho, col_n, dados_excel['dados'])
                    total_geral += total_bloco
                    resultados['blocos_com_valencio'] += 1
                else:
                    print(f"   ⚪ Bloco sem Valencio - Deixando vazio")
                    # Deixar linhas do bloco vazias na coluna N
                    for idx in bloco_atual:
                        linha_bloco = dados_excel['dados'][idx]
                        while len(linha_bloco) <= col_n:
                            linha_bloco.append('')
                        linha_bloco[col_n] = ''
                
                resultados['blocos_processados'] += 1
            
            bloco_atual = []
            linha_cabecalho = None
            
            # Próxima linha deve ser cabeçalho
            if i + 1 < len(dados_excel['dados']):
                linha_cabecalho = i + 1
            
        elif col_a_valor.startswith("Total - Manifesto:"):
            # Final de bloco
            if bloco_atual and linha_cabecalho is not None:
                tem_valencio = verificar_se_bloco_tem_valencio(bloco_atual, linha_cabecalho, dados_excel['dados'])
                
                if tem_valencio:
                    print(f"   🏭 PROCESSANDO BLOCO COM VALENCIO...")
                    total_bloco = processar_bloco_valencio_real(bloco_atual, linha_cabecalho, col_n, dados_excel['dados'])
                    
                    # Colocar total na linha de Total - Manifesto
                    while len(linha) <= col_n:
                        linha.append('')
                    linha[col_n] = total_bloco
                    # Também modificar diretamente no dados original
                    dados_excel['dados'][i][col_n] = total_bloco
                    
                    total_geral += total_bloco
                    resultados['blocos_com_valencio'] += 1
                    print(f"   ✅ Bloco Valencio finalizado: R$ {total_bloco:.2f}")
                else:
                    print(f"   ⚪ Bloco sem Valencio - Total vazio")
                    # Deixar Total vazio também
                    while len(linha) <= col_n:
                        linha.append('')
                    linha[col_n] = ''
                    # Também modificar diretamente no dados original
                    dados_excel['dados'][i][col_n] = ''
                    
                    # Deixar linhas do bloco vazias na coluna N
                    for idx in bloco_atual:
                        linha_bloco = dados_excel['dados'][idx]
                        while len(linha_bloco) <= col_n:
                            linha_bloco.append('')
                        linha_bloco[col_n] = ''
                
                resultados['blocos_processados'] += 1
                
                resultados['calculos'].append({
                    "tipo": "bloco",
                    "manifesto": col_a_valor,
                    "tem_valencio": tem_valencio,
                    "total": total_bloco if tem_valencio else 0,
                    "linhas_no_bloco": len(bloco_atual)
                })
            
            bloco_atual = []
            linha_cabecalho = None
            
        elif col_a_valor == "Frete" and linha_cabecalho is not None:
            # Linha de dados dentro de um bloco
            bloco_atual.append(i)
        
        i += 1
    
    # Processar último bloco se não terminou com Total
    if bloco_atual and linha_cabecalho is not None:
        tem_valencio = verificar_se_bloco_tem_valencio(bloco_atual, linha_cabecalho, dados_excel['dados'])
        
        if tem_valencio:
            total_bloco = processar_bloco_valencio_real(bloco_atual, linha_cabecalho, col_n, dados_excel['dados'])
            total_geral += total_bloco
            resultados['blocos_com_valencio'] += 1
            print(f"   ✅ Último bloco Valencio finalizado: R$ {total_bloco:.2f}")
        else:
            # Deixar vazio
            for idx in bloco_atual:
                linha_bloco = dados_excel['dados'][idx]
                while len(linha_bloco) <= col_n:
                    linha_bloco.append('')
                linha_bloco[col_n] = ''
        
        resultados['blocos_processados'] += 1
    
    resultados['total_geral'] = total_geral
    resultados['colunas'] = dados_excel['colunas']  # Atualizar com nova coluna
    
    print(f"🎉 VALENCIO PROCESSADO:")
    print(f"   {resultados['blocos_processados']} blocos analisados")
    print(f"   {resultados['blocos_com_valencio']} blocos com Valencio processados")
    print(f"   Total geral (só Valencio): R$ {total_geral:.2f}")
    
    return resultados

def verificar_se_bloco_tem_valencio(indices_linhas, linha_cabecalho, dados):
    """
    Verifica se um bloco contém FRIGORIFICO VALENCIO LTDA
    """
    # Identificar coluna do Remetente
    cabecalho = dados[linha_cabecalho]
    col_d = None
    
    for i, header in enumerate(cabecalho):
        header_str = str(header).strip().lower()
        if 'remetente' in header_str:
            col_d = i
            break
    
    if col_d is None:
        return False
    
    # Verificar se alguma linha tem FRIGORIFICO VALENCIO
    for idx in indices_linhas:
        if idx >= len(dados):
            continue
        
        linha = dados[idx]
        cliente = str(linha[col_d]).strip().upper() if col_d < len(linha) else ""
        
        if "FRIGORIFICO VALENCIO LTDA" in cliente or "VALENCIO" in cliente:
            return True
    
    return False

def processar_bloco_valencio_real(indices_linhas, linha_cabecalho, col_n, dados):
    """
    Processa um bloco individual do manifesto Valencio
    Em blocos que TÊM Valencio:
    - FRIGORIFICO VALENCIO LTDA: J * 0,67 / 0,88
    - Outros clientes: COPIAR valor da coluna J
    """
    # Identificar colunas do cabeçalho
    cabecalho = dados[linha_cabecalho]
    col_d = None  # Remetente
    col_j = None  # Kg Taxado
    
    for i, header in enumerate(cabecalho):
        header_str = str(header).strip().lower()
        if 'remetente' in header_str:
            col_d = i
        elif 'kg taxado' in header_str:
            col_j = i
    
    if col_d is None or col_j is None:
        print(f"❌ Erro: Não encontrou colunas necessárias no cabeçalho")
        return 0
    
    print(f"   Colunas: D(Remetente)={col_d}, J(Kg Taxado)={col_j}")
    
    total_bloco = 0
    
    for idx in indices_linhas:
        if idx >= len(dados):
            continue
            
        linha = dados[idx]
        
        # Expandir linha se necessário
        while len(linha) <= col_n:
            linha.append('')
        
        # Pegar valores das colunas
        cliente = str(linha[col_d]).strip().upper() if col_d < len(linha) else ""
        kg_taxado_str = str(linha[col_j]).strip() if col_j < len(linha) else "0"
        
        try:
            # Converter Kg Taxado para número
            if kg_taxado_str and kg_taxado_str != "None":
                kg_taxado = float(kg_taxado_str.replace(',', '.'))
            else:
                kg_taxado = 0
            
            if kg_taxado <= 0:
                valor_real = 0
            elif "FRIGORIFICO VALENCIO LTDA" in cliente or "VALENCIO" in cliente:
                # Aplicar fórmula: J * 0,67 / 0,88
                valor_real = kg_taxado * 0.67 / 0.88
                print(f"     🏭 VALENCIO: {kg_taxado} * 0.67/0.88 = {valor_real:.2f}")
            else:
                # Outros clientes no bloco com Valencio: COPIAR valor de J
                valor_real = kg_taxado
                print(f"     📋 {cliente[:20]}...: {kg_taxado} (cópia de J)")
            
            # Colocar valor na coluna N
            linha[col_n] = round(valor_real, 2)
            # Também modificar diretamente no dados original para garantir
            dados[idx][col_n] = round(valor_real, 2)
            print(f"       ✓ Salvou {round(valor_real, 2)} na linha {idx}, coluna {col_n}")
            total_bloco += valor_real
            
        except (ValueError, TypeError) as e:
            print(f"❌ Erro convertendo Kg Taxado '{kg_taxado_str}': {e}")
            linha[col_n] = 0
    
    return round(total_bloco, 2)

def salvar_valencio_csv(calculos, nome_original, arquivo_original_path=None):
    """
    Modifica o arquivo Excel ORIGINAL com os cálculos do Valencio
    NÃO cria arquivo novo - sobrescreve o original
    """
    import openpyxl
    
    if arquivo_original_path and os.path.exists(arquivo_original_path):
        # Usar o arquivo original
        caminho_arquivo = arquivo_original_path
    else:
        # Fallback para pasta uploads/valencio
        pasta_valencio = "financeiro/uploads/valencio"
        caminho_arquivo = os.path.join(pasta_valencio, nome_original)
    
    try:
        # Abrir arquivo original
        print(f"📝 Modificando arquivo original: {caminho_arquivo}")
        wb = openpyxl.load_workbook(caminho_arquivo)
        ws = wb.active
        
        # Escrever dados modificados de volta
        for row_idx, linha in enumerate(calculos['dados_originais'], 1):  # Excel começa em 1
            for col_idx, valor in enumerate(linha, 1):  # Excel começa em 1
                if valor != '':  # Só escrever se não for vazio
                    ws.cell(row=row_idx, column=col_idx, value=valor)
        
        # Salvar arquivo original modificado
        wb.save(caminho_arquivo)
        
        print(f"✅ Arquivo original modificado com sucesso!")
        
        return {
            "arquivo": nome_original,
            "caminho": caminho_arquivo,
            "linhas": len(calculos['dados_originais']),
            "blocos": calculos['blocos_processados'],
            "blocos_valencio": calculos['blocos_com_valencio'],
            "total": calculos['total_geral']
        }
        
    except Exception as e:
        print(f"❌ Erro ao modificar arquivo original: {e}")
        # Fallback: criar arquivo novo se não conseguir modificar original
        pasta_saida = "financeiro/uploads/valencio"
        os.makedirs(pasta_saida, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_backup = f"VALENCIO_MODIFICADO_{timestamp}_{nome_original}"
        caminho_backup = os.path.join(pasta_saida, nome_backup)
        
        wb_novo = openpyxl.Workbook()
        ws_novo = wb_novo.active
        
        # Escrever dados
        for row_idx, linha in enumerate(calculos['dados_originais'], 1):
            for col_idx, valor in enumerate(linha, 1):
                ws_novo.cell(row=row_idx, column=col_idx, value=valor)
        
        wb_novo.save(caminho_backup)
        
        return {
            "arquivo": nome_backup,
            "caminho": caminho_backup,
            "linhas": len(calculos['dados_originais']),
            "blocos": calculos['blocos_processados'],
            "blocos_valencio": calculos['blocos_com_valencio'],
            "total": calculos['total_geral']
        }

def validar_estrutura_valencio(dados_excel):
    """
    Valida se o arquivo tem estrutura de valencio
    """
    colunas_esperadas = ['valor', 'frete', 'preco', 'total']
    colunas_encontradas = [col.lower() for col in dados_excel['colunas']]
    
    tem_coluna_valor = any(
        any(esperada in col for esperada in colunas_esperadas) 
        for col in colunas_encontradas
    )
    
    if not tem_coluna_valor:
        return False, "Nenhuma coluna de valor/frete encontrada"
    
    return True, "Estrutura válida para cálculos"