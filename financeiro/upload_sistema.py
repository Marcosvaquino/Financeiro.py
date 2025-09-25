from flask import Blueprint, render_template, request, flash, redirect, url_for
import os
import re
from werkzeug.utils import secure_filename
from datetime import datetime
from .manifesto import processar_manifesto, extrair_mes_ano_de_arquivo
import threading
import importlib.util
import traceback


def extrair_mes_ano_de_nome_arquivo(nome_arquivo):
    """
    Extrai mês-ano a partir do NOME do arquivo procurando por nomes/abreviações de meses.
    Retorna no formato MM-YY (ex: '09-25') ou None se não encontrar.
    Regras:
    - Aceita variações em português: jan, janeiro, fev, fevereiro, mar, março, etc.
    - Se encontrar ano com 4 dígitos (ex: 2025) usa ele; se encontrar apenas 2 dígitos usa-os;
      se não encontrar, usa o ano atual.
    """
    if not nome_arquivo:
        return None

    nome = os.path.basename(nome_arquivo).lower()

    # Mapas de meses (abreviações e nomes comuns em pt)
    meses_map = {
        'jan': 1, 'janeiro': 1,
        'fev': 2, 'fevereiro': 2,
        'mar': 3, 'marco': 3, 'março': 3,
        'abr': 4, 'abril': 4,
        'mai': 5, 'maio': 5,
        'jun': 6, 'junho': 6,
        'jul': 7, 'julho': 7,
        'ago': 8, 'agosto': 8,
        'set': 9, 'setembro': 9, 'setem': 9,
        'out': 10, 'outubro': 10,
        'nov': 11, 'novembro': 11,
        'dez': 12, 'dezembro': 12
    }

    # Regex para encontrar um token de mês (palavra)
    mes_regex = re.compile(r'\b(' + '|'.join(re.escape(k) for k in sorted(meses_map.keys(), key=len, reverse=True)) + r')\b', re.IGNORECASE)
    m = mes_regex.search(nome)
    mes = None
    if m:
        chave = m.group(1).lower()
        mes = meses_map.get(chave)

    # Tentar encontrar ano (4 dígitos preferencialmente)
    ano = None
    m4 = re.search(r'20\d{2}', nome)
    if m4:
        ano = int(m4.group(0))
    else:
        # procurar dois dígitos que pareçam ano (ex: _25, -25, 25.xlsx)
        m2 = re.search(r'[^0-9](\d{2})[^0-9]', f' {nome} ')
        if m2:
            ano_candidate = int(m2.group(1))
            # Heurística: se candidato entre 0 e 30 -> assumir 2000+xx, senão 1900+xx
            if 0 <= ano_candidate <= 99:
                ano = 2000 + ano_candidate if ano_candidate <= 50 else 1900 + ano_candidate

    # Se não encontrou mês, tentar extrair par numérico MM-YY ou MM-YYYY
    if mes is None:
        m_num = re.search(r'\b(0[1-9]|1[0-2])[\-_/]?(\d{2,4})\b', nome)
        if m_num:
            try:
                mes = int(m_num.group(1))
                yy = m_num.group(2)
                if len(yy) == 2:
                    ano = 2000 + int(yy)
                else:
                    ano = int(yy)
            except Exception:
                pass

    # Se ainda não tiver ano, usar o ano atual
    if ano is None:
        ano = datetime.now().year

    if mes is None:
        return None

    return f"{mes:02d}-{str(ano)[-2:]}"



def schedule_acumulador_background():
    """Dispara criar_manifesto_acumulado.py em background. Usa um lock por arquivo para evitar concorrência."""
    try:
        uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads'))
        lockfile = os.path.join(uploads_dir, 'manifesto_acumulado.lock')
        # tentar criar lock atomically
        fd = None
        try:
            fd = os.open(lockfile, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(fd, str(os.getpid()).encode('utf-8'))
            os.close(fd)
            fd = None
        except FileExistsError:
            # já em execução
            return

        def _runner():
            try:
                # carregar o módulo pelo caminho absoluto
                script_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'criar_manifesto_acumulado.py'))
                spec = importlib.util.spec_from_file_location('criar_manifesto_acumulado', script_path)
                mod = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(mod)
                # chamar função principal
                try:
                    mod.criar_manifesto_acumulado()
                except Exception:
                    # se módulo tiver comportamento diferente, tentar chamar via main
                    if hasattr(mod, '__main__'):
                        pass
            except Exception:
                traceback.print_exc()
            finally:
                try:
                    if os.path.exists(lockfile):
                        os.remove(lockfile)
                except Exception:
                    pass

        t = threading.Thread(target=_runner, daemon=True)
        t.start()
    finally:
        try:
            if fd:
                os.close(fd)
        except Exception:
            pass

bp = Blueprint('upload_sistema', __name__, url_prefix='/upload')

@bp.route('/')
def index():
    return render_template('upload.html')

@bp.route('/processar', methods=['POST'])
def processar():
    if 'arquivo' not in request.files:
        flash('Nenhum arquivo foi selecionado!')
        return redirect(url_for('upload_sistema.index'))
    
    arquivo = request.files['arquivo']
    if arquivo.filename == '':
        flash('Nenhum arquivo foi selecionado!')
        return redirect(url_for('upload_sistema.index'))
    
    # Capturar o tipo selecionado pelo usuário
    tipo_usuario = request.form.get('tipo_arquivo', 'manifesto')
    
    # DETECÇÃO AUTOMÁTICA pelo nome do arquivo
    nome_arquivo = arquivo.filename.lower()
    tipo_detectado = None
    
    if 'manifesto' in nome_arquivo or 'manifest' in nome_arquivo:
        tipo_detectado = 'manifesto'
    elif 'valencio' in nome_arquivo or 'valen' in nome_arquivo or 'calculo' in nome_arquivo:
        tipo_detectado = 'valencio'
    elif 'pamplona' in nome_arquivo or 'pamp' in nome_arquivo:
        tipo_detectado = 'pamplona'
    
    # Verificar se há conflito entre seleção e detecção
    if tipo_detectado and tipo_detectado != tipo_usuario:
        flash(f'⚠️ ATENÇÃO: Você selecionou "{tipo_usuario.upper()}" mas o arquivo parece ser "{tipo_detectado.upper()}" (pelo nome). Processando como "{tipo_detectado.upper()}"!')
        tipo_final = tipo_detectado
    else:
        tipo_final = tipo_usuario

    # Se for manifesto, tentar extrair mês/ano e renomear antes de processar
    saved_path = None
    if tipo_final == 'manifesto':
        # Extrair mês/ano do arquivo (usa file-like)
        try:
            mes_ano = extrair_mes_ano_de_arquivo(arquivo)
        except Exception:
            mes_ano = None

        # Vamos usar o nome curto solicitado: Manifesto_Frete_{MM-YY}.xlsx
        ext = os.path.splitext(secure_filename(arquivo.filename))[1] or '.xlsx'
        novo_nome = f"Manifesto_Frete_{mes_ano}{ext}" if mes_ano else f"Manifesto_Frete_unknown{ext}"

        uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'manifestos'))
        os.makedirs(uploads_dir, exist_ok=True)
        destino = os.path.join(uploads_dir, novo_nome)

        # Se existir, sobrescrevemos (usuário quer apenas um arquivo por mês)
        arquivo.stream.seek(0)
        arquivo.save(destino)
        saved_path = destino
    
    # Por enquanto só mostra que recebeu o arquivo com o tipo
    if tipo_final == 'manifesto':
        # Integração direta no arquivo original (inline)
        def integrar_manifesto_inline(arquivo_path):
            import shutil
            import openpyxl
            from datetime import datetime
            from .veiculo_helper import VeiculoHelper
            from .cliente_helper import ClienteHelper
            
            try:
                # Backup
                backup_dir = os.path.join(os.path.dirname(arquivo_path), '..', '..', 'backups')
                os.makedirs(backup_dir, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = os.path.join(backup_dir, f"backup_{timestamp}_{os.path.basename(arquivo_path)}")
                shutil.copy2(arquivo_path, backup_path)
                
                print(f"💾 Backup: {os.path.basename(backup_path)}")
                print(f'🚛 INTEGRANDO: {os.path.basename(arquivo_path)}')
                
                # Carregar workbook
                wb = openpyxl.load_workbook(arquivo_path, data_only=True)
                ws = wb.active
                
                # Adicionar colunas
                nova_col_status = ws.max_column + 1
                nova_col_tipologia = nova_col_status + 1
                nova_col_cliente = nova_col_status + 2
                
                ws.cell(1, nova_col_status, 'Status_Veiculo')
                ws.cell(1, nova_col_tipologia, 'Tipologia') 
                ws.cell(1, nova_col_cliente, 'Cliente_Real')
                
                # Encontrar colunas de dados - usando lógica exata do temp_integrar.py
                col_placa = col_cliente = None
                print(f'🔍 ANALISANDO CABEÇALHOS:')
                
                for col in range(1, ws.max_column):
                    valor = ws.cell(1, col).value
                    if valor:
                        valor_upper = str(valor).upper()
                        print(f'   Col {col}: "{valor}" -> "{valor_upper}"')
                        
                        # Para placa: exatamente "VEÍCULO" ou "PLACA" (sem acentos nos dados)
                        if valor_upper in ['VEÍCULO', 'VEICULO', 'PLACA']:
                            if not col_placa:  # Pegar a primeira ocorrência
                                col_placa = col
                                print(f'   🚚 PLACA ENCONTRADA: Col {col}')
                        
                        # Para cliente: exatamente "CLASSIFICAÇÃO"
                        elif valor_upper in ['CLASSIFICAÇÃO', 'CLASSIFICACAO']:
                            if not col_cliente:  # Pegar a primeira ocorrência
                                col_cliente = col
                                print(f'   👥 CLIENTE ENCONTRADO: Col {col}')
                
                print(f'📍 RESULTADO: Placa=Col{col_placa}, Cliente=Col{col_cliente}')
                
                # Coletar dados únicos
                placas = set()
                clientes = set() 
                print(f'📊 COLETANDO DADOS ÚNICOS...')
                for row in range(2, min(ws.max_row + 1, 6)):  # Amostra pequena para debug
                    linha_dados = []
                    for col in range(1, min(ws.max_column + 1, 15)):
                        valor = ws.cell(row, col).value
                        linha_dados.append(f'Col{col}:"{valor}"')
                    print(f'   Linha {row}: {" | ".join(linha_dados)}')
                
                for row in range(2, ws.max_row + 1):
                    if col_placa:
                        placa = ws.cell(row, col_placa).value
                        if placa: placas.add(str(placa).upper().strip())
                    if col_cliente:
                        cliente = ws.cell(row, col_cliente).value
                        if cliente: clientes.add(str(cliente).upper().strip())
                
                print(f'📊 DADOS ÚNICOS: {len(placas)} placas, {len(clientes)} clientes')
                if placas:
                    print(f'   Placas: {list(placas)[:3]}...')
                if clientes:
                    print(f'   Clientes: {list(clientes)[:3]}...')
                
                # Buscar dados usando nomes REAIS (não ajustados)
                dados_veiculos = VeiculoHelper.buscar_multiplas_placas(list(placas)) if placas else {}
                dados_clientes = ClienteHelper.buscar_multiplos_nomes_reais(list(clientes)) if clientes else {}
                
                veiculos_encontrados = sum(1 for v in dados_veiculos.values() if v.get('encontrado'))
                clientes_encontrados = sum(1 for c in dados_clientes.values() if c.get('encontrado'))
                print(f'✅ HELPERS: {veiculos_encontrados} veículos, {clientes_encontrados} clientes encontrados')
                
                # Integrar
                print(f'🔄 INTEGRANDO {ws.max_row-1} LINHAS...')
                for row in range(2, ws.max_row + 1):
                    if col_placa:
                        placa = ws.cell(row, col_placa).value
                        if placa:
                            placa_norm = str(placa).upper().strip()
                            veiculo = dados_veiculos.get(placa_norm, {})
                            status_val = veiculo.get('status') or '0'
                            tipologia_val = veiculo.get('tipologia') or '0'
                            ws.cell(row, nova_col_status, status_val)
                            ws.cell(row, nova_col_tipologia, tipologia_val)
                        else:
                            ws.cell(row, nova_col_status, '0')
                            ws.cell(row, nova_col_tipologia, '0')
                    else:
                        ws.cell(row, nova_col_status, '0')
                        ws.cell(row, nova_col_tipologia, '0')
                    
                    if col_cliente:
                        cliente = ws.cell(row, col_cliente).value
                        if cliente:
                            cliente_norm = str(cliente).upper().strip() 
                            cliente_dados = dados_clientes.get(cliente_norm, {})
                            # Usar nome_ajustado em vez de nome_real
                            cliente_val = cliente_dados.get('nome_ajustado') or '0'
                            ws.cell(row, nova_col_cliente, cliente_val)
                        else:
                            ws.cell(row, nova_col_cliente, '0')
                    else:
                        ws.cell(row, nova_col_cliente, '0')
                    
                    if row % 100 == 0:
                        print(f'   Processadas {row-1} linhas...')
                
                print('💾 SALVANDO ARQUIVO...')
                
                wb.save(arquivo_path)
                
                veiculos_encontrados = sum(1 for v in dados_veiculos.values() if v.get('encontrado'))
                clientes_encontrados = sum(1 for c in dados_clientes.values() if c.get('encontrado'))
                
                return {
                    "success": True,
                    "message": f"✅ {ws.max_row-1} registros | 🚚 {veiculos_encontrados} veículos | 👥 {clientes_encontrados} clientes",
                    "backup_path": backup_path
                }
                
            except Exception as e:
                return {"success": False, "message": f"❌ Erro: {str(e)}", "backup_path": None}
        
        if saved_path:
            resultado = integrar_manifesto_inline(saved_path)
        else:
            uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'manifestos'))
            os.makedirs(uploads_dir, exist_ok=True)
            temp_path = os.path.join(uploads_dir, secure_filename(arquivo.filename))
            arquivo.save(temp_path)
            resultado = integrar_manifesto_inline(temp_path)
        
        if resultado['success']:
            backup_info = f" (backup: {os.path.basename(resultado['backup_path'])})" if resultado['backup_path'] else ""
            flash(f'✅ MANIFESTO INTEGRADO: {resultado["message"]}{backup_info}')
            # Agendar atualização do Manifesto_Acumulado em background
            try:
                schedule_acumulador_background()
            except Exception:
                # não bloquear a resposta principal
                pass
        else:
            flash(f'❌ ERRO INTEGRAÇÃO: {resultado["message"]}')
    elif tipo_final == 'valencio':
        # Para Valencio: extrair mês/ano e renomear para Valencio_Frete_{MM-YY}
        try:
            mes_ano = extrair_mes_ano_de_nome_arquivo(arquivo.filename if hasattr(arquivo, 'filename') else str(arquivo))
        except Exception:
            mes_ano = None

        ext = os.path.splitext(secure_filename(arquivo.filename))[1] or '.xlsx'
        novo_nome = f"Valencio_Frete_{mes_ano}{ext}" if mes_ano else f"Valencio_Frete_unknown{ext}"

        uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'valencio'))
        os.makedirs(uploads_dir, exist_ok=True)
        destino = os.path.join(uploads_dir, novo_nome)

        # salvar (sobrescrever se necessário)
        arquivo.stream.seek(0)
        arquivo.save(destino)

        from .valencio import processar_valencio
        resultado = processar_valencio(destino)
        if resultado['success']:
            flash(f'✅ VALENCIO: {resultado["message"]}')
        else:
            flash(f'❌ ERRO VALENCIO: {resultado["message"]}')
    elif tipo_final == 'pamplona':
        # Para Pamplona: extrair mês/ano e renomear para Pamplona_Frete_{MM-YY}
        try:
            mes_ano = extrair_mes_ano_de_nome_arquivo(arquivo.filename if hasattr(arquivo, 'filename') else str(arquivo))
        except Exception:
            mes_ano = None

        ext = os.path.splitext(secure_filename(arquivo.filename))[1] or '.xlsx'
        novo_nome = f"Pamplona_Frete_{mes_ano}{ext}" if mes_ano else f"Pamplona_Frete_unknown{ext}"

        uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads', 'pamplona'))
        os.makedirs(uploads_dir, exist_ok=True)
        destino = os.path.join(uploads_dir, novo_nome)

        # salvar (sobrescrever se necessário)
        arquivo.stream.seek(0)
        arquivo.save(destino)

        from .pamplona import processar_pamplona
        resultado = processar_pamplona(destino)
        if resultado['success']:
            flash(f'✅ PAMPLONA: {resultado["message"]}')
        else:
            flash(f'❌ ERRO PAMPLONA: {resultado["message"]}')
    else:
        flash(f'❓ Arquivo "{arquivo.filename}" de tipo indefinido foi recebido!')
    
    return redirect(url_for('upload_sistema.index'))