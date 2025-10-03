from flask import Blueprint, render_template, request, jsonify
import sqlite3
import openpyxl
import os
from datetime import datetime, timedelta
from collections import defaultdict
from decimal import Decimal
from .database import get_connection

bp = Blueprint('painel_frete', __name__, url_prefix='/frete/painel')

@bp.route('/v2')
def index_v2():
    """Página do Dashboard V2 - Versão Limpa"""
    return render_template('painel_frete_v2.html')

def get_db_connection():
    """Retorna conexão com o banco de dados usando o wrapper central."""
    conn = get_connection()
    conn.row_factory = sqlite3.Row
    return conn

def extrair_dados_manifesto_real(filtros=None):
    """Extrai dados reais do manifesto para gráfico Frete Correto vs Despesas Gerais
    
    Args:
        filtros (dict): Dicionário com filtros opcionais:
            - perfil: Status_Veiculo (Col 25)
            - clientes: Cliente_Real (Col 27) 
            - veiculos: Veículo (Col 4)
            - mes: Data - mês (Col 3)
            - ano: Data - ano (Col 3)
    """
    arquivo_manifesto = os.path.join('financeiro', 'uploads', 'Manifesto_Acumulado.xlsx')
    
    if not os.path.exists(arquivo_manifesto):
        print(f"Arquivo não encontrado: {arquivo_manifesto}")
        return None
        
    try:
        wb = openpyxl.load_workbook(arquivo_manifesto, read_only=True)
        ws = wb.active
        
        # Headers esperados
        # Col 30: Frete Correto, Col 29: Despesas Gerais, Col 3: Data
        dados_por_dia = defaultdict(lambda: {'frete_correto': 0, 'despesas_gerais': 0})
        totais_mensais = {'frete_correto': 0, 'despesas_gerais': 0}
        
        linha_inicial = 2  # Pula header
        for row in ws.iter_rows(min_row=linha_inicial, values_only=True):
            if not row or len(row) < 30:
                continue
                
            try:
                # Extrair dados para filtragem
                status_veiculo = str(row[24]).upper() if row[24] else ""  # Col 25 (Status_Veiculo)
                cliente_real = str(row[26]).upper() if row[26] else ""    # Col 27 (Cliente_Real)
                veiculo = str(row[3]).upper() if row[3] else ""           # Col 4 (Veículo)
                
                # Extrair data (Col 3) - formato datetime ou string
                data_obj = row[2] if row[2] else None
                if not data_obj:
                    continue
                    
                # Parse da data
                if hasattr(data_obj, 'day'):  # É datetime
                    dia = data_obj.day
                    mes = data_obj.month
                    ano = data_obj.year
                elif "/" in str(data_obj):  # String formato dd/mm/yyyy
                    partes = str(data_obj).split("/")
                    if len(partes) >= 3:
                        dia = int(partes[0])
                        mes = int(partes[1])
                        ano = int(partes[2])
                    else:
                        continue
                elif "-" in str(data_obj):  # String formato yyyy-mm-dd
                    partes = str(data_obj).split("-")
                    if len(partes) >= 3:
                        ano = int(partes[0])
                        mes = int(partes[1])
                        dia = int(partes[2].split()[0])  # Remove time se houver
                    else:
                        continue
                else:
                    continue
                
                # Aplicar filtros se fornecidos
                if filtros:
                    # Filtro por perfil (Status_Veiculo)
                    if filtros.get('perfil') and filtros['perfil'].upper() not in ['TODOS', 'AGREGADO']:
                        if status_veiculo != filtros['perfil'].upper():
                            continue
                    
                    # Filtro por clientes (Cliente_Real)
                    if filtros.get('clientes') and len(filtros['clientes']) > 0 and filtros['clientes'][0] != '':
                        clientes_filtro = [c.upper() for c in filtros['clientes'] if c]
                        if clientes_filtro and cliente_real not in clientes_filtro:
                            continue
                    
                    # Filtro por veículos
                    if filtros.get('veiculos') and len(filtros['veiculos']) > 0 and filtros['veiculos'][0] != '':
                        veiculos_filtro = [v.upper() for v in filtros['veiculos'] if v]
                        if veiculos_filtro and veiculo not in veiculos_filtro:
                            continue
                    
                    # Filtro por mês
                    if filtros.get('mes') and filtros['mes']:
                        mes_nomes = {
                            'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4, 'MAI': 5, 'JUN': 6,
                            'JUL': 7, 'AGO': 8, 'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12
                        }
                        mes_filtro = mes_nomes.get(filtros['mes'].upper())
                        if mes_filtro and mes != mes_filtro:
                            continue
                    
                    # Filtro por ano
                    if filtros.get('ano') and filtros['ano']:
                        try:
                            ano_filtro = int(filtros['ano'])
                            if ano != ano_filtro:
                                continue
                        except (ValueError, TypeError):
                            continue
                
                # Extrair valores - Col 30: Frete Correto, Col 29: Despesas Gerais
                def safe_float(valor):
                    if valor is None or valor == '':
                        return 0.0
                    try:
                        return float(valor)
                    except (ValueError, TypeError):
                        return 0.0
                
                frete_correto = safe_float(row[29])  # Col 30 (index 29)
                despesas_gerais = safe_float(row[28])  # Col 29 (index 28)
                
                # Agrupar por dia
                dados_por_dia[dia]['frete_correto'] += frete_correto
                dados_por_dia[dia]['despesas_gerais'] += despesas_gerais
                
                # Totais mensais
                totais_mensais['frete_correto'] += frete_correto
                totais_mensais['despesas_gerais'] += despesas_gerais
                
            except (ValueError, IndexError, TypeError) as e:
                continue
                
        wb.close()
        
        # Organizar dados por dia (1-31) - ACUMULADOS
        dados_finais = []
        frete_acumulado = 0
        despesas_acumuladas = 0
        
        for dia in range(1, 32):
            # Acumular valores até o dia atual
            frete_acumulado += dados_por_dia[dia]['frete_correto']
            despesas_acumuladas += dados_por_dia[dia]['despesas_gerais']
            
            dados_finais.append({
                'dia': dia,
                'frete_correto': frete_acumulado,
                'despesas_gerais': despesas_acumuladas
            })
        
        return {
            'dados_diarios': dados_finais,
            'totais_mensais': totais_mensais
        }
        
    except Exception as e:
        print(f"Erro ao ler manifesto: {e}")
        return None

def obter_opcoes_filtros():
    """Obtém as opções disponíveis para os filtros baseado nos dados do manifesto"""
    arquivo_manifesto = os.path.join('financeiro', 'uploads', 'Manifesto_Acumulado.xlsx')
    
    if not os.path.exists(arquivo_manifesto):
        # Retorna opções padrão se não conseguir ler o arquivo
        return {
            'perfis': ['AGREGADO', 'FIXO', 'SPOT'],
            'clientes': ['ADORO', 'GOLD PAO', 'Marfrig', 'MEGGS', 'FRZ Log'],
            'veiculos': ['AMX0E01', 'ARS5C20', 'ASN5A61'],
            'anos': [2024, 2025]
        }
    
    try:
        wb = openpyxl.load_workbook(arquivo_manifesto, read_only=True)
        ws = wb.active
        
        perfis = set()
        clientes = set() 
        veiculos = set()
        anos = set()
        
        linha_inicial = 2  # Pula header
        for row in ws.iter_rows(min_row=linha_inicial, values_only=True):
            if not row or len(row) < 30:
                continue
                
            try:
                # Status_Veiculo (Col 25)
                if row[24]:
                    perfis.add(str(row[24]).strip())
                
                # Cliente_Real (Col 27)
                if row[26]:
                    clientes.add(str(row[26]).strip())
                
                # Veículo (Col 4)
                if row[3]:
                    veiculos.add(str(row[3]).strip())
                
                # Ano da Data (Col 3)
                data_obj = row[2]
                if data_obj:
                    if hasattr(data_obj, 'year'):
                        anos.add(data_obj.year)
                    elif "/" in str(data_obj):
                        partes = str(data_obj).split("/")
                        if len(partes) >= 3:
                            anos.add(int(partes[2]))
                    elif "-" in str(data_obj):
                        partes = str(data_obj).split("-")
                        if len(partes) >= 3:
                            anos.add(int(partes[0]))
                            
            except (ValueError, IndexError, TypeError):
                continue
        
        wb.close()
        
        return {
            'perfis': ['TODOS'] + sorted([p for p in perfis if p and p != '0']),  # Remove valores vazios/zero
            'clientes': sorted([c for c in clientes if c and c != '0']),
            'veiculos': sorted([v for v in veiculos if v and v != '0']),
            'anos': sorted(list(anos), reverse=True)
        }
        
    except Exception as e:
        print(f"Erro ao obter opções de filtros: {e}")
        return {
            'perfis': ['AGREGADO', 'FIXO', 'SPOT'],
            'clientes': ['ADORO', 'GOLD PAO', 'Marfrig', 'MEGGS'],
            'veiculos': ['AMX0E01', 'ARS5C20', 'ASN5A61'],
            'anos': [2024, 2025]
        }

@bp.route('/')
def index():
    """Página principal do painel de frete - dashboard completo"""
    # Obter opções dinâmicas dos filtros
    opcoes_filtros = obter_opcoes_filtros()
    
    dados_dashboard = {
        'metricas': {
            'frete_receber': 2007716.00,
            'frete_pagar': 1357920.00,
            'diferenca': 649795.00,
            'produtividade': 92.36
        },
        'filtros': {
            'mes_atual': 'AGO',  # Mês atual do manifesto
            'ano_atual': max(opcoes_filtros['anos']) if opcoes_filtros['anos'] else 2025,
            'perfil_selecionado': 'TODOS',
            'opcoes_perfis': opcoes_filtros['perfis'],
            'opcoes_clientes': opcoes_filtros['clientes'],
            'opcoes_veiculos': opcoes_filtros['veiculos'],
            'opcoes_anos': opcoes_filtros['anos']
        },
        'frete_diario': gerar_dados_frete_diario(),
        'frete_mensal': gerar_dados_frete_mensal(),
        'clientes_participacao': gerar_dados_clientes()
    }
    
    return render_template('painel_frete.html', dados=dados_dashboard)

def gerar_dados_frete_diario(filtros=None):
    """Gera dados reais para o gráfico Frete Correto vs Despesas Gerais"""
    dados_manifesto = extrair_dados_manifesto_real(filtros)
    
    if dados_manifesto:
        # Usar dados reais do manifesto
        return {
            'dados_diarios': dados_manifesto['dados_diarios'],
            'totais_mensais': dados_manifesto['totais_mensais']
        }
    else:
        # Fallback para dados simulados se não conseguir ler o manifesto
        print("Usando dados simulados - manifesto não disponível")
        dados = []
        totais = {'frete_correto': 0, 'despesas_gerais': 0}
        
        for dia in range(1, 32):
            # Simula variação diária
            variacao = (dia * 0.02) + 0.8
            frete_correto = 65000 * variacao
            despesas_gerais = 45000 * variacao
            
            dados.append({
                'dia': dia,
                'frete_correto': frete_correto,
                'despesas_gerais': despesas_gerais
            })
            
            totais['frete_correto'] += frete_correto
            totais['despesas_gerais'] += despesas_gerais
        
        return {
            'dados_diarios': dados,
            'totais_mensais': totais
        }

def gerar_dados_frete_mensal():
    """Gera dados simulados para o gráfico mensal"""
    meses = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET']
    dados = []
    
    base_valores = [
        {'receber': 2224294, 'pagar': 1564138, 'produtividade': 29.7},
        {'receber': 2148782, 'pagar': 1451867, 'produtividade': 32.3},
        {'receber': 2297528, 'pagar': 1525256, 'produtividade': 33.6},
        {'receber': 2597444, 'pagar': 1701265, 'produtividade': 34.5},
        {'receber': 2571763, 'pagar': 1756192, 'produtividade': 31.7},
        {'receber': 2778534, 'pagar': 1909397, 'produtividade': 31.3},
        {'receber': 3343969, 'pagar': 2211157, 'produtividade': 33.9},
        {'receber': 2927822, 'pagar': 1888939, 'produtividade': 35.5},
        {'receber': 2007716, 'pagar': 1357920, 'produtividade': 32.4}
    ]
    
    for i, mes in enumerate(meses):
        if i < len(base_valores):
            dados.append({
                'mes': mes,
                'frete_receber': base_valores[i]['receber'],
                'frete_pagar': base_valores[i]['pagar'],
                'produtividade': base_valores[i]['produtividade']
            })
    
    return dados

def gerar_dados_clientes():
    """Gera dados de participação por cliente"""
    return [
        {'nome': 'FRZ Log', 'percentual': 31.76, 'cor': '#FF6B35'},
        {'nome': 'ADORO', 'percentual': 22.78, 'cor': '#F7931E'},
        {'nome': 'Marfrig ( BRF )', 'percentual': 19.20, 'cor': '#FFD23F'},
        {'nome': 'MINERVA', 'percentual': 14.18, 'cor': '#06D6A0'},
        {'nome': 'Friboi', 'percentual': 5.58, 'cor': '#118AB2'},
        {'nome': 'Transferência', 'percentual': 3.65, 'cor': '#073B4C'},
        {'nome': 'MEGGS', 'percentual': 2.14, 'cor': '#8B5CF6'},
        {'nome': 'GOLD PAO', 'percentual': 0.64, 'cor': '#EF4444'}
    ]

def gerar_metricas():
    """Gera métricas para os cards do dashboard"""
    import random
    
    frete_receber = random.uniform(180000, 220000)
    frete_pagar = random.uniform(150000, 180000)
    diferenca = frete_receber - frete_pagar
    produtividade = (diferenca / frete_receber) * 100
    
    return {
        'frete_receber': frete_receber,
        'frete_pagar': frete_pagar,
        'diferenca': diferenca,
        'produtividade': produtividade
    }

@bp.route('/api/dados', methods=['POST'])
def api_dados_post():
    """API endpoint para atualização dinâmica dos dados"""
    try:
        filtros = request.get_json()
        print(f"🔍 Filtros recebidos: {filtros}")  # Debug
        
        # Gerar dados baseados nos filtros reais
        dados_filtrados = gerar_dados_frete_diario(filtros)
        print(f"📊 Dados filtrados gerados: {bool(dados_filtrados)}")
        
        if dados_filtrados and 'totais_mensais' in dados_filtrados:
            totais = dados_filtrados['totais_mensais']
            print(f"💰 Totais: FC=R${totais['frete_correto']:,.2f}, DG=R${totais['despesas_gerais']:,.2f}")
        
        dados = {
            'frete_diario': dados_filtrados,
            'frete_mensal': gerar_dados_frete_mensal(), 
            'clientes_participacao': gerar_dados_clientes(),
            'metricas': gerar_metricas()
        }
        
        print(f"✅ Enviando resposta da API")
        return jsonify(dados)
        
    except Exception as e:
        print(f"❌ Erro na API: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@bp.route('/api/dados/<tipo>')
def api_dados(tipo):
    """API para buscar dados específicos via AJAX"""
    if tipo == 'frete_diario':
        return jsonify(gerar_dados_frete_diario())
    elif tipo == 'frete_mensal':
        return jsonify(gerar_dados_frete_mensal())
    elif tipo == 'clientes':
        return jsonify(gerar_dados_clientes())
    elif tipo == 'metricas':
        return jsonify(gerar_metricas())
    else:
        return jsonify({'error': 'Tipo de dados não encontrado'}), 404