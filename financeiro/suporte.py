from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
import sqlite3
from datetime import datetime
from pathlib import Path
from .database import get_connection

bp = Blueprint('suporte', __name__, url_prefix='/frete/suporte')

def get_db_connection():
    """Retorna conexão com o banco de dados usando o wrapper central (com timeout/PRAGMA)."""
    conn = get_connection()
    conn.row_factory = sqlite3.Row
    return conn

@bp.route('/')
def index():
    """Página principal do suporte"""
    return render_template('suporte.html')

@bp.route('/veiculos')
def veiculos():
    """Lista todos os veículos cadastrados"""
    conn = get_db_connection()
    veiculos = conn.execute('''
        SELECT placa, status, tipologia, data_cadastro, ativo 
        FROM veiculos_suporte 
        ORDER BY placa
    ''').fetchall()
    conn.close()
    return render_template('suporte_veiculos.html', veiculos=veiculos)

@bp.route('/veiculos/add', methods=['POST'])
def add_veiculo():
    """Adiciona um novo veículo"""
    placa = request.form.get('placa', '').strip().upper()
    status = request.form.get('status', '').strip().upper()
    tipologia = request.form.get('tipologia', '').strip().upper()
    
    if not placa or not status or not tipologia:
        flash('Todos os campos são obrigatórios', 'error')
        return redirect(url_for('suporte.veiculos'))
    
    if status not in ['FIXO', 'SPOT']:
        flash('Status deve ser FIXO ou SPOT', 'error')
        return redirect(url_for('suporte.veiculos'))
    
    conn = get_db_connection()
    try:
        conn.execute('''
            INSERT INTO veiculos_suporte (placa, status, tipologia, data_cadastro, ativo)
            VALUES (?, ?, ?, ?, ?)
        ''', (placa, status, tipologia, datetime.now(), True))
        conn.commit()
        flash(f'Veículo {placa} cadastrado com sucesso', 'success')
    except sqlite3.IntegrityError:
        flash(f'Veículo {placa} já está cadastrado', 'error')
    except Exception as e:
        flash(f'Erro ao cadastrar veículo: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('suporte.veiculos'))

@bp.route('/veiculos/edit/<placa>', methods=['POST'])
def edit_veiculo(placa):
    """Edita um veículo existente"""
    status = request.form.get('status', '').strip().upper()
    tipologia = request.form.get('tipologia', '').strip().upper()
    ativo = request.form.get('ativo') == 'on'
    
    if not status or not tipologia:
        flash('Status e tipologia são obrigatórios', 'error')
        return redirect(url_for('suporte.veiculos'))
    
    if status not in ['FIXO', 'SPOT']:
        flash('Status deve ser FIXO ou SPOT', 'error')
        return redirect(url_for('suporte.veiculos'))
    
    conn = get_db_connection()
    try:
        conn.execute('''
            UPDATE veiculos_suporte 
            SET status = ?, tipologia = ?, ativo = ?
            WHERE placa = ?
        ''', (status, tipologia, ativo, placa))
        conn.commit()
        flash(f'Veículo {placa} atualizado com sucesso', 'success')
    except Exception as e:
        flash(f'Erro ao atualizar veículo: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('suporte.veiculos'))

@bp.route('/veiculos/delete/<placa>', methods=['POST'])
def delete_veiculo(placa):
    """Remove um veículo"""
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM veiculos_suporte WHERE placa = ?', (placa,))
        conn.commit()
        flash(f'Veículo {placa} removido com sucesso', 'success')
    except Exception as e:
        flash(f'Erro ao remover veículo: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('suporte.veiculos'))

@bp.route('/veiculos/detect-novas')
def detect_placas_novas():
    """Detecta placas novas no Manifesto_Frete.xlsx"""
    try:
        import openpyxl
        manifesto_path = Path(__file__).parent / 'uploads' / 'frete' / 'Manifesto_Frete.xlsx'
        
        if not manifesto_path.exists():
            return jsonify({'error': 'Arquivo Manifesto_Frete.xlsx não encontrado'}), 404
        
        # Ler placas do manifesto
        wb = openpyxl.load_workbook(str(manifesto_path))
        ws = wb.active
        
        placas_manifesto = set()
        # Assumindo que a coluna D contém as placas (Veículos)
        for row in range(2, ws.max_row + 1):  # Skip header
            placa = ws.cell(row=row, column=4).value  # Coluna D
            if placa and isinstance(placa, str):
                placa_clean = placa.strip().upper()
                if placa_clean:
                    placas_manifesto.add(placa_clean)
        
        # Verificar quais placas não estão cadastradas
        conn = get_db_connection()
        placas_cadastradas = set()
        for row in conn.execute('SELECT placa FROM veiculos_suporte').fetchall():
            placas_cadastradas.add(row['placa'])
        conn.close()
        
        placas_novas = placas_manifesto - placas_cadastradas
        
        return jsonify({
            'placas_novas': list(placas_novas),
            'total_manifesto': len(placas_manifesto),
            'total_cadastradas': len(placas_cadastradas),
            'total_novas': len(placas_novas)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# =============================================================================
# ROTAS PARA GERENCIAMENTO DE CLIENTES
# =============================================================================

@bp.route('/clientes')
def clientes():
    """Lista todos os clientes cadastrados"""
    conn = get_db_connection()
    clientes = conn.execute('''
        SELECT nome_real, nome_ajustado, data_cadastro, ativo 
        FROM clientes_suporte 
        ORDER BY nome_real
    ''').fetchall()
    conn.close()
    return render_template('suporte_clientes.html', clientes=clientes)

@bp.route('/clientes/add', methods=['POST'])
def add_cliente():
    """Adiciona um novo cliente"""
    nome_real = request.form.get('nome_real', '').strip()
    nome_ajustado = request.form.get('nome_ajustado', '').strip().upper()
    
    if not nome_real or not nome_ajustado:
        flash('Nome real e nome ajustado são obrigatórios', 'error')
        return redirect(url_for('suporte.clientes'))
    
    conn = get_db_connection()
    try:
        conn.execute('''
            INSERT INTO clientes_suporte (nome_real, nome_ajustado, data_cadastro, ativo)
            VALUES (?, ?, ?, ?)
        ''', (nome_real, nome_ajustado, datetime.now(), True))
        conn.commit()
        flash(f'Cliente "{nome_real}" cadastrado com sucesso', 'success')
    except sqlite3.IntegrityError:
        flash(f'Cliente "{nome_real}" já está cadastrado', 'error')
    except Exception as e:
        flash(f'Erro ao cadastrar cliente: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('suporte.clientes'))

@bp.route('/clientes/edit/<nome_real>', methods=['POST'])
def edit_cliente(nome_real):
    """Edita um cliente existente"""
    novo_nome_real = request.form.get('nome_real', '').strip()
    nome_ajustado = request.form.get('nome_ajustado', '').strip().upper()
    ativo = request.form.get('ativo') == 'on'
    
    if not novo_nome_real or not nome_ajustado:
        flash('Nome real e nome ajustado são obrigatórios', 'error')
        return redirect(url_for('suporte.clientes'))
    
    conn = get_db_connection()
    try:
        conn.execute('''
            UPDATE clientes_suporte 
            SET nome_real = ?, nome_ajustado = ?, ativo = ?
            WHERE nome_real = ?
        ''', (novo_nome_real, nome_ajustado, ativo, nome_real))
        conn.commit()
        flash(f'Cliente "{novo_nome_real}" atualizado com sucesso', 'success')
    except Exception as e:
        flash(f'Erro ao atualizar cliente: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('suporte.clientes'))

@bp.route('/clientes/delete/<nome_real>', methods=['POST'])
def delete_cliente(nome_real):
    """Remove um cliente"""
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM clientes_suporte WHERE nome_real = ?', (nome_real,))
        conn.commit()
        flash(f'Cliente "{nome_real}" removido com sucesso', 'success')
    except Exception as e:
        flash(f'Erro ao remover cliente: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('suporte.clientes'))