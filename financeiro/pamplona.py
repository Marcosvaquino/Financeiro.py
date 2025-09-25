"""
Módulo para processamento de arquivos Pamplona
Estrutura similar ao valencio.py
"""

import os
import re
import openpyxl
from datetime import datetime
import pandas as pd
import traceback
import shutil


def tentar_converter_numero_brasileiro(texto):
    """
    Converte texto em número, lidando com formatos brasileiros e americanos.
    Exemplo: '1.234,56' -> 1234.56 | '1,234.56' -> 1234.56 | '1234.56' -> 1234.56
    """
    if not texto or pd.isna(texto):
        return None
    
    if isinstance(texto, (int, float)):
        return float(texto)
    
    texto_str = str(texto).strip()
    if not texto_str or texto_str.lower() in ['', 'nan', 'none']:
        return None
    
    try:
        # Remove espaços e caracteres não numéricos (exceto . , -)
        texto_limpo = re.sub(r'[^\d.,-]', '', texto_str)
        
        # Se tem vírgula como último separador decimal (formato BR)
        if ',' in texto_limpo and '.' in texto_limpo:
            # Formato: 1.234.567,89 -> remover pontos e trocar vírgula por ponto
            texto_limpo = texto_limpo.replace('.', '').replace(',', '.')
        elif ',' in texto_limpo:
            # Se só tem vírgula, assumir que é decimal brasileiro
            texto_limpo = texto_limpo.replace(',', '.')
        
        return float(texto_limpo)
    except (ValueError, TypeError):
        return None


def processar_pamplona(arquivo_upload):
    """
    Processa um arquivo Pamplona (similar ao valencio.py).
    
    Args:
        arquivo_upload: Caminho para o arquivo Excel/CSV
        
    Returns:
        dict: {"success": bool, "message": str, "detalhes": dict}
    """
    try:
        print(f"🏢 INICIANDO PROCESSAMENTO PAMPLONA: {os.path.basename(arquivo_upload)}")

        # Abrir workbook e processar blocos diretamente (para preservar formato)
        wb = openpyxl.load_workbook(arquivo_upload)
        ws = wb.active

        try:
            backup_dir = os.path.join(os.path.dirname(__file__), 'uploads', 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_name = f"backup_pamplona_{timestamp}_{os.path.basename(arquivo_upload)}"
            backup_path = os.path.join(backup_dir, backup_name)
            shutil.copy2(arquivo_upload, backup_path)
            print(f"💾 Backup criado: {backup_name}")
        except Exception as e:
            print(f"⚠️ Não conseguiu criar backup automático: {e}")

        # Processar blocos: acumular PESO (coluna H -> índice 8), detectar TOTAL em coluna C (índice 3)
        blocos = 0
        linhas_processadas = 0
        soma_atual = 0.0
        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row))
        for idx, row in enumerate(rows, start=2):
                cell_nf = row[2].value  # coluna C
                cell_peso = row[7].value  # coluna H

                # Normalizar cell_nf para string
                nf_text = str(cell_nf).strip() if cell_nf is not None else ''

                # Tentar converter peso
                peso_val = None
                if cell_peso is None or (isinstance(cell_peso, str) and str(cell_peso).strip() == ''):
                    peso_val = 0.0
                else:
                    peso_val = tentar_converter_numero_brasileiro(cell_peso)
                    if peso_val is None:
                        # tentar conversão direta se for número em outro formato
                        try:
                            peso_val = float(cell_peso)
                        except Exception:
                            peso_val = 0.0

                # Se não for linha TOTAL, acumular
                if not re.search(r'\bTOTAL\b', nf_text, re.IGNORECASE):
                    soma_atual += float(peso_val or 0.0)
                    linhas_processadas += 1
                else:
                    # Linha TOTAL: gravar soma na coluna H e resultado na coluna I
                    soma_round = round(soma_atual, 2)
                    # Escrever soma em H (coluna 8 -> index 7)
                    ws.cell(row=idx, column=8, value=soma_round)

                    # Calcular resultado = soma * 0.63 / 0.97
                    resultado = 0.0
                    try:
                        resultado = (soma_atual * 0.63) / 0.97
                    except Exception:
                        resultado = 0.0
                    resultado_round = round(resultado, 2)

                    # Escrever resultado em I (coluna 9 -> index 8)
                    ws.cell(row=idx, column=9, value=resultado_round)

                    blocos += 1
                    soma_atual = 0.0

        # Salvar workbook (substituindo o original)
        try:
            wb.save(arquivo_upload)
            print(f"✅ PAMPLONA PROCESSADO: {blocos} blocos, {linhas_processadas} linhas analisadas. Arquivo salvo: {os.path.basename(arquivo_upload)}")
            return {"success": True, "message": f"{blocos} blocos, {linhas_processadas} linhas analisadas", "detalhes": {"arquivo": os.path.basename(arquivo_upload), "backup": os.path.basename(backup_path) if 'backup_path' in locals() else None}}
        except Exception as e:
            traceback.print_exc()
            return {"success": False, "message": f"Erro ao salvar: {e}", "detalhes": {}}
            
    except Exception as e:
        traceback.print_exc()
        return {
            "success": False,
            "message": f"❌ ERRO GERAL: {str(e)}",
            "detalhes": {}
        }


def ler_excel_pamplona(arquivo_upload):
    """
    Lê o arquivo Excel do Pamplona e retorna estrutura de dados.
    """
    try:
        print(f"📖 Lendo arquivo: {os.path.basename(arquivo_upload)}")
        
        # Carregar workbook
        wb = openpyxl.load_workbook(arquivo_upload, data_only=True)
        ws = wb.active
        
        # Extrair cabeçalhos
        colunas = []
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(1, col).value
            colunas.append(str(valor) if valor else f"Col_{col}")
        
        # Extrair dados
        dados = []
        for row in range(2, ws.max_row + 1):
            linha = []
            for col in range(1, ws.max_column + 1):
                valor = ws.cell(row, col).value
                # Converter datetime para string se necessário
                if isinstance(valor, datetime):
                    valor = valor.strftime('%d/%m/%Y')
                # Tentar converter números
                elif valor and isinstance(valor, str):
                    numero = tentar_converter_numero_brasileiro(valor)
                    if numero is not None:
                        valor = numero
                linha.append(valor)
            dados.append(linha)
        
        print(f"✅ Arquivo lido: {len(colunas)} colunas, {len(dados)} linhas de dados")
        
        return {
            "colunas": colunas,
            "dados": dados
        }
        
    except Exception as e:
        print(f"❌ Erro ao ler arquivo: {e}")
        raise


def processar_calculos_pamplona(dados_excel):
    """
    Função legada (mantida apenas para compatibilidade). A nova lógica
    processa diretamente o workbook em `processar_pamplona`.
    """
    raise NotImplementedError("Use processar_pamplona() que processa o workbook diretamente.")


def salvar_pamplona_csv(arquivo_original, dados_processados):
    """
    Salva os dados processados de volta no arquivo Excel original.
    """
    try:
        print(f"💾 Salvando dados processados em: {os.path.basename(arquivo_original)}")
        
        # Carregar workbook original
        wb = openpyxl.load_workbook(arquivo_original, data_only=True)
        ws = wb.active
        
        # Atualizar dados (começando da linha 2, preservando cabeçalho)
        for row_idx, linha in enumerate(dados_processados, start=2):
            for col_idx, valor in enumerate(linha, start=1):
                if col_idx <= ws.max_column:
                    ws.cell(row_idx, col_idx, valor)
        
        # Salvar arquivo
        wb.save(arquivo_original)
        
        print(f"✅ Arquivo salvo com sucesso!")
        
        return {
            "success": True,
            "message": f"Arquivo salvo: {os.path.basename(arquivo_original)}"
        }
        
    except Exception as e:
        print(f"❌ Erro ao salvar: {e}")
        
        # Tentar salvar backup em caso de erro
        try:
            backup_path = arquivo_original + ".backup"
            wb.save(backup_path)
            return {
                "success": False,
                "message": f"Erro ao salvar original, backup criado: {os.path.basename(backup_path)}"
            }
        except:
            return {
                "success": False,
                "message": f"Erro ao salvar: {str(e)}"
            }


# Função auxiliar para compatibilidade com o padrão do sistema
def processar_bloco_pamplona_real(dados_bloco):
    """
    Processa um bloco específico do Pamplona.
    TODO: Implementar lógica específica quando você definir as regras.
    """
    # Por enquanto, retorna os dados sem modificação
    return dados_bloco