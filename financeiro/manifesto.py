"""
Módulo para processamento de arquivos MANIFESTO
Lógica CTRL+C + CTRL+V - preserva formatação exata do Excel
COM INTEGRAÇÃO: enriquece com dados de veículos e clientes
"""
import os
import openpyxl
import csv
import re
from datetime import datetime
from collections import Counter

# Imports para integração
try:
    from .veiculo_helper import VeiculoHelper
    from .cliente_helper import ClienteHelper
    from .custo_frota import CustoFrotaHelper
except ImportError:
    # Fallback para quando executar standalone
    import sys
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    try:
        from financeiro.veiculo_helper import VeiculoHelper
        from financeiro.cliente_helper import ClienteHelper
        from financeiro.custo_frota import CustoFrotaHelper
    except ImportError:
        print("⚠️ Módulos de integração não encontrados. Modo básico ativado.")
        VeiculoHelper = None
        ClienteHelper = None
        CustoFrotaHelper = None

def tentar_converter_numero_brasileiro(texto):
    """
    Tenta converter texto que pode ser um número em formato brasileiro
    Exemplos: "1.234,56", "1234.56", "1,234.56", "1234"
    """
    if not texto or not isinstance(texto, str):
        return None
    
    texto_limpo = texto.strip().replace(' ', '')
    
    # Se está vazio após limpeza
    if not texto_limpo:
        return None
    
    # Verificar se tem apenas números, vírgulas e pontos
    if not re.match(r'^-?\d{1,3}(?:[.,]\d{3})*(?:[.,]\d+)?$', texto_limpo):
        return None
    
    try:
        # Caso 1: Formato brasileiro "1.234,56"
        if ',' in texto_limpo and '.' in texto_limpo:
            # Se vírgula vem depois do ponto, é formato brasileiro
            pos_virgula = texto_limpo.rfind(',')
            pos_ponto = texto_limpo.rfind('.')
            
            if pos_virgula > pos_ponto:
                # Formato brasileiro: 1.234,56
                numero_str = texto_limpo.replace('.', '').replace(',', '.')
                return float(numero_str)
            else:
                # Formato americano: 1,234.56  
                numero_str = texto_limpo.replace(',', '')
                return float(numero_str)
        
        # Caso 2: Apenas vírgula "1234,56" (brasileiro)
        elif ',' in texto_limpo:
            numero_str = texto_limpo.replace(',', '.')
            return float(numero_str)
        
        # Caso 3: Apenas ponto "1234.56" ou "1.234" (pode ser americano ou separador de milhares)
        elif '.' in texto_limpo:
            # Se tem mais de 3 dígitos após o último ponto, é separador decimal
            partes = texto_limpo.split('.')
            if len(partes[-1]) <= 2:  # Última parte tem 1-2 dígitos = decimal
                return float(texto_limpo)
            else:  # Mais de 2 dígitos = separador de milhares brasileiro
                numero_str = texto_limpo.replace('.', '')
                return float(numero_str)
        
        # Caso 4: Apenas números "1234"
        else:
            return float(texto_limpo)
            
    except (ValueError, AttributeError):
        return None

def processar_manifesto(arquivo_upload):
    """
    Processa arquivo de manifesto usando lógica CTRL+C + CTRL+V
    COM INTEGRAÇÃO COMPLETA: enriquece com dados de veículos e clientes
    
    Args:
        arquivo_upload: FileStorage object do Flask ou caminho string
        
    Returns:
        dict: {"success": bool, "message": str, "dados": dict}
    """
    try:
        # Determinar nome original quando arquivo_upload for caminho string
        if isinstance(arquivo_upload, str):
            nome_original = os.path.basename(arquivo_upload)
        else:
            # FileStorage-like
            nome_original = getattr(arquivo_upload, 'filename', 'uploaded_file.xlsx')

        print(f"🚛 PROCESSANDO MANIFESTO: {nome_original}")
        
        # CTRL+C - Ler Excel preservando formatação
        dados_excel = ler_excel_ctrl_c(arquivo_upload)
        
        # INTEGRAÇÃO COMPLETA: Enriquecer com veículos e clientes
        dados_enriquecidos = integrar_manifesto_completo(dados_excel)
        
        # Salvar resultado integrado
        resultado_csv = salvar_manifesto_integrado(dados_enriquecidos, nome_original)
        
        # Calcular custos estimados
        custos = calcular_custos_manifesto(dados_enriquecidos)
        
        # Se recebemos um caminho (arquivo salvo), retornamos esse caminho como arquivo final.
        arquivo_final = nome_original
        if isinstance(arquivo_upload, str):
            arquivo_final = os.path.basename(arquivo_upload)

        return {
            "success": True,
            "message": f"✅ Manifesto integrado processado: {dados_enriquecidos['resumo']}",
            "dados": {
                "linhas_processadas": len(dados_enriquecidos['dados_processados']),
                "arquivo_final": arquivo_final,
                "arquivo_csv_integrado": resultado_csv['arquivo'],
                "colunas": dados_excel['colunas'],
                "integracao": {
                    "veiculos_encontrados": dados_enriquecidos['veiculos_encontrados'],
                    "clientes_encontrados": dados_enriquecidos['clientes_encontrados'],
                    "resumo": dados_enriquecidos['resumo'],
                    "estatisticas": dados_enriquecidos['estatisticas']
                },
                "custos_estimados": {
                    "custo_total": f"R$ {custos['custo_total']:.2f}",
                    "economia_potencial": f"R$ {custos['economia_se_todos_fixo']:.2f}",
                    "breakdown": {
                        "fixo": f"{custos['veiculos_fixo']} veículos = R$ {custos['custo_fixo_total']:.2f}",
                        "spot": f"{custos['veiculos_spot']} veículos = R$ {custos['custo_spot_total']:.2f}",
                        "desconhecidos": f"{custos['veiculos_desconhecidos']} veículos = R$ {custos['custo_desconhecido_total']:.2f}"
                    }
                }
            }
        }
        
    except Exception as e:
        return {
            "success": False,
            "message": f"❌ Erro ao processar manifesto: {str(e)}",
            "dados": {}
        }

def ler_excel_ctrl_c(arquivo_upload):
    """
    CTRL+C - Lê Excel preservando formatação exata
    """
    # Carregar workbook com data_only=True para valores calculados
    wb = openpyxl.load_workbook(arquivo_upload, data_only=True)
    ws = wb.active
    
    dados = []
    colunas = []
    
    # Detectar colunas pela primeira linha
    primeira_linha = True
    for linha in ws.iter_rows(values_only=True):
        if not any(linha):  # Pular linhas vazias
            continue
            
        if primeira_linha:
            colunas = [str(cel) if cel is not None else f"Coluna_{i+1}" for i, cel in enumerate(linha)]
            primeira_linha = False
            continue
            
        # Processar dados preservando formato
        linha_processada = []
        for cel in linha:
            if cel is None:
                linha_processada.append("")
            elif isinstance(cel, datetime):
                # Preservar datas no formato dd/mm/yyyy
                linha_processada.append(cel.strftime("%d/%m/%Y"))
            elif isinstance(cel, (int, float)):
                # NÚMEROS: converter para inteiro (sem casa decimal)
                linha_processada.append(int(cel))
            else:
                # TEXTO: só limpar, não converter
                valor_str = str(cel).replace('\n', ' ').replace('\r', '').strip()
                linha_processada.append(valor_str)
        
        dados.append(linha_processada)
    
    return {
        "colunas": colunas,
        "dados": dados
    }

def salvar_csv_ctrl_v(dados_excel, nome_original):
    """
    CTRL+V - Salva como CSV com encoding UTF-8-BOM
    """
    # Criar diretório se não existir
    pasta_saida = "financeiro/uploads/manifestos"
    os.makedirs(pasta_saida, exist_ok=True)
    # Determinar nome do CSV a partir do nome_original fornecido
    # Se foi passada a versão já renomeada (ex: Manifesto_Frete_09-25_xxx.xlsx)
    # usamos o mesmo base trocando extensão para .csv; caso contrário,
    # usamos nome_original com .csv. Em ambos os casos, evitamos prefixo MANIFESTO_.
    base = os.path.basename(nome_original)
    if base.lower().endswith('.xlsx'):
        nome_csv_base = base[:-5] + '.csv'
    else:
        # garantir extensão .csv
        nome_csv_base = os.path.splitext(base)[0] + '.csv'

    caminho_completo = os.path.join(pasta_saida, nome_csv_base)
    # Se já existir, acrescentar timestamp curinga
    if os.path.exists(caminho_completo):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_csv_base = f"{os.path.splitext(nome_csv_base)[0]}_{timestamp}.csv"
        caminho_completo = os.path.join(pasta_saida, nome_csv_base)
    
    # Salvar CSV com UTF-8-BOM e delimitador ponto-e-vírgula
    with open(caminho_completo, 'w', encoding='utf-8-sig', newline='') as arquivo_csv:
        writer = csv.writer(arquivo_csv, delimiter=';')
        
        # Escrever cabeçalho
        writer.writerow(dados_excel['colunas'])
        
        # Escrever dados
        for linha in dados_excel['dados']:
            writer.writerow(linha)
    
    return {
        "arquivo": nome_csv_base,
        "caminho": caminho_completo,
        "linhas": len(dados_excel['dados'])
    }

def validar_estrutura_manifesto(dados_excel):
    """
    Valida se o arquivo tem a estrutura esperada de manifesto
    """
    colunas_obrigatorias = ['manifesto', 'data', 'veiculo', 'motorista']
    colunas_encontradas = [col.lower() for col in dados_excel['colunas']]
    
    for col_obrig in colunas_obrigatorias:
        if not any(col_obrig in col for col in colunas_encontradas):
            return False, f"Coluna '{col_obrig}' não encontrada"
    
    return True, "Estrutura válida"


def extrair_mes_ano_de_arquivo(file_like):
    """
    Extrai o mês e ano predominante (modo) a partir de uma coluna de data no arquivo Excel.
    Retorna uma string no formato MM-YY (ex: '05-25') ou None se não for possível extrair.
    Aceita file-like (FileStorage) ou caminho para arquivo.
    """
    # Nomes prováveis de coluna de data
    candidatos = [
        'data', 'data_saida', 'data_emissao', 'data_nf', 'dt', 'data_saida', 'data_entrega',
        'data_coleta', 'data_saida_prevista'
    ]

    # Tentar abrir como XLSX primeiro
    try:
        wb = openpyxl.load_workbook(file_like, data_only=True)
        ws = wb.active
    except Exception:
        wb = None
        ws = None

    # Função auxiliar para retornar formato MM-YY a partir de lista de tuplas (mes, ano)
    def _modo_mes_ano(lista):
        if not lista:
            return None
        cnt = Counter(lista)
        (mes, ano), _ = cnt.most_common(1)[0]
        return f"{mes:02d}-{str(ano)[-2:]}"

    candidatos = candidatos if 'candidatos' in locals() else [
        'data', 'data_saida', 'data_emissao', 'data_nf', 'dt', 'data_entrega',
        'data_coleta', 'data_saida_prevista'
    ]

    # Se for XLSX, tentar extrair pela coluna de cabeçalho
    if ws is not None:
        header = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header = [str(c).strip().lower() if c is not None else '' for c in row]
            break

        date_col_idx = None
        if header:
            for i, col_name in enumerate(header):
                for cand in candidatos:
                    if cand in col_name:
                        date_col_idx = i
                        break
                if date_col_idx is not None:
                    break

        # Se não encontrou por nome, tentar detectar coluna que parece ter datas
        if date_col_idx is None:
            max_cols = min(10, ws.max_column)
            max_rows = min(30, ws.max_row)
            scores = [0] * max_cols
            for r in range(2, max_rows + 1):
                for c in range(1, max_cols + 1):
                    val = ws.cell(row=r, column=c).value
                    if isinstance(val, datetime):
                        scores[c - 1] += 1
                    else:
                        try:
                            if isinstance(val, str) and re.search(r"\d{1,2}[-\/\.]\d{1,2}[-\/\.]\d{2,4}", val):
                                scores[c - 1] += 1
                        except Exception:
                            pass
            if scores and max(scores) > 0:
                date_col_idx = scores.index(max(scores))

        # Se ainda não encontrou, buscar padrões embutidos nas primeiras células
        if date_col_idx is None:
            patrones = [r"(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})", r"(\d{4}[\-]\d{1,2}[\-]\d{1,2})"]
            encontrados = []
            max_rows = min(30, ws.max_row)
            max_cols = min(10, ws.max_column)
            for r in range(1, max_rows + 1):
                for c in range(1, max_cols + 1):
                    val = ws.cell(row=r, column=c).value
                    if val is None:
                        continue
                    s = str(val)
                    for pat in patrones:
                        m = re.search(pat, s)
                        if m:
                            date_str = m.group(1)
                            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
                                try:
                                    dt = datetime.strptime(date_str, fmt)
                                    encontrados.append((dt.month, dt.year))
                                    break
                                except Exception:
                                    continue
            if encontrados:
                return _modo_mes_ano(encontrados)

        # Se temos uma coluna de data identificada, coletar os meses
        if date_col_idx is not None:
            meses = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if date_col_idx < len(row):
                    val = row[date_col_idx]
                    if val is None:
                        continue
                    if isinstance(val, datetime):
                        meses.append((val.month, val.year))
                    else:
                        s = str(val).strip()
                        for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d-%b-%Y'):
                            try:
                                dt = datetime.strptime(s, fmt)
                                meses.append((dt.month, dt.year))
                                break
                            except Exception:
                                continue
            if meses:
                return _modo_mes_ano(meses)

    # Se não foi possível extrair via XLSX, tentar CSV/texto (precisa de um caminho)
    path = None
    if isinstance(file_like, str):
        path = file_like
    else:
        # tentar obter caminho do FileStorage (pode ser apenas o filename sem path)
        filename = getattr(file_like, 'filename', None)
        if filename and os.path.isabs(filename) and os.path.exists(filename):
            path = filename

    if not path:
        return None

    try:
        with open(path, 'r', encoding='utf-8-sig', errors='ignore') as f:
            sample = f.read(2048)
            f.seek(0)
            delim = ','
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=';,\t|')
                delim = dialect.delimiter
            except Exception:
                if ';' in sample:
                    delim = ';'

            reader = csv.reader(f, delimiter=delim)
            encontrados = []
            patrones = [r"(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})", r"(\d{4}[\-]\d{1,2}[\-]\d{1,2})"]
            for i, row in enumerate(reader):
                if i >= 30:
                    break
                for cell in row[:10]:
                    if not cell:
                        continue
                    s = str(cell)
                    for pat in patrones:
                        m = re.search(pat, s)
                        if m:
                            date_str = m.group(1)
                            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
                                try:
                                    dt = datetime.strptime(date_str, fmt)
                                    encontrados.append((dt.month, dt.year))
                                    break
                                except Exception:
                                    continue

            if not encontrados:
                return None
            return _modo_mes_ano(encontrados)
    except Exception:
        return None

    # Fim da função - se nada foi encontrado nas tentativas acima, retornamos None
    return None


def integrar_manifesto_completo(dados_excel):
    """
    INTEGRAÇÃO COMPLETA: Enriquece dados do manifesto com veículos e clientes
    
    Args:
        dados_excel: Dict com 'colunas' e 'dados' do Excel
        
    Returns:
        Dict com dados enriquecidos e estatísticas
    """
    if not VeiculoHelper or not ClienteHelper:
        print("⚠️ Integração não disponível - módulos helpers não encontrados")
        return {
            'dados_processados': dados_excel['dados'],
            'veiculos_encontrados': 0,
            'clientes_encontrados': 0,
            'resumo': 'Integração não disponível'
        }
    
    # Converter dados do Excel para formato de dicionários
    colunas = dados_excel['colunas']
    dados_lista = []
    
    for linha in dados_excel['dados']:
        item = {}
        for i, valor in enumerate(linha):
            if i < len(colunas):
                item[colunas[i].lower()] = valor
        dados_lista.append(item)
    
    # Extrair placas e clientes únicos
    placas = []
    clientes = []
    
    for item in dados_lista:
        # Buscar coluna de placa (vários nomes possíveis)
        placa = None
        for col in ['placa', 'veiculo', 'veiculo_placa', 'placa_veiculo']:
            if col in item and item[col]:
                placa = str(item[col]).upper().strip()
                break
        if placa and placa not in placas:
            placas.append(placa)
        
        # Buscar coluna de cliente
        cliente = None  
        for col in ['cliente', 'cliente_nome', 'destinatario', 'remetente']:
            if col in item and item[col]:
                cliente = str(item[col]).upper().strip()
                break
        if cliente and cliente not in clientes:
            clientes.append(cliente)
    
    print(f"🚛 INTEGRAÇÃO MANIFESTO COMPLETO")
    print(f"📊 Dados encontrados: {len(dados_lista)} registros, {len(placas)} placas únicas, {len(clientes)} clientes únicos")
    
    # Buscar dados dos veículos
    dados_veiculos = {}
    veiculos_encontrados = 0
    if placas:
        print(f"🚚 Buscando dados de {len(placas)} veículos...")
        dados_veiculos = VeiculoHelper.buscar_multiplas_placas(placas)
        veiculos_encontrados = sum(1 for v in dados_veiculos.values() if v.get('encontrado', False))
        print(f"✅ Veículos encontrados: {veiculos_encontrados}/{len(placas)}")
    
    # Buscar dados dos clientes usando nova função melhorada
    dados_clientes = {}
    clientes_encontrados = 0
    if clientes:
        print(f"👥 Buscando dados de {len(clientes)} clientes...")
        dados_clientes = ClienteHelper.buscar_multiplos_nomes_manifesto(clientes)
        clientes_encontrados = sum(1 for c in dados_clientes.values() if c.get('encontrado', False))
        print(f"✅ Clientes encontrados: {clientes_encontrados}/{len(clientes)}")
        
        # Debug: mostrar métodos de busca usados
        metodos = {}
        for cliente_dados in dados_clientes.values():
            metodo = cliente_dados.get('metodo', 'unknown')
            metodos[metodo] = metodos.get(metodo, 0) + 1
        print(f"📊 Métodos de busca: {metodos}")
    
    # Enriquecer dados
    dados_enriquecidos = []
    for item in dados_lista:
        item_enriquecido = item.copy()
        
        # Dados do veículo
        placa = None
        for col in ['placa', 'veiculo', 'veiculo_placa', 'placa_veiculo']:
            if col in item and item[col]:
                placa = str(item[col]).upper().strip()
                break
        
        if placa and placa in dados_veiculos:
            veiculo = dados_veiculos[placa]
            item_enriquecido.update({
                'veiculo_status': veiculo.get('status'),
                'veiculo_tipologia': veiculo.get('tipologia'),
                'veiculo_encontrado': veiculo.get('encontrado', False),
                'veiculo_ativo': veiculo.get('ativo', False)
            })
        else:
            item_enriquecido.update({
                'veiculo_status': None,
                'veiculo_tipologia': None,
                'veiculo_encontrado': False,
                'veiculo_ativo': False
            })
        
        # Dados do cliente
        cliente = None
        for col in ['cliente', 'cliente_nome', 'destinatario', 'remetente']:
            if col in item and item[col]:
                cliente = str(item[col]).upper().strip()
                break
        
        if cliente and cliente in dados_clientes:
            cliente_dados = dados_clientes[cliente]
            item_enriquecido.update({
                'cliente_nome_real': cliente_dados.get('nome_real'),
                'cliente_encontrado': cliente_dados.get('encontrado', False)
            })
        else:
            item_enriquecido.update({
                'cliente_nome_real': None,
                'cliente_encontrado': False
            })
        
        # NOVA FUNCIONALIDADE: Custo Frota Fixa (só para veículos FIXOS)
        custo_frota_fixa = 0.0
        if (item_enriquecido.get('veiculo_status') == 'FIXO' and 
            item_enriquecido.get('veiculo_tipologia') and 
            CustoFrotaHelper):
            
            # Buscar KM da viagem
            km_viagem = 0
            for col_km in ['km', 'km_saida', 'km_chegada', 'distancia']:
                if col_km in item and item[col_km]:
                    try:
                        km_viagem = float(item[col_km])
                        break
                    except (ValueError, TypeError):
                        continue
            
            # Calcular custo se tiver tipologia e KM
            if km_viagem > 0:
                custo_frota_fixa = CustoFrotaHelper.calcular_custo_frota_fixa(
                    item_enriquecido.get('veiculo_tipologia'), 
                    km_viagem
                )
        
        item_enriquecido.update({
            'custo_frota_fixa': custo_frota_fixa
        })
        
        dados_enriquecidos.append(item_enriquecido)
    
    # Estatísticas
    veiculos_fixo = sum(1 for item in dados_enriquecidos if item.get('veiculo_status') == 'FIXO')
    veiculos_spot = sum(1 for item in dados_enriquecidos if item.get('veiculo_status') == 'SPOT')
    
    resumo = f"✅ {len(dados_enriquecidos)} registros | 🚚 {veiculos_encontrados} veículos ({veiculos_fixo} FIXO, {veiculos_spot} SPOT) | 👥 {clientes_encontrados} clientes"
    
    print(f"📋 {resumo}")
    
    return {
        'dados_processados': dados_enriquecidos,
        'veiculos_encontrados': veiculos_encontrados,
        'clientes_encontrados': clientes_encontrados,
        'resumo': resumo,
        'estatisticas': {
            'total_registros': len(dados_enriquecidos),
            'veiculos_fixo': veiculos_fixo,
            'veiculos_spot': veiculos_spot,
            'placas_unicas': len(placas),
            'clientes_unicos': len(clientes)
        }
    }


def salvar_manifesto_integrado(dados_enriquecidos, nome_original):
    """
    Salva resultado do manifesto integrado como CSV
    """
    pasta_saida = "financeiro/uploads/manifestos"
    os.makedirs(pasta_saida, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_nome = os.path.splitext(nome_original)[0]
    nome_csv = f"MANIFESTO_INTEGRADO_{timestamp}_{base_nome}.csv"
    caminho_completo = os.path.join(pasta_saida, nome_csv)
    
    # Preparar dados para CSV
    dados_para_csv = dados_enriquecidos['dados_processados']
    
    if not dados_para_csv:
        return {"arquivo": nome_csv, "caminho": caminho_completo, "linhas": 0}
    
    # Obter todas as chaves para o cabeçalho
    todas_chaves = set()
    for item in dados_para_csv:
        todas_chaves.update(item.keys())
    
    colunas_ordenadas = []
    # Primeiro as colunas originais
    for item in dados_para_csv:
        for chave in item.keys():
            if not chave.startswith('veiculo_') and not chave.startswith('cliente_') and chave not in colunas_ordenadas:
                colunas_ordenadas.append(chave)
        break
    
    # Depois as colunas de integração
    colunas_integracao = [k for k in todas_chaves if k.startswith('veiculo_') or k.startswith('cliente_')]
    colunas_ordenadas.extend(sorted(colunas_integracao))
    
    with open(caminho_completo, 'w', encoding='utf-8-sig', newline='') as arquivo_csv:
        writer = csv.writer(arquivo_csv, delimiter=';')
        
        # Cabeçalho de resumo
        writer.writerow(['=== MANIFESTO INTEGRADO ==='])
        writer.writerow(['Resumo:', dados_enriquecidos['resumo']])
        writer.writerow([])
        
        # Cabeçalho das colunas
        writer.writerow(colunas_ordenadas)
        
        # Dados
        for item in dados_para_csv:
            linha = [item.get(col, '') for col in colunas_ordenadas]
            writer.writerow(linha)
    
    return {
        "arquivo": nome_csv,
        "caminho": caminho_completo,
        "linhas": len(dados_para_csv)
    }


def calcular_custos_manifesto(dados_enriquecidos, distancia_media_km=150):
    """
    Calcula custos estimados baseado nos dados enriquecidos
    """
    # Custos por km
    custo_fixo_km = 2.50
    custo_spot_km = 3.20
    custo_desconhecido_km = 3.50
    
    veiculos_fixo = 0
    veiculos_spot = 0
    veiculos_desconhecidos = 0
    
    dados = dados_enriquecidos.get('dados_processados', [])
    
    for item in dados:
        if item.get('veiculo_encontrado'):
            if item.get('veiculo_status') == 'FIXO':
                veiculos_fixo += 1
            elif item.get('veiculo_status') == 'SPOT':
                veiculos_spot += 1
        else:
            veiculos_desconhecidos += 1
    
    custo_fixo_total = veiculos_fixo * custo_fixo_km * distancia_media_km
    custo_spot_total = veiculos_spot * custo_spot_km * distancia_media_km
    custo_desconhecido_total = veiculos_desconhecidos * custo_desconhecido_km * distancia_media_km
    custo_total = custo_fixo_total + custo_spot_total + custo_desconhecido_total
    
    economia_potencial = (veiculos_spot * (custo_spot_km - custo_fixo_km) + 
                         veiculos_desconhecidos * (custo_desconhecido_km - custo_fixo_km)) * distancia_media_km
    
    return {
        'custo_total': custo_total,
        'custo_fixo_total': custo_fixo_total,
        'custo_spot_total': custo_spot_total,
        'custo_desconhecido_total': custo_desconhecido_total,
        'economia_se_todos_fixo': economia_potencial,
        'veiculos_fixo': veiculos_fixo,
        'veiculos_spot': veiculos_spot,
        'veiculos_desconhecidos': veiculos_desconhecidos
    }