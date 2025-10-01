import openpyxl

wb = openpyxl.load_workbook('d:/OneDrive/PROJETOFINANCEIRO.PY/financeiro/uploads/Manifesto_Acumulado_novo.xlsx', data_only=True)
ws = wb.active

print('Verificando cálculos Despesas Gerais:')
print('Linha | Veiculo | Vale_Frete | Custo_Frota | Despesas_Gerais | Status')
print('-' * 80)

count = 0
for i in range(2, 50):  # Verificar primeiras 50 linhas
    vale = float(ws.cell(i, 11).value or 0)
    custo = float(ws.cell(i, 28).value or 0)
    despesas = float(ws.cell(i, 29).value or 0)
    veiculo = str(ws.cell(i, 4).value or '')[:7]
    
    if vale + custo > 0:
        status = "✅" if abs(despesas - (vale + custo)) < 0.01 else "❌"
        print(f'{i:4d} | {veiculo:7s} | {vale:10.2f} | {custo:11.2f} | {despesas:15.2f} | {status}')
        count += 1
        
        if count >= 10:  # Mostrar apenas 10 exemplos
            break

print(f'\n✅ Verificados {count} exemplos de cálculo')