import sys, os
root = r'D:\OneDrive\PROJETOFINANCEIRO.PY'
sys.path.insert(0, root)

import openpyxl

# Verificar estrutura real do arquivo Valencio
arquivo = r'D:\OneDrive\PROJETOFINANCEIRO.PY\financeiro\uploads\valencio\Valencio_Frete_08-25.xlsx'

print("🔍 ANALISANDO ESTRUTURA DO ARQUIVO VALENCIO")
print("=" * 60)

wb = openpyxl.load_workbook(arquivo, data_only=True)
ws = wb.active

# Mostrar cabeçalhos
print("📋 CABEÇALHOS (Linha 1):")
for col in range(1, min(20, ws.max_column + 1)):
    header = ws.cell(1, col).value
    letra = chr(64 + col)  # A=65, B=66, etc
    print(f"  {letra} (col {col}): '{header}'")

# Mostrar algumas linhas de dados
print(f"\n📄 PRIMEIRAS LINHAS DE DADOS:")
for row in range(2, min(12, ws.max_row + 1)):
    col_a = ws.cell(row, 1).value
    col_d = ws.cell(row, 4).value
    col_j = ws.cell(row, 10).value
    col_n = ws.cell(row, 14).value
    
    print(f"  Linha {row}:")
    print(f"    A: '{col_a}'")
    print(f"    D: '{col_d}'")
    print(f"    J: '{col_j}'")
    print(f"    N: '{col_n}'")
    print()

# Procurar por linhas de Manifesto
print("🔍 PROCURANDO MANIFESTOS:")
manifestos_encontrados = 0
for row in range(1, min(50, ws.max_row + 1)):
    col_a = str(ws.cell(row, 1).value or "").strip()
    if "manifesto" in col_a.lower():
        manifestos_encontrados += 1
        print(f"  Linha {row}: '{col_a}'")

print(f"\n📊 TOTAL MANIFESTOS ENCONTRADOS: {manifestos_encontrados}")

# Procurar por FRIGORIFICO VALENCIO
print("\n🏭 PROCURANDO FRIGORIFICO VALENCIO:")
valencio_encontrados = 0
for row in range(1, min(100, ws.max_row + 1)):
    col_d = str(ws.cell(row, 4).value or "").strip().upper()
    if "FRIGORIFICO VALENCIO" in col_d:
        valencio_encontrados += 1
        kg_taxado = ws.cell(row, 10).value
        print(f"  Linha {row}: '{col_d}' | Kg Taxado: {kg_taxado}")

print(f"\n🏭 TOTAL FRIGORIFICO VALENCIO ENCONTRADOS: {valencio_encontrados}")