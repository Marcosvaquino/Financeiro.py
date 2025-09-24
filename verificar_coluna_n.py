import sys, os
root = r'D:\OneDrive\PROJETOFINANCEIRO.PY'
sys.path.insert(0, root)

import openpyxl

# Abrir o arquivo do manifesto
arquivo = os.path.join(root, 'financeiro', 'uploads', 'manifestos', 'Manifesto_Frete_09-25.xlsx')
wb = openpyxl.load_workbook(arquivo, data_only=True)
ws = wb.active

print("IDENTIFICANDO COLUNA N no MANIFESTO:")
print("-" * 50)

# Coluna N = coluna 14
coluna_n = 14
print(f"Coluna N (14): '{ws.cell(1, coluna_n).value}'")

# Ver alguns valores da coluna N
print(f"\nPrimeiros valores da Coluna N:")
for row in range(2, min(12, ws.max_row + 1)):
    valor = ws.cell(row, coluna_n).value
    print(f"  Linha {row}: {valor}")

# Verificar se são números
print(f"\nVerificando se são números:")
valores_numericos = []
for row in range(2, min(50, ws.max_row + 1)):
    valor = ws.cell(row, coluna_n).value
    if valor is not None:
        try:
            num = float(valor)
            valores_numericos.append(num)
        except:
            print(f"  Linha {row}: '{valor}' - NAO É NUMERO")

if valores_numericos:
    print(f"Encontrados {len(valores_numericos)} valores numéricos")
    print(f"Primeiros valores: {valores_numericos[:10]}")
    print(f"Soma dos primeiros 10: {sum(valores_numericos[:10])}")
    
    # Soma de TODA a coluna N
    soma_total = sum(valores_numericos)
    print(f"\nSOMA TOTAL da Coluna N: {soma_total}")
else:
    print("Nenhum valor numérico encontrado na Coluna N")

# Mostrar estrutura das colunas para confirmar
print(f"\nESTRUTURA DAS COLUNAS:")
for col in range(1, 26):  # Colunas A-Z
    header = ws.cell(1, col).value
    letra = chr(64 + col)  # A=65, B=66, etc
    print(f"  {letra} (col {col}): '{header}'")