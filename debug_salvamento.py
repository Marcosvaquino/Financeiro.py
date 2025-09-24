"""
Debug detalhado do salvamento
"""

import openpyxl
from openpyxl import load_workbook
import os
import sys

sys.path.append('financeiro')
from valencio import processar_calculos_valencio, ler_excel_valencio

def debug_salvamento():
    arquivo_path = r"D:\OneDrive\PROJETOFINANCEIRO.PY\financeiro\uploads\valencio\Valencio_Frete_08-25.xlsx"
    
    print(f"🔬 DEBUG DETALHADO DO SALVAMENTO")
    
    # 1. Ler dados
    dados_excel = ler_excel_valencio(arquivo_path)
    print(f"   Linhas originais: {len(dados_excel['dados'])}")
    print(f"   Colunas originais: {len(dados_excel['colunas'])}")
    
    # Verificar algumas linhas antes
    print(f"\n📋 ANTES - Primeiras 5 linhas:")
    for i, linha in enumerate(dados_excel['dados'][:5]):
        print(f"   Linha {i}: {len(linha)} colunas - {linha}")
    
    # 2. Processar
    resultados = processar_calculos_valencio(dados_excel)
    
    # Verificar linhas depois
    print(f"\n📋 DEPOIS - Primeiras 5 linhas:")
    for i, linha in enumerate(resultados['dados_originais'][:5]):
        print(f"   Linha {i}: {len(linha)} colunas - Coluna 14={linha[14] if len(linha) > 14 else 'N/A'}")
    
    # Buscar linhas com valores na coluna 14
    print(f"\n🔍 PROCURANDO VALORES NA COLUNA 14...")
    valores_encontrados = 0
    for i, linha in enumerate(resultados['dados_originais']):
        if len(linha) > 14 and linha[14] and isinstance(linha[14], (int, float)) and linha[14] != '':
            valores_encontrados += 1
            print(f"   Linha {i}: Coluna 14 = {linha[14]} (tipo: {type(linha[14])})")
            if valores_encontrados >= 10:
                print("   ...")
                break
    
    print(f"   Total com valores: {valores_encontrados}")
    
    # 3. Testar salvamento manual
    print(f"\n💾 SALVAMENTO MANUAL...")
    wb = load_workbook(arquivo_path)
    ws = wb.active
    
    # Testar escrever um valor manual na coluna N (14)
    print(f"   Escrevendo valor teste na linha 10, coluna 14...")
    ws.cell(row=10, column=14, value=999.99)
    
    # Salvar
    wb.save(arquivo_path)
    wb.close()
    
    # Verificar se foi salvo
    wb2 = load_workbook(arquivo_path)
    ws2 = wb2.active
    valor_teste = ws2.cell(row=10, column=14).value
    print(f"   Valor lido de volta: {valor_teste}")
    wb2.close()
    
    print(f"\n🎉 DEBUG CONCLUÍDO!")

if __name__ == "__main__":
    debug_salvamento()