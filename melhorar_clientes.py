import sys, os
import unicodedata
import re
root = r'D:\OneDrive\PROJETOFINANCEIRO.PY'
sys.path.insert(0, root)

import openpyxl
from financeiro.veiculo_helper import VeiculoHelper
from financeiro.database import get_connection

def normalizar_nome_busca(nome):
    """Normaliza nome para busca flexível"""
    if not nome:
        return ""
    
    # Converter para string e upper
    nome = str(nome).upper().strip()
    
    # Remover acentos
    nome = unicodedata.normalize('NFD', nome)
    nome = ''.join(c for c in nome if unicodedata.category(c) != 'Mn')
    
    # Normalizar espaços múltiplos
    nome = re.sub(r'\s+', ' ', nome).strip()
    
    return nome

def buscar_cliente_melhorado(nome_procurado, clientes_banco):
    """Busca cliente com algoritmo melhorado"""
    if not nome_procurado:
        return '0'
    
    nome_norm = normalizar_nome_busca(nome_procurado)
    
    # 1. Busca exata primeiro
    for nome_real, nome_ajustado in clientes_banco:
        if normalizar_nome_busca(nome_real) == nome_norm:
            return nome_real
        if normalizar_nome_busca(nome_ajustado) == nome_norm:
            return nome_real
    
    # 2. Busca por contenção
    for nome_real, nome_ajustado in clientes_banco:
        banco_real_norm = normalizar_nome_busca(nome_real)
        banco_ajust_norm = normalizar_nome_busca(nome_ajustado)
        
        if nome_norm in banco_real_norm or banco_real_norm in nome_norm:
            return nome_real
        if nome_norm in banco_ajust_norm or banco_ajust_norm in nome_norm:
            return nome_real
    
    # 3. Busca por palavras-chave
    palavras_procuradas = set(nome_norm.split())
    melhor_match = None
    melhor_score = 0
    
    for nome_real, nome_ajustado in clientes_banco:
        # Testar com nome real
        palavras_real = set(normalizar_nome_busca(nome_real).split())
        if palavras_procuradas and palavras_real:
            intersecao = len(palavras_procuradas.intersection(palavras_real))
            if intersecao >= 2:  # Pelo menos 2 palavras em comum
                score = intersecao / len(palavras_procuradas)
                if score > melhor_score:
                    melhor_score = score
                    melhor_match = nome_real
        
        # Testar com nome ajustado
        palavras_ajust = set(normalizar_nome_busca(nome_ajustado).split())
        if palavras_procuradas and palavras_ajust:
            intersecao = len(palavras_procuradas.intersection(palavras_ajust))
            if intersecao >= 1:  # Pelo menos 1 palavra em comum para nome ajustado
                score = intersecao / len(palavras_procuradas)
                if score > melhor_score:
                    melhor_score = score
                    melhor_match = nome_real
    
    return melhor_match if melhor_match else '0'

# Carregar clientes do banco uma vez
conn = get_connection()
cursor = conn.cursor()
cursor.execute("SELECT nome_real, nome_ajustado FROM clientes_suporte WHERE ativo = 1")
clientes_banco = cursor.fetchall()
conn.close()

print(f"BUSCA MELHORADA DE CLIENTES")
print(f"Clientes no banco: {len(clientes_banco)}")

# Abrir o arquivo do manifesto
arquivo = os.path.join(root, 'financeiro', 'uploads', 'manifestos', 'Manifesto_Frete_09-25.xlsx')
wb = openpyxl.load_workbook(arquivo, data_only=True)
ws = wb.active

# Coletar clientes únicos da coluna S (19) do manifesto
clientes_manifesto = set()
for row in range(2, ws.max_row + 1):
    cliente = ws.cell(row, 19).value  # Coluna S
    if cliente and str(cliente).strip():
        clientes_manifesto.add(str(cliente).strip())

print(f"Clientes únicos no manifesto: {len(clientes_manifesto)}")

# Criar mapeamento
mapeamento = {}
print(f"\nTESTANDO MAPEAMENTO:")
for cliente_manifesto in sorted(clientes_manifesto):
    resultado = buscar_cliente_melhorado(cliente_manifesto, clientes_banco)
    mapeamento[cliente_manifesto.upper().strip()] = resultado
    print(f"'{cliente_manifesto}' -> '{resultado}'")

# Aplicar o mapeamento no arquivo
print(f"\nAPLICANDO MAPEAMENTO NO ARQUIVO...")
linhas_atualizadas = 0

for row in range(2, ws.max_row + 1):
    cliente = ws.cell(row, 19).value  # Coluna S = Classificação
    if cliente and str(cliente).strip():
        cliente_key = str(cliente).upper().strip()
        cliente_real = mapeamento.get(cliente_key, '0')
        ws.cell(row, 25, cliente_real)  # Coluna 25 = Cliente_Real
        linhas_atualizadas += 1
    else:
        ws.cell(row, 25, '0')

# Salvar
wb.save(arquivo)

print(f"\nINTEGRACAO DE CLIENTES FINALIZADA!")
print(f"Linhas atualizadas: {linhas_atualizadas}")
print(f"Mapeamentos aplicados: {len([v for v in mapeamento.values() if v != '0'])}")
print(f"Arquivo salvo: {arquivo}")