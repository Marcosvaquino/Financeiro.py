#!/usr/bin/env python3
"""
Script para criar um arquivo acumulado de manifestos.

Comportamento:
- Procura arquivos em financeiro/uploads/manifestos (xlsx)
- Usa, se existir, um arquivo que contenha "_08" ou "08-" no nome como base de colunas;
  caso contrário, usa o primeiro arquivo encontrado como base.
- Monta lista de colunas: começa com colunas do arquivo base, depois adiciona colunas extras
  encontradas nos outros arquivos (na ordem encontrada).
- Para cada arquivo, lê as linhas (a partir da linha 2) e mapeia valores para as colunas
  do arquivo acumulado (preenchendo com None quando a coluna não existe no arquivo fonte).
- Salva o resultado em financeiro/uploads/Manifesto_Acumulado.xlsx

Uso: execute o script na raiz do projeto (ele assume paths relativos ao workspace).
"""
import os
import glob
import sys
from openpyxl import load_workbook, Workbook
from datetime import datetime

# Adicionar o diretório financeiro ao path para imports
current_dir = os.path.dirname(os.path.abspath(__file__))
financeiro_dir = os.path.join(current_dir, 'financeiro')
if financeiro_dir not in sys.path:
    sys.path.insert(0, financeiro_dir)

# Imports para integração com veículos, clientes e custos
try:
    from financeiro.veiculo_helper import VeiculoHelper
    from financeiro.cliente_helper import ClienteHelper
    from financeiro.custo_frota import CustoFrotaHelper
    INTEGRACAO_DISPONIVEL = True
except ImportError as e:
    print(f"⚠️ Integração não disponível: {e}")
    VeiculoHelper = None
    ClienteHelper = None
    CustoFrotaHelper = None
    INTEGRACAO_DISPONIVEL = False


def integrar_dados_manifesto(ws_out):
    """
    Integra dados de veículos, clientes e custos no manifesto acumulado.
    Preenche as colunas: Status_Veiculo, Tipologia, Cliente_Real, Frete Final, Custo Frota Fixa
    """
    if not INTEGRACAO_DISPONIVEL:
        return

    # Buscar índices das colunas
    headers = [ws_out.cell(1, col).value for col in range(1, ws_out.max_column + 1)]
    
    # Mapear headers para índices (1-based)
    header_map = {}
    for i, header in enumerate(headers, 1):
        if header:
            header_map[header.strip()] = i
    
    # Índices das colunas importantes
    col_veiculo = header_map.get('Veículo', 4)  # Coluna D por padrão
    col_classificacao = header_map.get('Cliente', 31)  # Usar coluna Cliente que tem os dados
    col_km_saida = header_map.get('Km saída', 17)
    col_km_chegada = header_map.get('Km chegada', 18)
    col_status_veiculo = header_map.get('Status_Veiculo')
    col_tipologia = header_map.get('Tipologia')
    col_cliente_real = header_map.get('Cliente_Real')
    col_frete_final = header_map.get('Frete Final')
    col_custo_frota_fixa = header_map.get('Custo Frota Fixa')
    col_despesas_gerais = header_map.get('Despesas Gerais')
    col_vale_frete = header_map.get('Vale frete', 11)
    
    # Se as colunas não existem, não podemos preencher
    if not all([col_status_veiculo, col_tipologia, col_cliente_real]):
        print("⚠️ Colunas de integração não encontradas no manifesto")
        return
    
    # Coletar todas as placas do manifesto
    placas_manifesto = set()
    clientes_manifesto = set()
    
    for row in range(2, ws_out.max_row + 1):
        # Coletar placas
        placa = ws_out.cell(row, col_veiculo).value
        if placa and str(placa).strip():
            placas_manifesto.add(str(placa).upper().strip())
        
        # Coletar clientes
        cliente = ws_out.cell(row, col_classificacao).value
        if cliente and str(cliente).strip():
            clientes_manifesto.add(str(cliente).upper().strip())
    
    # Buscar dados dos veículos e clientes
    print(f"🚚 Buscando dados de {len(placas_manifesto)} veículos...")
    dados_veiculos = VeiculoHelper.buscar_multiplas_placas(list(placas_manifesto))
    
    print(f"👥 Buscando dados de {len(clientes_manifesto)} clientes...")
    dados_clientes = ClienteHelper.buscar_multiplos_nomes_ajustados(list(clientes_manifesto))
    
    # Processar cada linha do manifesto
    linhas_processadas = 0
    for row in range(2, ws_out.max_row + 1):
        # === 1. DADOS DO VEÍCULO ===
        placa = ws_out.cell(row, col_veiculo).value
        if placa and str(placa).strip():
            placa_norm = str(placa).upper().strip()
            veiculo = dados_veiculos.get(placa_norm, {})
            
            # Status_Veiculo
            ws_out.cell(row, col_status_veiculo, veiculo.get('status', '0'))
            
            # Tipologia  
            ws_out.cell(row, col_tipologia, veiculo.get('tipologia', '0'))
        else:
            ws_out.cell(row, col_status_veiculo, '0')
            ws_out.cell(row, col_tipologia, '0')
        
        # === 2. DADOS DO CLIENTE ===
        cliente = ws_out.cell(row, col_classificacao).value
        if cliente and str(cliente).strip():
            cliente_norm = str(cliente).upper().strip()
            cliente_dados = dados_clientes.get(cliente_norm, {})
            ws_out.cell(row, col_cliente_real, cliente_dados.get('nome_real', '0'))
        else:
            ws_out.cell(row, col_cliente_real, '0')
        
        # === 3. CUSTO FROTA FIXA (apenas para veículos FIXOS) ===
        if col_custo_frota_fixa:
            status_veiculo = ws_out.cell(row, col_status_veiculo).value
            tipologia = ws_out.cell(row, col_tipologia).value
            
            if (status_veiculo == 'FIXO' and tipologia and tipologia != '0' and 
                col_km_saida and col_km_chegada):
                
                # Calcular KM da viagem
                km_saida = ws_out.cell(row, col_km_saida).value
                km_chegada = ws_out.cell(row, col_km_chegada).value
                
                try:
                    km_saida_num = float(km_saida) if km_saida else 0
                    km_chegada_num = float(km_chegada) if km_chegada else 0
                    km_viagem = abs(km_chegada_num - km_saida_num)
                    
                    if km_viagem > 0:
                        custo_frota = CustoFrotaHelper.calcular_custo_frota_fixa(tipologia, km_viagem)
                        ws_out.cell(row, col_custo_frota_fixa, custo_frota)
                    else:
                        ws_out.cell(row, col_custo_frota_fixa, 0)
                except (ValueError, TypeError):
                    ws_out.cell(row, col_custo_frota_fixa, 0)
            else:
                # SPOT ou dados inválidos = vazio
                ws_out.cell(row, col_custo_frota_fixa, '')
        
        # === NOVA FUNCIONALIDADE: Calcular Despesas Gerais ===
        # Despesas Gerais = Vale frete (Col 11) + Custo Frota Fixa (Col 28/29)
        if col_despesas_gerais:
            try:
                vale_frete = ws_out.cell(row, col_vale_frete).value or 0
                custo_frota_fixa = ws_out.cell(row, col_custo_frota_fixa).value or 0
                
                # Converter para float se necessário
                vale_frete_num = float(vale_frete) if vale_frete != '' else 0
                custo_frota_fixa_num = float(custo_frota_fixa) if custo_frota_fixa != '' else 0
                
                despesas_total = vale_frete_num + custo_frota_fixa_num
                ws_out.cell(row, col_despesas_gerais, despesas_total)
                
            except (ValueError, TypeError):
                ws_out.cell(row, col_despesas_gerais, 0)
        
        linhas_processadas += 1
        
        # Progress feedback
        if linhas_processadas % 500 == 0:
            print(f"  Integradas {linhas_processadas} linhas...")
    
    print(f"✅ Integração concluída: {linhas_processadas} linhas processadas")


def criar_manifesto_acumulado(upload_dir=None, output_name='Manifesto_Acumulado.xlsx'):
    if upload_dir is None:
        upload_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'financeiro', 'uploads', 'manifestos'))

    if not os.path.isdir(upload_dir):
        return {'success': False, 'message': f'Pasta de manifestos não encontrada: {upload_dir}'}

    arquivos = sorted([p for p in glob.glob(os.path.join(upload_dir, '*.xlsx')) if not os.path.basename(p).startswith('~$')])
    if not arquivos:
        return {'success': False, 'message': f'Nenhum arquivo .xlsx encontrado em {upload_dir}'}

    # Usar cabeçalho fixo (extraído do Manifesto_Acumulado existente)
    # FIXED_HEADERS foi extraído e salvo aqui para não depender de um arquivo base dinâmico
    FIXED_HEADERS = [
        'Manifesto', 'Filial', 'Data', 'Veículo', 'Destino', 'Serviços', 'NFs', 'Kg Real', 'Kg Taxado', 'M3',
        'Vale frete', 'Valor NF', 'Valor Fretes', 'valor final', 'Capacidade Veículo', '% Aprov. Veículo',
        'Km saída', 'Km chegada', 'Km final', 'Valor Frete', 'Classificação', 'Observaçoes operacionais',
        'Status', 'Usuário', 'Status_Veiculo', 'Tipologia', 'Cliente_Real', 'Custo Frota Fixa', 'Despesas Gerais'
    ]

    # Usaremos FIXED_HEADERS como full_headers de saída (mantendo a normalização interna)
    base_headers = list(FIXED_HEADERS)

    # Função para normalizar cabeçalhos (remover acentos, maiúsculas, strip)
    import unicodedata
    def normalizar_texto(texto):
        if not texto:
            return ''
        t = str(texto).upper().strip()
        t = ''.join(c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn')
        # colapsar espaços múltiplos
        t = ' '.join(t.split())
        return t

    # normalizar chaves km para uso durante a construção dos headers
    km_saida_key = normalizar_texto('Km saída')
    km_chegada_key = normalizar_texto('Km chegada')
    km_final_key = normalizar_texto('Km final')

    # full_headers começa com colunas do base, MAS vamos colapsar duplicatas normalizadas
    full_headers = []
    normalized_seen = set()
    duplicatas_base = []
    for idx, h in enumerate(base_headers):
        hn = normalizar_texto(h)
        # Tratar cabeçalhos vazios como únicos (para preservar posição se necessário)
        if not hn:
            unique_key = f'__EMPTY_{idx}'
        else:
            unique_key = hn

        # Ignorar qualquer cabeçalho existente que seja 'Km final' — vamos criar a nossa própria coluna
        if hn == km_final_key:
            # registrar como duplicata removida para informação, mas não incluir no full_headers
            duplicatas_base.append((idx + 1, h))
            continue

        if unique_key not in normalized_seen:
            full_headers.append(h)
            normalized_seen.add(unique_key)
        else:
            # registro de duplicata para log
            duplicatas_base.append((idx + 1, h))

    if duplicatas_base:
        print(f"⚠️ Duplicatas de cabeçalho detectadas no cabeçalho fixo:")
        for pos, nome in duplicatas_base:
            print(f"  - Coluna {pos}: '{nome}' (ignorada na criação do acumulado)")

    # Examinar outros arquivos e coletar colunas extras (apenas se normalized não existir)
    # Examinar outros arquivos e coletar colunas extras (apenas se normalized não existir)
    for a in arquivos:
        wb = load_workbook(a, data_only=True)
        ws = wb.active
        for col in range(1, ws.max_column + 1):
            h = ws.cell(1, col).value
            hname = '' if h is None else str(h)
            hn = normalizar_texto(hname)
            if not hn:
                continue
            # Ignorar colunas 'Km final' vindas de outros arquivos — sempre criaremos somente a nossa coluna
            if hn == km_final_key:
                continue
            if hn not in normalized_seen:
                full_headers.append(hname)
                normalized_seen.add(hn)

    # Criar workbook acumulado
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Manifesto_Acumulado'

    # Precompute normalized keys for full_headers (mantendo alinhamento)
    full_headers_norm = [normalizar_texto(h) for h in full_headers]

    # Inserir UMA UNICA coluna 'Km final' imediatamente após 'Km chegada'
    # Primeiro, remover eventuais ocorrências de 'Km final' (já ignoradas antes, mas garantir)
    # (fazendo uma cópia porque vamos modificar a lista)
    fh = list(full_headers)
    fhn = [normalizar_texto(x) for x in fh]
    # remover todas as ocorrências de km_final_key
    indices_to_remove = [i for i, v in enumerate(fhn) if v == km_final_key]
    for i in reversed(indices_to_remove):
        fh.pop(i)
        fhn.pop(i)

    # agora inserir 'Km final' logo após 'Km chegada' se existir, senão anexar ao final
    if km_chegada_key in fhn:
        insert_pos = fhn.index(km_chegada_key) + 1
        fh.insert(insert_pos, 'Km final')
        fhn.insert(insert_pos, km_final_key)
    else:
        fh.append('Km final')
        fhn.append(km_final_key)

    # substituir full_headers pelas versões ajustadas
    full_headers = fh
    full_headers_norm = fhn

    # Garantir que 'Km saída' esteja antes de 'Km chegada' se ambos existirem.
    # Queremos a ordem: Km saída, Km chegada, Km final
    if km_saida_key in full_headers_norm and km_chegada_key in full_headers_norm:
        idx_saida = full_headers_norm.index(km_saida_key)
        idx_chegada = full_headers_norm.index(km_chegada_key)
        # se Km saída vem depois de Km chegada, mover Km saída para antes
        if idx_saida > idx_chegada:
            h = full_headers.pop(idx_saida)
            hn = full_headers_norm.pop(idx_saida)
            # inserir antes de chegada (posição atual de chegada pode ter mudado se saída removida antes)
            insert_at = full_headers_norm.index(km_chegada_key)
            full_headers.insert(insert_at, h)
            full_headers_norm.insert(insert_at, hn)

    # Escrever cabeçalho final (após todos os ajustes)
    for idx, h in enumerate(full_headers, start=1):
        ws_out.cell(1, idx, h)


    total_rows = 0
    files_processed = 0

    # Selecionar um arquivo por mês/ano: agrupar arquivos por padrão no nome (MM/YY ou MM/YYYY)
    import re
    month_re = re.compile(r'(0[1-9]|1[0-2])[-_]?((?:\d{2})|(?:\d{4}))')

    arquivos_por_mes = {}
    for a in arquivos:
        nome = os.path.basename(a)
        m = month_re.search(nome)
        if not m:
            # arquivo sem mês detectado: colocar em chave None (tratamento posterior)
            key = None
        else:
            mm = m.group(1)
            yy = m.group(2)
            # normalizar ano para 4 dígitos
            if len(yy) == 2:
                ano = '20' + yy
            else:
                ano = yy
            key = f"{ano}-{mm}"
        arquivos_por_mes.setdefault(key, []).append(a)

    # Para cada mês (incluindo None), escolher o arquivo mais recente (mtime)
    arquivos_selecionados = []
    for key, lista in arquivos_por_mes.items():
        # escolher mais recente por last modified
        chosen = max(lista, key=lambda p: os.path.getmtime(p))
        arquivos_selecionados.append(chosen)

    # Agora processar apenas os arquivos selecionados (um por mês ou os sem mês)
    for a in arquivos_selecionados:
        wb = load_workbook(a, data_only=True)
        ws = wb.active

        # Construir mapping normalizado -> lista de colunas fonte (mantendo ordem)
        source_header_cols = {}
        for col in range(1, ws.max_column + 1):
            h = ws.cell(1, col).value
            hn = normalizar_texto(h)
            if hn not in source_header_cols:
                source_header_cols[hn] = []
            source_header_cols[hn].append(col)

        # Construir dest_map: para cada posição de full_headers (index i), escolher a coluna fonte correspondente
        # se existir (ao consumir da lista source_header_cols, preservamos múltiplas ocorrências)
        dest_map = {}
        # copiar listas para que possamos pop sem afetar outros arquivos
        source_cols_copy = {k: list(v) for k, v in source_header_cols.items()}

        for dest_idx, hn in enumerate(full_headers_norm, start=1):
            if hn and hn in source_cols_copy and source_cols_copy[hn]:
                dest_map[dest_idx] = source_cols_copy[hn].pop(0)
            else:
                dest_map[dest_idx] = None

        # Agora ler linhas e preencher de acordo com dest_map
        # Precisamos identificar o índice destino para Km saída, Km chegada e Km final
        dest_idx_saida = None
        dest_idx_chegada = None
        dest_idx_final = None
        for dest_idx, hn in enumerate(full_headers_norm, start=1):
            if hn == km_saida_key:
                dest_idx_saida = dest_idx
            elif hn == km_chegada_key:
                dest_idx_chegada = dest_idx
            elif hn == km_final_key:
                dest_idx_final = dest_idx

        def try_parse_number(v):
            """Tenta converter para float/int, retornando None se não for possível."""
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return v
            s = str(v).strip()
            if s == '':
                return None
            # substituir vírgula por ponto se for decimal com vírgula
            s = s.replace(',', '.')
            try:
                if '.' in s:
                    return float(s)
                return int(s)
            except Exception:
                try:
                    return float(s)
                except Exception:
                    return None

        for row in range(2, ws.max_row + 1):
            row_values = []
            empty = True
            for dest_idx in range(1, len(full_headers) + 1):
                src_col = dest_map.get(dest_idx)
                if src_col:
                    val = ws.cell(row, src_col).value
                else:
                    val = None
                if val is not None:
                    empty = False
                row_values.append(val)

            if not empty:
                # calcular Km final se tivermos colunas de saída/chegada e destino final
                km_final_val = None
                if dest_idx_final is not None and dest_idx_chegada is not None and dest_idx_saida is not None:
                    # índices em row_values são baseados em 0
                    try:
                        chegada_val = try_parse_number(row_values[dest_idx_chegada - 1])
                        saida_val = try_parse_number(row_values[dest_idx_saida - 1])
                        if chegada_val is not None and saida_val is not None:
                            km_final_val = chegada_val - saida_val
                        else:
                            km_final_val = None
                    except Exception:
                        km_final_val = None

                # Sobrescrever o valor na posição de destino de Km final dentro do row_values
                if dest_idx_final is not None:
                    try:
                        row_values[dest_idx_final - 1] = km_final_val
                    except Exception:
                        # caso o índice esteja fora de alcance, ignorar
                        pass

                ws_out.append(row_values)
                total_rows += 1

        files_processed += 1

    # Pós-processamento: garantir que em todas as linhas do arquivo de saída a coluna 'Km final'
    # seja recalculada como (Km chegada - Km saída) e sobrescrita.
    try:
        # localizar índices no arquivo de saída (1-based)
        out_headers = [ws_out.cell(1, c).value for c in range(1, ws_out.max_column + 1)]
        out_headers_norm = [normalizar_texto(h) for h in out_headers]
        try:
            out_idx_saida = out_headers_norm.index(km_saida_key) + 1
        except ValueError:
            out_idx_saida = None
        try:
            out_idx_chegada = out_headers_norm.index(km_chegada_key) + 1
        except ValueError:
            out_idx_chegada = None
        try:
            out_idx_final = out_headers_norm.index(km_final_key) + 1
        except ValueError:
            out_idx_final = None

        def try_parse_number_local(v):
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return v
            s = str(v).strip()
            if s == '':
                return None
            s = s.replace(',', '.')
            try:
                if '.' in s:
                    return float(s)
                return int(s)
            except Exception:
                try:
                    return float(s)
                except Exception:
                    return None

        if out_idx_final and out_idx_chegada and out_idx_saida:
            for r in range(2, ws_out.max_row + 1):
                v_saida = ws_out.cell(r, out_idx_saida).value
                v_chegada = ws_out.cell(r, out_idx_chegada).value
                s_val = try_parse_number_local(v_saida)
                c_val = try_parse_number_local(v_chegada)
                if s_val is not None and c_val is not None:
                    ws_out.cell(r, out_idx_final, c_val - s_val)
                else:
                    # limpar/zerar quando não for possível calcular
                    ws_out.cell(r, out_idx_final, None)
    except Exception:
        # Não falhar o processo caso o pós-processamento não funcione; apenas continuar
        pass

    # --- Nova lógica: criar coluna 'Frete Correto' e preencher a partir dos arquivos Valencio
    try:
        # localizar/garantir coluna 'Frete Correto' no final
        out_headers = [ws_out.cell(1, c).value for c in range(1, ws_out.max_column + 1)]
        out_headers_norm = [normalizar_texto(h) for h in out_headers]
        if normalizar_texto('Frete Correto') in out_headers_norm:
            out_idx_frete = out_headers_norm.index(normalizar_texto('Frete Correto')) + 1
        else:
            out_idx_frete = ws_out.max_column + 1
            ws_out.cell(1, out_idx_frete, 'Frete Correto')

        # Função para parsear número com formatação BR (1.234,56) ou US (1234.56)
        def parse_num_br(v):
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return float(v)
            s = str(v).strip()
            if s == '':
                return None
            # remover espaços e pontos de milhar, trocar vírgula por ponto
            s2 = s.replace(' ', '')
            # detectar se usa vírgula como decimal
            if ',' in s2 and s2.count(',') >= 1:
                # remover pontos (milhares)
                s2 = s2.replace('.', '')
                s2 = s2.replace(',', '.')
            else:
                s2 = s2.replace(',', '')
            try:
                return float(s2)
            except Exception:
                return None

        # Construir mapa manifesto->valor a partir dos arquivos Valencio
        valencio_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'financeiro', 'uploads', 'valencio'))
        # val_map_by_month: { 'YYYY-MM' or None: { manifesto_num: (mtime, value) } }
        val_map_by_month = {}
        if os.path.isdir(valencio_dir):
            val_files = sorted([p for p in glob.glob(os.path.join(valencio_dir, '*.xlsx')) if not os.path.basename(p).startswith('~$')])

            # função helper para extrair mês/ano do NOME do arquivo (aceita nomes/abreviações pt)
            def extrair_periodo_de_nome(nome_arquivo):
                if not nome_arquivo:
                    return None
                nome = os.path.basename(nome_arquivo).lower()
                meses_map = {
                    'jan': 1, 'janeiro': 1,
                    'fev': 2, 'fevereiro': 2,
                    'mar': 3, 'marco': 3, 'março': 3,
                    'abr': 4, 'abril': 4,
                    'mai': 5, 'maio': 5,
                    'jun': 6, 'junho': 6,
                    'jul': 7, 'julho': 7,
                    'ago': 8, 'agosto': 8,
                    'set': 9, 'setembro': 9, 'setem': 9,
                    'out': 10, 'outubro': 10,
                    'nov': 11, 'novembro': 11,
                    'dez': 12, 'dezembro': 12
                }
                import re
                mes_regex = re.compile(r'\b(' + '|'.join(re.escape(k) for k in sorted(meses_map.keys(), key=len, reverse=True)) + r')\b', re.IGNORECASE)
                m = mes_regex.search(nome)
                mes = None
                ano = None
                if m:
                    chave = m.group(1).lower()
                    mes = meses_map.get(chave)

                m4 = re.search(r'20\d{2}', nome)
                if m4:
                    ano = int(m4.group(0))
                else:
                    m2 = re.search(r'[^0-9](\d{2})[^0-9]', f' {nome} ')
                    if m2:
                        ano_candidate = int(m2.group(1))
                        ano = 2000 + ano_candidate if ano_candidate <= 50 else 1900 + ano_candidate

                # fallback: tentar MM-YY numérico
                if mes is None:
                    m_num = re.search(r'\b(0[1-9]|1[0-2])[\-_/]?(\d{2,4})\b', nome)
                    if m_num:
                        try:
                            mes = int(m_num.group(1))
                            yy = m_num.group(2)
                            if len(yy) == 2:
                                ano = 2000 + int(yy)
                            else:
                                ano = int(yy)
                        except Exception:
                            pass

                if mes is None:
                    return None
                if ano is None:
                    ano = datetime.now().year
                return f"{ano}-{mes:02d}"

            # varrer arquivos e coletar linhas que contenham 'Total' e 'Manifesto' na coluna A
            import re
            key_re = re.compile(r'\bTOTAL\b.*\bMANIFESTO\b', flags=re.IGNORECASE)
            num_re = re.compile(r'(\d{3,})')
            for vf in val_files:
                periodo = extrair_periodo_de_nome(vf)
                wbv = None
                try:
                    wbv = load_workbook(vf, data_only=True)
                    wsv = wbv.active
                    for r in range(1, wsv.max_row + 1):
                        a_val = wsv.cell(r, 1).value
                        if not a_val:
                            continue
                        a_text = str(a_val)
                        if not key_re.search(a_text):
                            continue
                        # extrair número do manifesto da célula
                        mnum = num_re.search(a_text)
                        if not mnum:
                            continue
                        manifesto_num = mnum.group(1)
                        # pegar valor na coluna N (14)
                        try:
                            raw_val = wsv.cell(r, 14).value
                        except Exception:
                            raw_val = None
                        parsed = parse_num_br(raw_val)
                        # priorizar arquivo mais recente por mtime dentro do mesmo periodo
                        mtime = os.path.getmtime(vf)
                        month_bucket = periodo
                        if month_bucket not in val_map_by_month:
                            val_map_by_month[month_bucket] = {}
                        prev = val_map_by_month[month_bucket].get(manifesto_num)
                        if prev is None or mtime > prev[0]:
                            val_map_by_month[month_bucket][manifesto_num] = (mtime, parsed if parsed is not None else raw_val)
                except Exception:
                    # ignorar arquivo Valencio com problemas
                    continue

        # Agora preencher cada linha do acumulado
        # identificar índice da coluna 'Data' no arquivo de saída para casar por mês
        out_headers = [ws_out.cell(1, c).value for c in range(1, ws_out.max_column + 1)]
        out_headers_norm = [normalizar_texto(h) for h in out_headers]
        try:
            data_idx = out_headers_norm.index(normalizar_texto('Data')) + 1
        except ValueError:
            data_idx = None

        def parse_ano_mes_de_data(v):
            # espera formatos como 'dd/mm/yyyy' ou 'dd/mm/yy' ou datetime
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return None
            s = str(v).strip()
            if not s:
                return None
            # procurar padrão dd/mm/yyyy ou dd/mm/yy
            import re
            m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
            if m:
                dd = int(m.group(1)); mm = int(m.group(2)); yy = m.group(3)
                if len(yy) == 2:
                    ano = 2000 + int(yy)
                else:
                    ano = int(yy)
                return f"{ano}-{mm:02d}"
            # tentar extrair mês por palavra no campo Data (menos provável)
            m2 = re.search(r'([A-Za-zçãé]+)', s)
            if m2:
                token = m2.group(1)
                # reutilizar lógica de nomes de meses simplificada
                meses_short = {'jan':1,'fev':2,'mar':3,'abr':4,'mai':5,'jun':6,'jul':7,'ago':8,'set':9,'out':10,'nov':11,'dez':12}
                t = token.lower()[:3]
                if t in meses_short:
                    ano = datetime.now().year
                    return f"{ano}-{meses_short[t]:02d}"
            return None

        for r in range(2, ws_out.max_row + 1):
            manifest_val = ws_out.cell(r, 1).value
            if manifest_val is None:
                ws_out.cell(r, out_idx_frete, None)
                continue

            # determinar periodo da linha usando coluna Data
            periodo_linha = None
            if data_idx:
                try:
                    data_val = ws_out.cell(r, data_idx).value
                    periodo_linha = parse_ano_mes_de_data(data_val)
                except Exception:
                    periodo_linha = None

            # normalizar número do manifesto (apenas dígitos)
            mstr = ''.join(ch for ch in str(manifest_val) if ch.isdigit())
            chosen = None

            # tentar lookup no bucket do mesmo período
            if mstr:
                if periodo_linha and periodo_linha in val_map_by_month:
                    chosen = val_map_by_month[periodo_linha].get(mstr)
                    if chosen:
                        chosen = chosen[1]
                # fallback para bucket None (arquivos Valencio sem mês no nome)
                if chosen is None and None in val_map_by_month:
                    prev = val_map_by_month[None].get(mstr)
                    if prev:
                        chosen = prev[1]

            # último fallback: procurar em qualquer bucket e escolher valor com maior mtime
            if chosen is None and mstr:
                best = None
                for bucket, mapping in val_map_by_month.items():
                    if mstr in mapping:
                        cand = mapping[mstr]
                        if best is None or cand[0] > best[0]:
                            best = cand
                if best:
                    chosen = best[1]

            if chosen is None:
                # fallback para valor na coluna N do próprio acumulado (coluna 14)
                try:
                    fallback_raw = ws_out.cell(r, 14).value
                    chosen = parse_num_br(fallback_raw)
                except Exception:
                    chosen = None

            # escrever como número (ou None)
            ws_out.cell(r, out_idx_frete, chosen)
    except Exception:
        # não falhar inteiramente se houver erro aqui
        pass

    # Salvar arquivo
    output_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'financeiro', 'uploads', output_name))
    try:
        # se arquivo já existe, fazer backup antes de sobrescrever
        if os.path.exists(output_path):
            backup_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'financeiro', 'uploads', 'backups'))
            os.makedirs(backup_dir, exist_ok=True)
            import time
            stamp = time.strftime('%Y%m%d_%H%M%S')
            backup_path = os.path.join(backup_dir, f"Manifesto_Acumulado_{stamp}.xlsx")
            try:
                os.replace(output_path, backup_path)
            except Exception:
                # se não conseguir mover, tentar copiar
                try:
                    from shutil import copy2
                    copy2(output_path, backup_path)
                except Exception:
                    # não bloquear a execução apenas por falha no backup
                    pass

        # === NOVA FUNCIONALIDADE: INTEGRAÇÃO COM DADOS DE VEÍCULOS, CLIENTES E CUSTOS ===
        if INTEGRACAO_DISPONIVEL:
            try:
                integrar_dados_manifesto(ws_out)
                print("✅ Integração com dados de veículos, clientes e custos aplicada")
            except Exception as e:
                print(f"⚠️ Erro na integração: {e}")

        wb_out.save(output_path)
    except PermissionError:
        # Arquivo pode estar aberto no Excel; tentar salvar com sufixo
        alt_path = output_path.replace('.xlsx', '_novo.xlsx')
        try:
            wb_out.save(alt_path)
            output_path = alt_path
        except Exception as e:
            return {
                'success': False,
                'message': f'Falha ao salvar arquivo (PermissionError). Tentativa alternativa falhou: {e}',
            }
    except Exception as e:
        return {
            'success': False,
            'message': f'Falha ao salvar arquivo: {e}',
        }

    return {
        'success': True,
        'message': f'Processados {files_processed} arquivos, {total_rows} linhas escritas em {output_path}',
        'output_path': output_path,
        'files': arquivos_selecionados if 'arquivos_selecionados' in locals() else arquivos
    }


if __name__ == '__main__':
    res = criar_manifesto_acumulado()
    if res.get('success'):
        print('OK:', res['message'])
    else:
        print('ERRO:', res['message'])
