import sys, os
root = r'D:\OneDrive\PROJETOFINANCEIRO.PY'
sys.path.insert(0, root)

import openpyxl
from financeiro.veiculo_helper import VeiculoHelper
from financeiro.cliente_helper import ClienteHelper
from financeiro.custo_frota import CustoFrotaHelper

arquivo = os.path.join(root, 'financeiro', 'uploads', 'manifestos', 'Manifesto_Frete_07-25.xlsx')
print('🚛 INTEGRAÇÃO CORRETA - Usando Coluna D (Veículo) e Coluna T (Cliente)')

wb = openpyxl.load_workbook(arquivo, data_only=True)
ws = wb.active

# Coletar dados das colunas de referência
todas_placas = set()  # Coluna D (4)
todos_clientes = set()  # Coluna S (19)

for row in range(2, ws.max_row + 1):
    # Coluna D = Veículo (placas)
    placa = ws.cell(row, 4).value
    if placa and str(placa).strip():
        todas_placas.add(str(placa).upper().strip())
    
    # Coluna S = Classificação (clientes) - Col S = coluna 19
    cliente = ws.cell(row, 19).value  
    if cliente and str(cliente).strip():
        todos_clientes.add(str(cliente).upper().strip())

print(f'📊 Dados coletados:')
print(f'🚚 {len(todas_placas)} placas únicas da Coluna D')
print(f'👥 {len(todos_clientes)} clientes únicos da Coluna S')

if todas_placas:
    print(f'  Primeiras placas: {list(todas_placas)[:5]}')

if todos_clientes:
    print(f'  Primeiros clientes: {list(todos_clientes)[:5]}')

# Buscar dados dos veículos (para Status e Tipologia)
dados_veiculos = {}
if todas_placas:
    print('\n🚚 Buscando dados dos veículos...')
    dados_veiculos = VeiculoHelper.buscar_multiplas_placas(list(todas_placas))
    encontrados_v = sum(1 for v in dados_veiculos.values() if v.get('encontrado', False))
    print(f'✅ {encontrados_v}/{len(todas_placas)} veículos encontrados')

# Buscar dados dos clientes (para Cliente_Real)
dados_clientes = {}
if todos_clientes:
    print('👥 Buscando dados dos clientes...')
    dados_clientes = ClienteHelper.buscar_multiplos_nomes_ajustados(list(todos_clientes))
    encontrados_c = sum(1 for c in dados_clientes.values() if c.get('encontrado', False))
    print(f'✅ {encontrados_c}/{len(todos_clientes)} clientes encontrados')

# Integrar TODAS as linhas
print(f'\n🔄 Integrando {ws.max_row - 1} linhas...')
linhas_processadas = 0

for row in range(2, ws.max_row + 1):
    # COLUNA D (Veículo) -> Status_Veiculo (Col 23) e Tipologia (Col 24)
    placa = ws.cell(row, 4).value
    if placa and str(placa).strip():
        placa_norm = str(placa).upper().strip()
        veiculo = dados_veiculos.get(placa_norm, {})
        ws.cell(row, 23, veiculo.get('status') or '0')  # Status_Veiculo
        ws.cell(row, 24, veiculo.get('tipologia') or '0')  # Tipologia
    else:
        ws.cell(row, 23, '0')
        ws.cell(row, 24, '0')
    
    # COLUNA T (Cliente) -> Cliente_Real (Col 25)
    cliente = ws.cell(row, 20).value  # Coluna T = 20
    if cliente and str(cliente).strip():
        cliente_norm = str(cliente).upper().strip()
        cliente_dados = dados_clientes.get(cliente_norm, {})
        # Se encontrou no ClienteHelper, usa nome_real, senão usa o nome original
        nome_real = cliente_dados.get('nome_real')
        ws.cell(row, 25, nome_real if nome_real else str(cliente).strip())  # Cliente_Real
    else:
        ws.cell(row, 25, '0')
    
    # NOVA COLUNA: Custo Frota Fixa (Col 26) - só para veículos FIXOS
    status_veiculo = ws.cell(row, 23).value  # Status já calculado acima
    if status_veiculo == 'FIXO':
        tipologia = ws.cell(row, 24).value  # Tipologia já calculada acima
        km_saida = ws.cell(row, 6).value  # Coluna F = Km saida
        km_chegada = ws.cell(row, 7).value  # Coluna G = Km chegada
        
        # Calcular KM da viagem
        km_viagem = 0
        try:
            if km_saida and km_chegada:
                km_viagem = abs(float(km_chegada) - float(km_saida))
        except (ValueError, TypeError):
            # Se não conseguir calcular KM, usar 0
            km_viagem = 0
        
        # Calcular custo frota fixa
        if tipologia and tipologia != '0':
            custo_calculado = CustoFrotaHelper.calcular_custo_frota_fixa(tipologia, km_viagem)
            ws.cell(row, 26, custo_calculado)  # Custo Frota Fixa
        else:
            ws.cell(row, 26, 0)
    else:
        # Não é FIXO, deixar vazio
        ws.cell(row, 26, '')
    
    linhas_processadas += 1
    
    if linhas_processadas % 200 == 0:
        print(f'  Processadas {linhas_processadas} linhas...')

# Salvar
print('💾 Salvando arquivo...')
wb.save(arquivo)

print(f'\n🎉 INTEGRAÇÃO FINAL COMPLETA!')
print(f'📁 Arquivo: Manifesto_Frete_09-25.xlsx')
print(f'📊 {linhas_processadas} linhas integradas')
print(f'🚚 Status_Veiculo (Col 23): baseado na Coluna D (Veículo)')
print(f'🔧 Tipologia (Col 24): baseado na Coluna D (Veículo)')
print(f'👥 Cliente_Real (Col 25): baseado na Coluna T (Cliente)')
print(f'💰 Custo Frota Fixa (Col 26): calculado para veículos FIXOS usando Tipologia e KM')
print(f'✅ {len(dados_veiculos)} placas processadas, {len(dados_clientes)} clientes processados')