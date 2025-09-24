"""
Verifica o resultado final do processamento Valencio
"""

import openpyxl
from openpyxl import load_workbook
import os

def verificar_resultado():
    arquivo_path = r"D:\OneDrive\PROJETOFINANCEIRO.PY\financeiro\uploads\valencio\Valencio_Frete_08-25.xlsx"
    
    print(f"🔍 VERIFICANDO RESULTADO FINAL")
    print(f"Arquivo: {arquivo_path}")
    
    # Carrega planilha
    wb = load_workbook(arquivo_path)
    ws = wb.active
    
    print(f"\n📊 ESTRUTURA DO ARQUIVO:")
    print(f"   Total de linhas: {ws.max_row}")
    print(f"   Total de colunas: {ws.max_column}")
    
    # Verifica algumas linhas importantes
    print(f"\n📋 VERIFICANDO ALGUMAS LINHAS:")
    
    # Procura por linhas com "Total - Manifesto"
    total_lines = 0
    for row_num in range(1, min(ws.max_row + 1, 50)):  # Primeiras 50 linhas
        cell_d = ws.cell(row_num, 4)  # Coluna D
        cell_j = ws.cell(row_num, 10)  # Coluna J 
        cell_n = ws.cell(row_num, 14)  # Coluna N
        
        if cell_d.value and "Total - Manifesto" in str(cell_d.value):
            total_lines += 1
            print(f"   Linha {row_num}: {cell_d.value} | J={cell_j.value} | N={cell_n.value}")
        elif cell_d.value and "VALENCIO" in str(cell_d.value):
            print(f"   Linha {row_num}: VALENCIO | J={cell_j.value} | N={cell_n.value}")
        elif cell_d.value and "PAMPLONA" in str(cell_d.value):
            print(f"   Linha {row_num}: PAMPLONA | J={cell_j.value} | N={cell_n.value}")
    
    print(f"\n✅ Linhas 'Total - Manifesto' encontradas: {total_lines}")
    
    # Verifica se há valores na coluna N
    valores_n = 0
    for row_num in range(1, ws.max_row + 1):
        cell_n = ws.cell(row_num, 14)
        if cell_n.value and isinstance(cell_n.value, (int, float)) and cell_n.value > 0:
            valores_n += 1
    
    print(f"✅ Células com valores na coluna N: {valores_n}")
    
    wb.close()
    print(f"\n🎉 VERIFICAÇÃO CONCLUÍDA!")

if __name__ == "__main__":
    verificar_resultado()