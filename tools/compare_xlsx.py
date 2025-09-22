import openpyxl
from pathlib import Path

orig = Path(r'D:\OneDrive\PROJETOFINANCEIRO.PY\financeiro\uploads\frete\manifestos_22-09-2025_01-32 VALENCIO ATE 15-09.xlsx')
proc = Path(r'D:\OneDrive\PROJETOFINANCEIRO.PY\financeiro\uploads\frete\manifestos_22-09-2025_01-32 VALENCIO ATE 15-09_processed.xlsx')

w1 = openpyxl.load_workbook(str(orig), data_only=True)
ws1 = w1.active
w2 = openpyxl.load_workbook(str(proc), data_only=True)
ws2 = w2.active

max_row = max(ws1.max_row, ws2.max_row)
max_col = max(ws1.max_column, ws2.max_column)

differences = []
for r in range(1, max_row+1):
    for c in range(1, max_col+1):
        v1 = ws1.cell(row=r, column=c).value
        v2 = ws2.cell(row=r, column=c).value
        if v1 is None: v1 = ''
        if v2 is None: v2 = ''
        if str(v1).strip() != str(v2).strip():
            differences.append((r, c, v1, v2))
            if len(differences) >= 200:
                break
    if len(differences) >= 200:
        break

print(f'Found {len(differences)} differences (showing up to 200):')
for d in differences[:200]:
    r,c,v1,v2 = d
    col = openpyxl.utils.get_column_letter(c)
    print(f'Row {r} Col {col}: orig="{v1}"  proc="{v2}"')
