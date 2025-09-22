from pathlib import Path
import os
from datetime import datetime
from openpyxl import Workbook
import sys
from pathlib import Path as _P
# ensure workspace root is on sys.path so `financeiro` package is importable
workspace_root = str(_P(__file__).resolve().parents[1])
if workspace_root not in sys.path:
    sys.path.insert(0, workspace_root)

from financeiro.frete import process_frete_file

uploads = Path('financeiro') / 'uploads' / 'frete'
uploads.mkdir(parents=True, exist_ok=True)

orig_name = 'Teste_Valecio.xlsx'
name_lower = orig_name.lower()
if 'valecio' in name_lower or 'valencio' in name_lower:
    base = 'Valencio_Ajustes'
    ext = os.path.splitext(orig_name)[1]
    save_name = f"{base}{ext}"
else:
    save_name = orig_name

save_path = uploads / save_name

# create a minimal workbook expected by process_frete_file
wb = Workbook()
ws = wb.active
# header discovery expects a row equal to 'Tipo' so add that
ws['A1'] = 'Tipo'
ws['B1'] = 'Número'
# add a data row to avoid empty processing
ws['A2'] = 'Frete'
ws['B2'] = '123'
wb.save(str(save_path))

print('Saved test file as:', save_path)

# Call the processor and print output path
out = process_frete_file(str(save_path))
print('Processor output:', out)

# List files in uploads folder
print('\nUploads folder listing:')
for f in sorted(uploads.iterdir()):
    print('-', f.name)