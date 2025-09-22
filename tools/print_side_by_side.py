from openpyxl import load_workbook
orig_path = r'D:\OneDrive\PROJETOFINANCEIRO.PY\financeiro\uploads\frete\manifestos_22-09-2025_01-32 VALENCIO ATE 15-09.xlsx'
proc_path = r'D:\OneDrive\PROJETOFINANCEIRO.PY\financeiro\uploads\frete\manifestos_22-09-2025_01-32 VALENCIO ATE 15-09_processed.xlsx'
wo = load_workbook(orig_path, data_only=True).active
wp = load_workbook(proc_path, data_only=True).active

maxr = max(wo.max_row, wp.max_row, 60)
print('Row | Original (A..P) | Processed (A..P)')
for r in range(1, maxr+1):
    o_vals = [str(wo.cell(row=r, column=c).value or '') for c in range(1,17)]
    p_vals = [str(wp.cell(row=r, column=c).value or '') for c in range(1,17)]
    print(f"{r:02d} | " + ' | '.join(o_vals[:16]) + ' || ' + ' | '.join(p_vals[:16]))
    if r==60:
        break
