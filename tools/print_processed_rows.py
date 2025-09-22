from openpyxl import load_workbook
p_proc = r'D:\OneDrive\PROJETOFINANCEIRO.PY\financeiro\uploads\frete\manifestos_22-09-2025_01-32 VALENCIO ATE 15-09_processed.xlsx'
wp = load_workbook(p_proc, data_only=True).active
for r in range(30,43):
    vals = [str(wp.cell(row=r, column=c).value or '') for c in range(1,16)]
    print(r, vals)
