from openpyxl import load_workbook
p = r'D:\OneDrive\PROJETOFINANCEIRO.PY\financeiro\uploads\frete\manifestos_22-09-2025_01-32 VALENCIO ATE 15-09.xlsx'
wb = load_workbook(p, data_only=True)
ws = wb.active
# find header row
header_row = None
for r in range(1, ws.max_row+1):
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip().lower() == 'tipo':
            header_row = r
            break
    if header_row:
        break
print('header_row =', header_row)
headers = [str(ws.cell(row=header_row, column=c).value or '').strip() for c in range(1, ws.max_column+1)]
print('Headers:')
for i,h in enumerate(headers, start=1):
    print(i, h)

print('\nRows 30..42:')
for r in range(30,43):
    vals = [str(ws.cell(row=r, column=c).value or '') for c in range(1, ws.max_column+1)]
    print(r, vals[:20])
