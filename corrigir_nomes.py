import sys, os
import unicodedata
import re
root = r'D:\OneDrive\PROJETOFINANCEIRO.PY'
sys.path.insert(0, root)

import openpyxl
from financeiro.database import get_connection

def normalizar_nome_busca(nome):
    """Normaliza nome para busca flexível"""
    if not nome:
        return ""
    
    nome = str(nome).upper().strip()
    
    # Remover acentos
    nome = unicodedata.normalize('NFD', nome)
    nome = ''.join(c for c in nome if unicodedata.category(c) != 'Mn')
    
    nome = re.sub(r'\s+', ' ', nome).strip()
    return nome

def buscar_nome_ajustado(nome_real_manifesto, clientes_banco):
    """Busca o NOME AJUSTADO baseado no nome real do manifesto"""
    if not nome_real_manifesto:
        return '0'
    
    nome_norm = normalizar_nome_busca(nome_real_manifesto)
    
    # 1. Busca exata no Nome Real do banco
    for nome_real_banco, nome_ajustado_banco in clientes_banco:
        if normalizar_nome_busca(nome_real_banco) == nome_norm:
            return nome_ajustado_banco  # RETORNA O NOME AJUSTADO!
    
    # 2. Busca por contenção no Nome Real
    for nome_real_banco, nome_ajustado_banco in clientes_banco:
        banco_real_norm = normalizar_nome_busca(nome_real_banco)
        
        if nome_norm in banco_real_norm or banco_real_norm in nome_norm:
            return nome_ajustado_banco  # RETORNA O NOME AJUSTADO!
    
    # 3. Busca por palavras-chave no Nome Real
    palavras_procuradas = set(nome_norm.split())
    melhor_match = None
    melhor_score = 0
    
    for nome_real_banco, nome_ajustado_banco in clientes_banco:
        palavras_real = set(normalizar_nome_busca(nome_real_banco).split())
        
        if palavras_procuradas and palavras_real:
            intersecao = len(palavras_procuradas.intersection(palavras_real))
            if intersecao >= 1:  # Pelo menos 1 palavra em comum
                score = intersecao / len(palavras_procuradas)
                if score > melhor_score:
                    melhor_score = score
                    melhor_match = nome_ajustado_banco  # RETORNA O NOME AJUSTADO!
    
    return melhor_match if melhor_match else '0'

# Carregar clientes do banco
conn = get_connection()
cursor = conn.cursor()
cursor.execute("SELECT nome_real, nome_ajustado FROM clientes_suporte WHERE ativo = 1")
clientes_banco = cursor.fetchall()
conn.close()

print(f"CORRIGINDO BUSCA - NOME REAL -> NOME AJUSTADO")
print(f"Clientes no banco: {len(clientes_banco)}")

# Testar o mapeamento correto
print(f"\nTESTANDO MAPEAMENTO CORRETO:")
nomes_teste = [
    'Adoro Varzea Paulista',  # deve retornar 'ADORO'
    'Marfrig ( BRF )',        # deve retornar 'MARFRIG'
    'MINERVA',                # deve retornar 'MINERVA'
    'GOLD PAO',               # deve retornar 'GOLD PAO'
    'Transferência'           # deve retornar 'TRANSFERENCIA'
]

for nome_manifesto in nomes_teste:
    nome_ajustado = buscar_nome_ajustado(nome_manifesto, clientes_banco)
    print(f"'{nome_manifesto}' -> '{nome_ajustado}'")

# Aplicar no arquivo
arquivo = os.path.join(root, 'financeiro', 'uploads', 'manifestos', 'Manifesto_Frete_09-25.xlsx')
wb = openpyxl.load_workbook(arquivo, data_only=True)
ws = wb.active

print(f"\nAPLICANDO CORRECAO NO ARQUIVO...")
linhas_atualizadas = 0

for row in range(2, ws.max_row + 1):
    nome_real_manifesto = ws.cell(row, 19).value  # Coluna S = Classificação
    
    if nome_real_manifesto and str(nome_real_manifesto).strip():
        nome_ajustado = buscar_nome_ajustado(nome_real_manifesto, clientes_banco)
        ws.cell(row, 25, nome_ajustado)  # Coluna 25 = Cliente_Real
        linhas_atualizadas += 1
    else:
        ws.cell(row, 25, '0')

# Salvar
wb.save(arquivo)

print(f"\nCORRECAO APLICADA COM SUCESSO!")
print(f"Linhas atualizadas: {linhas_atualizadas}")
print(f"Agora a coluna Cliente_Real contem os NOMES AJUSTADOS!")
print(f"Arquivo salvo: {arquivo}")