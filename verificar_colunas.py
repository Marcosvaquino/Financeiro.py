"""
Verificação específica das colunas no arquivo Excel
"""

import openpyxl
from openpyxl import load_workbook
import os

def verificar_colunas_especificas():
    arquivo_path = r"D:\OneDrive\PROJETOFINANCEIRO.PY\financeiro\uploads\valencio\Valencio_Frete_08-25.xlsx"
    
    print(f"🔍 VERIFICANDO COLUNAS ESPECÍFICAS")
    print(f"Arquivo: {arquivo_path}")
    
    # Carrega planilha
    wb = load_workbook(arquivo_path)
    ws = wb.active
    
    print(f"\n📊 ESTRUTURA DO ARQUIVO:")
    print(f"   Total de linhas: {ws.max_row}")
    print(f"   Total de colunas: {ws.max_column}")
    
    # Verificar cabeçalhos
    print(f"\n📋 CABEÇALHOS:")
    for col in range(1, min(ws.max_column + 1, 20)):
        header = ws.cell(1, col).value
        print(f"   Col {col} ({chr(64+col)}): {header}")
    
    # Verificar valores nas últimas colunas
    print(f"\n🔍 VERIFICANDO VALORES NAS COLUNAS N (14) e O (15):")
    valores_n = 0
    valores_o = 0
    
    for row_num in range(1, min(ws.max_row + 1, 50)):
        cell_n = ws.cell(row_num, 14)  # Coluna N
        cell_o = ws.cell(row_num, 15)  # Coluna O
        
        if cell_n.value and isinstance(cell_n.value, (int, float)) and cell_n.value != 0:
            valores_n += 1
            print(f"   Linha {row_num}: N={cell_n.value}")
            
        if cell_o.value and isinstance(cell_o.value, (int, float)) and cell_o.value != 0:
            valores_o += 1
            print(f"   Linha {row_num}: O={cell_o.value}")
            
        if valores_n + valores_o >= 10:
            break
    
    print(f"\n✅ Total valores coluna N (14): {valores_n}")
    print(f"✅ Total valores coluna O (15): {valores_o}")
    
    # Buscar em toda a planilha
    print(f"\n🔍 BUSCA COMPLETA:")
    for row_num in range(1, ws.max_row + 1):
        cell_o = ws.cell(row_num, 15)
        if cell_o.value and isinstance(cell_o.value, (int, float)) and cell_o.value != 0:
            cell_d = ws.cell(row_num, 4).value
            print(f"   Linha {row_num}: {cell_d} -> O={cell_o.value}")
    
    wb.close()
    print(f"\n🎉 VERIFICAÇÃO CONCLUÍDA!")

if __name__ == "__main__":
    verificar_colunas_especificas()