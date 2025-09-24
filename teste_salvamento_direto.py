"""
Teste direto de salvamento do arquivo Valencio
"""

import openpyxl
from openpyxl import load_workbook
import os
import sys

sys.path.append('financeiro')
from valencio import processar_calculos_valencio, ler_excel_valencio, salvar_valencio_csv

def teste_salvamento_direto():
    arquivo_path = r"D:\OneDrive\PROJETOFINANCEIRO.PY\financeiro\uploads\valencio\Valencio_Frete_08-25.xlsx"
    
    print(f"🧪 TESTE DIRETO DE SALVAMENTO")
    print(f"Arquivo: {arquivo_path}")
    
    # 1. Ler dados
    print(f"\n📖 1. Lendo dados...")
    dados_excel = ler_excel_valencio(arquivo_path)
    print(f"   Linhas lidas: {len(dados_excel['dados'])}")
    
    # 2. Processar cálculos 
    print(f"\n🔄 2. Processando cálculos...")
    resultados = processar_calculos_valencio(dados_excel)
    print(f"   Blocos processados: {resultados['blocos_processados']}")
    print(f"   Blocos Valencio: {resultados['blocos_com_valencio']}")
    print(f"   Total: R$ {resultados['total_geral']:.2f}")
    
    # 3. Verificar dados antes de salvar
    print(f"\n🔍 3. Verificando dados modificados...")
    valores_coluna_n = 0
    for linha in resultados['dados_originais']:
        if len(linha) >= 14 and linha[13] and isinstance(linha[13], (int, float)) and linha[13] > 0:
            valores_coluna_n += 1
            print(f"   Linha com valor N: {linha[13]}")
            if valores_coluna_n >= 5:  # Mostrar só os primeiros 5
                print("   ...")
                break
    
    print(f"   Total de linhas com valor N: {valores_coluna_n}")
    
    # 4. Salvar
    print(f"\n💾 4. Salvando arquivo...")
    resultado_salvar = salvar_valencio_csv(resultados, "Valencio_Frete_08-25.xlsx", arquivo_path)
    print(f"   Resultado: {resultado_salvar}")
    
    # 5. Verificar arquivo salvo
    print(f"\n✅ 5. Verificando arquivo salvo...")
    wb = load_workbook(arquivo_path)
    ws = wb.active
    
    valores_salvos = 0
    for row_num in range(1, ws.max_row + 1):
        cell_n = ws.cell(row_num, 14)  # Coluna N
        if cell_n.value and isinstance(cell_n.value, (int, float)) and cell_n.value > 0:
            valores_salvos += 1
            if valores_salvos <= 5:
                print(f"   Linha {row_num} coluna N: {cell_n.value}")
    
    print(f"   Total de valores salvos na coluna N: {valores_salvos}")
    wb.close()
    
    print(f"\n🎉 TESTE CONCLUÍDO!")

if __name__ == "__main__":
    teste_salvamento_direto()